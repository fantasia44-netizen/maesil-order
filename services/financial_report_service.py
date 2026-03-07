"""
financial_report_service.py -- CEO 재무현황 요약 및 세무사 전달용 리포트 생성.

데이터 소스:
  - 매출/영업이익: pnl_service (calculate_monthly_pnl, calculate_channel_pnl)
  - 미수금/미지급금: tax_invoices 테이블 (autotool_accounting 공유 Supabase)
  - 현금 잔액: bank_transactions 테이블 (있으면)
  - 비용: expenses 테이블
"""
import io
import logging
from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


# ──────────────────────────────────────────────
#  유틸리티
# ──────────────────────────────────────────────

def _month_range(year_month):
    """'2026-03' -> ('2026-03-01', '2026-03-31')."""
    parts = year_month.split('-')
    year, month = int(parts[0]), int(parts[1])
    from datetime import timedelta
    if month == 12:
        next_year, next_month = year + 1, 1
    else:
        next_year, next_month = year, month + 1
    last_day = (datetime(next_year, next_month, 1) - timedelta(days=1)).day
    return f"{year}-{month:02d}-01", f"{year}-{month:02d}-{last_day}"


def _prev_month(year_month):
    """'2026-03' -> '2026-02'."""
    parts = year_month.split('-')
    year, month = int(parts[0]), int(parts[1])
    if month == 1:
        return f"{year - 1}-12"
    return f"{year}-{month - 1:02d}"


def _safe_pct(numerator, denominator):
    if not denominator:
        return 0.0
    return round(numerator / denominator * 100, 1)


def _change_pct(current, previous):
    if not previous:
        return None
    return round((current - previous) / abs(previous) * 100, 1)


# ──────────────────────────────────────────────
#  미수금/미지급금 조회
# ──────────────────────────────────────────────

def _query_receivables(db, year_month):
    """미수금 잔액: tax_invoices 중 direction='sales' & 미매칭분 합계.
    tax_invoices 테이블이 없으면 0 반환."""
    try:
        date_from, date_to = _month_range(year_month)
        res = db.client.table("tax_invoices") \
            .select("total_amount,matched_amount,direction,status") \
            .eq("direction", "sales") \
            .gte("issue_date", date_from) \
            .lte("issue_date", date_to) \
            .execute()
        rows = res.data or []
        total = 0
        for r in rows:
            amt = float(r.get('total_amount', 0) or 0)
            matched = float(r.get('matched_amount', 0) or 0)
            total += max(amt - matched, 0)
        return round(total)
    except Exception as e:
        logger.warning(f"[재무리포트] 미수금 조회 실패 (테이블 없을 수 있음): {e}")
        return 0


def _query_payables(db, year_month):
    """미지급금 잔액: tax_invoices 중 direction='purchase' & 미매칭분 합계."""
    try:
        date_from, date_to = _month_range(year_month)
        res = db.client.table("tax_invoices") \
            .select("total_amount,matched_amount,direction,status") \
            .eq("direction", "purchase") \
            .gte("issue_date", date_from) \
            .lte("issue_date", date_to) \
            .execute()
        rows = res.data or []
        total = 0
        for r in rows:
            amt = float(r.get('total_amount', 0) or 0)
            matched = float(r.get('matched_amount', 0) or 0)
            total += max(amt - matched, 0)
        return round(total)
    except Exception as e:
        logger.warning(f"[재무리포트] 미지급금 조회 실패 (테이블 없을 수 있음): {e}")
        return 0


def _query_cash_balance(db):
    """현금 잔액: bank_transactions 최신 레코드의 balance. 테이블 없으면 None."""
    try:
        res = db.client.table("bank_transactions") \
            .select("balance,transaction_date") \
            .order("transaction_date", desc=True) \
            .limit(1) \
            .execute()
        rows = res.data or []
        if rows:
            return float(rows[0].get('balance', 0) or 0)
        return None
    except Exception as e:
        logger.warning(f"[재무리포트] 현금잔액 조회 실패 (테이블 없을 수 있음): {e}")
        return None


# ──────────────────────────────────────────────
#  매입 세금계산서 조회 (거래처별)
# ──────────────────────────────────────────────

def _query_tax_invoices_by_vendor(db, year_month, direction='purchase'):
    """세금계산서 거래처별 합계 조회."""
    try:
        date_from, date_to = _month_range(year_month)
        res = db.client.table("tax_invoices") \
            .select("vendor_name,supply_amount,tax_amount,total_amount,issue_date,matched_amount") \
            .eq("direction", direction) \
            .gte("issue_date", date_from) \
            .lte("issue_date", date_to) \
            .execute()
        rows = res.data or []

        by_vendor = defaultdict(lambda: {
            'supply_amount': 0, 'tax_amount': 0,
            'total_amount': 0, 'count': 0, 'matched_amount': 0,
        })
        for r in rows:
            vendor = r.get('vendor_name', '기타') or '기타'
            by_vendor[vendor]['supply_amount'] += float(r.get('supply_amount', 0) or 0)
            by_vendor[vendor]['tax_amount'] += float(r.get('tax_amount', 0) or 0)
            by_vendor[vendor]['total_amount'] += float(r.get('total_amount', 0) or 0)
            by_vendor[vendor]['matched_amount'] += float(r.get('matched_amount', 0) or 0)
            by_vendor[vendor]['count'] += 1

        return dict(by_vendor)
    except Exception as e:
        logger.warning(f"[재무리포트] 세금계산서({direction}) 조회 실패: {e}")
        return {}


# ══════════════════════════════════════════════
#  CEO 재무현황 요약
# ══════════════════════════════════════════════

def get_ceo_financial_summary(db, year_month):
    """CEO용 재무현황 요약 데이터 생성.

    Args:
        db: SupabaseDB instance
        year_month: '2026-03' 형식

    Returns:
        dict: CEO 대시보드에 필요한 모든 재무 요약 데이터
    """
    from services.pnl_service import (
        calculate_monthly_pnl, calculate_channel_pnl, calculate_pnl_trend,
    )

    # 1. 월별 손익
    pnl = calculate_monthly_pnl(db, year_month)

    revenue_total = pnl['revenue']['total']
    operating_profit = pnl['operating_profit']
    operating_margin = pnl['operating_margin']
    prev_comparison = pnl['prev_month_comparison']

    # 2. 미수금 / 미지급금
    receivables = _query_receivables(db, year_month)
    payables = _query_payables(db, year_month)

    # 전월 미수금/미지급금
    prev_ym = _prev_month(year_month)
    prev_receivables = _query_receivables(db, prev_ym)
    prev_payables = _query_payables(db, prev_ym)

    # 3. 현금 잔액
    cash_balance = _query_cash_balance(db)

    # 4. 채널별 수익성 TOP 5
    channel_pnl = calculate_channel_pnl(db, year_month)
    channels = channel_pnl.get('channels', [])
    # 이미 매출 내림차순 정렬되어 있으므로 profit_margin 기준 재정렬
    top_channels = sorted(channels, key=lambda c: c.get('channel_profit', 0), reverse=True)[:5]

    # 5. 비용 구성 비율
    expense_rows = db.query_expenses(month=year_month)
    expense_by_cat = defaultdict(float)
    expense_total = 0
    for row in expense_rows:
        cat = row.get('category', '기타')
        amt = float(row.get('amount', 0))
        expense_by_cat[cat] += amt
        expense_total += amt

    expense_composition = []
    for cat, amt in sorted(expense_by_cat.items(), key=lambda x: -x[1]):
        expense_composition.append({
            'category': cat,
            'amount': round(amt),
            'ratio': _safe_pct(amt, expense_total),
        })

    # 6. 6개월 추이
    trend = calculate_pnl_trend(db, months=6)

    return {
        'year_month': year_month,
        # KPI 카드
        'revenue': revenue_total,
        'revenue_change': prev_comparison.get('revenue_change'),
        'prev_revenue': prev_comparison.get('prev_revenue', 0),
        'operating_profit': operating_profit,
        'operating_margin': operating_margin,
        'profit_change': prev_comparison.get('profit_change'),
        'prev_operating_profit': prev_comparison.get('prev_operating_profit', 0),
        'gross_profit': pnl['gross_profit'],
        'gross_margin': pnl['gross_margin'],
        # 미수금/미지급금
        'receivables': receivables,
        'receivables_change': _change_pct(receivables, prev_receivables),
        'payables': payables,
        'payables_change': _change_pct(payables, prev_payables),
        # 현금
        'cash_balance': cash_balance,
        # 채널별 수익성 TOP 5
        'top_channels': top_channels,
        # 비용 구성
        'expense_composition': expense_composition,
        'expense_total': round(expense_total),
        # 추이 데이터
        'trend': trend,
        # 상세 P&L (필요 시)
        'cogs': pnl['cogs'],
        'sga': pnl['sga'],
    }


# ══════════════════════════════════════════════
#  세무사 전달용 엑셀 리포트 생성
# ══════════════════════════════════════════════

# 스타일 상수
_HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_SUBTOTAL_FILL = PatternFill(start_color="D6E9F8", end_color="D6E9F8", fill_type="solid")
_SUBTOTAL_FONT = Font(bold=True, size=10)
_THIN_BORDER = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)
_NUM_FMT = '#,##0'
_PCT_FMT = '0.0%'


def _apply_header_style(ws, row, col_count):
    """헤더 행 스타일 적용."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _apply_cell_border(ws, row, col_count):
    """셀 테두리 적용."""
    for col in range(1, col_count + 1):
        ws.cell(row=row, column=col).border = _THIN_BORDER


def _auto_column_width(ws, col_count, min_width=10, max_width=35):
    """컬럼 너비 자동 조정."""
    for col in range(1, col_count + 1):
        max_len = min_width
        letter = get_column_letter(col)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col, max_col=col):
            for cell in row:
                if cell.value is not None:
                    cell_len = len(str(cell.value))
                    # 한글은 2배 폭
                    try:
                        cell_len += sum(1 for c in str(cell.value) if ord(c) > 127)
                    except Exception:
                        pass
                    max_len = max(max_len, cell_len)
        ws.column_dimensions[letter].width = min(max_len + 2, max_width)


def generate_tax_report(db, year_month):
    """세무사 전달용 월간 엑셀 리포트 생성.

    시트 구성:
      1. 매입매출집계표: 매출(채널별), 매입(거래처별)
      2. 거래처별정산현황: 미수금/미지급금 현황
      3. 비용내역서: 카테고리별 비용 상세

    Args:
        db: SupabaseDB instance
        year_month: '2026-03' 형식

    Returns:
        io.BytesIO: 엑셀 파일 바이너리 (send_file 용)
    """
    from services.pnl_service import calculate_monthly_pnl, calculate_channel_pnl

    wb = Workbook()
    date_from, date_to = _month_range(year_month)
    display_month = year_month.replace('-', '년 ') + '월'

    # ── 시트 1: 매입매출 집계표 ──
    ws1 = wb.active
    ws1.title = "매입매출집계표"
    _build_sales_purchase_sheet(ws1, db, year_month, date_from, date_to, display_month)

    # ── 시트 2: 거래처별 정산 현황 ──
    ws2 = wb.create_sheet("거래처별정산현황")
    _build_settlement_sheet(ws2, db, year_month, display_month)

    # ── 시트 3: 비용 내역서 ──
    ws3 = wb.create_sheet("비용내역서")
    _build_expense_sheet(ws3, db, year_month, display_month)

    # 엑셀 파일 → BytesIO
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_sales_purchase_sheet(ws, db, year_month, date_from, date_to, display_month):
    """매입매출 집계표 시트 작성."""
    from services.pnl_service import calculate_channel_pnl

    # 제목
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f"{display_month} 매입매출 집계표"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # ── 매출 섹션 ──
    row = 3
    ws.cell(row=row, column=1, value="[매출]")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="2C3E50")

    row = 4
    headers = ["채널명", "매출액", "수수료", "배송비", "기여이익", "이익률(%)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_style(ws, row, len(headers))

    # 채널별 손익 데이터
    channel_pnl = calculate_channel_pnl(db, year_month)
    channels = channel_pnl.get('channels', [])
    total_ch = channel_pnl.get('total', {})

    row = 5
    for ch in channels:
        ws.cell(row=row, column=1, value=ch['channel'])
        ws.cell(row=row, column=2, value=ch['revenue']).number_format = _NUM_FMT
        ws.cell(row=row, column=3, value=ch['commission']).number_format = _NUM_FMT
        ws.cell(row=row, column=4, value=ch['shipping']).number_format = _NUM_FMT
        ws.cell(row=row, column=5, value=ch['channel_profit']).number_format = _NUM_FMT
        ws.cell(row=row, column=6, value=ch['profit_margin'] / 100 if ch['profit_margin'] else 0).number_format = _PCT_FMT
        _apply_cell_border(ws, row, len(headers))
        row += 1

    # 매출 소계
    ws.cell(row=row, column=1, value="매출 합계")
    ws.cell(row=row, column=2, value=total_ch.get('revenue', 0)).number_format = _NUM_FMT
    ws.cell(row=row, column=3, value=total_ch.get('commission', 0)).number_format = _NUM_FMT
    ws.cell(row=row, column=4, value=total_ch.get('shipping', 0)).number_format = _NUM_FMT
    ws.cell(row=row, column=5, value=total_ch.get('channel_profit', 0)).number_format = _NUM_FMT
    margin = total_ch.get('profit_margin', 0)
    ws.cell(row=row, column=6, value=margin / 100 if margin else 0).number_format = _PCT_FMT
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _SUBTOTAL_FILL
        cell.font = _SUBTOTAL_FONT
        cell.border = _THIN_BORDER
    row += 2

    # ── 매입 섹션 (세금계산서 기반) ──
    ws.cell(row=row, column=1, value="[매입]")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="2C3E50")
    row += 1

    purchase_headers = ["거래처명", "공급가액", "부가세", "합계금액", "건수"]
    for i, h in enumerate(purchase_headers, 1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_style(ws, row, len(purchase_headers))
    row += 1

    purchase_by_vendor = _query_tax_invoices_by_vendor(db, year_month, direction='purchase')
    purchase_total = {'supply': 0, 'tax': 0, 'total': 0, 'count': 0}

    if purchase_by_vendor:
        for vendor, data in sorted(purchase_by_vendor.items(), key=lambda x: -x[1]['total_amount']):
            ws.cell(row=row, column=1, value=vendor)
            ws.cell(row=row, column=2, value=round(data['supply_amount'])).number_format = _NUM_FMT
            ws.cell(row=row, column=3, value=round(data['tax_amount'])).number_format = _NUM_FMT
            ws.cell(row=row, column=4, value=round(data['total_amount'])).number_format = _NUM_FMT
            ws.cell(row=row, column=5, value=data['count'])
            _apply_cell_border(ws, row, len(purchase_headers))
            purchase_total['supply'] += data['supply_amount']
            purchase_total['tax'] += data['tax_amount']
            purchase_total['total'] += data['total_amount']
            purchase_total['count'] += data['count']
            row += 1
    else:
        ws.cell(row=row, column=1, value="(세금계산서 데이터 없음)")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1

    # 매입 소계
    ws.cell(row=row, column=1, value="매입 합계")
    ws.cell(row=row, column=2, value=round(purchase_total['supply'])).number_format = _NUM_FMT
    ws.cell(row=row, column=3, value=round(purchase_total['tax'])).number_format = _NUM_FMT
    ws.cell(row=row, column=4, value=round(purchase_total['total'])).number_format = _NUM_FMT
    ws.cell(row=row, column=5, value=purchase_total['count'])
    for col in range(1, len(purchase_headers) + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _SUBTOTAL_FILL
        cell.font = _SUBTOTAL_FONT
        cell.border = _THIN_BORDER

    _auto_column_width(ws, 6)


def _build_settlement_sheet(ws, db, year_month, display_month):
    """거래처별 정산 현황 시트 작성."""
    # 제목
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f"{display_month} 거래처별 정산 현황"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # ── 미수금 (매출 세금계산서) ──
    row = 3
    ws.cell(row=row, column=1, value="[미수금 (매출)]")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="E74C3C")
    row += 1

    headers = ["거래처명", "세금계산서 합계", "매칭(입금) 금액", "미수 잔액", "건수"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_style(ws, row, len(headers))
    row += 1

    sales_by_vendor = _query_tax_invoices_by_vendor(db, year_month, direction='sales')
    ar_total = {'total': 0, 'matched': 0, 'outstanding': 0, 'count': 0}

    if sales_by_vendor:
        for vendor, data in sorted(sales_by_vendor.items(), key=lambda x: -(x[1]['total_amount'] - x[1]['matched_amount'])):
            outstanding = data['total_amount'] - data['matched_amount']
            ws.cell(row=row, column=1, value=vendor)
            ws.cell(row=row, column=2, value=round(data['total_amount'])).number_format = _NUM_FMT
            ws.cell(row=row, column=3, value=round(data['matched_amount'])).number_format = _NUM_FMT
            ws.cell(row=row, column=4, value=round(outstanding)).number_format = _NUM_FMT
            ws.cell(row=row, column=5, value=data['count'])
            _apply_cell_border(ws, row, len(headers))
            ar_total['total'] += data['total_amount']
            ar_total['matched'] += data['matched_amount']
            ar_total['outstanding'] += outstanding
            ar_total['count'] += data['count']
            row += 1
    else:
        ws.cell(row=row, column=1, value="(매출 세금계산서 데이터 없음)")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1

    # 미수금 소계
    ws.cell(row=row, column=1, value="미수금 합계")
    ws.cell(row=row, column=2, value=round(ar_total['total'])).number_format = _NUM_FMT
    ws.cell(row=row, column=3, value=round(ar_total['matched'])).number_format = _NUM_FMT
    ws.cell(row=row, column=4, value=round(ar_total['outstanding'])).number_format = _NUM_FMT
    ws.cell(row=row, column=5, value=ar_total['count'])
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _SUBTOTAL_FILL
        cell.font = _SUBTOTAL_FONT
        cell.border = _THIN_BORDER
    row += 2

    # ── 미지급금 (매입 세금계산서) ──
    ws.cell(row=row, column=1, value="[미지급금 (매입)]")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="3498DB")
    row += 1

    for i, h in enumerate(headers, 1):
        h2 = h.replace("미수", "미지급").replace("매칭(입금)", "매칭(지급)")
        ws.cell(row=row, column=i, value=h2)
    _apply_header_style(ws, row, len(headers))
    row += 1

    purchase_by_vendor = _query_tax_invoices_by_vendor(db, year_month, direction='purchase')
    ap_total = {'total': 0, 'matched': 0, 'outstanding': 0, 'count': 0}

    if purchase_by_vendor:
        for vendor, data in sorted(purchase_by_vendor.items(), key=lambda x: -(x[1]['total_amount'] - x[1]['matched_amount'])):
            outstanding = data['total_amount'] - data['matched_amount']
            ws.cell(row=row, column=1, value=vendor)
            ws.cell(row=row, column=2, value=round(data['total_amount'])).number_format = _NUM_FMT
            ws.cell(row=row, column=3, value=round(data['matched_amount'])).number_format = _NUM_FMT
            ws.cell(row=row, column=4, value=round(outstanding)).number_format = _NUM_FMT
            ws.cell(row=row, column=5, value=data['count'])
            _apply_cell_border(ws, row, len(headers))
            ap_total['total'] += data['total_amount']
            ap_total['matched'] += data['matched_amount']
            ap_total['outstanding'] += outstanding
            ap_total['count'] += data['count']
            row += 1
    else:
        ws.cell(row=row, column=1, value="(매입 세금계산서 데이터 없음)")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1

    # 미지급금 소계
    ws.cell(row=row, column=1, value="미지급금 합계")
    ws.cell(row=row, column=2, value=round(ap_total['total'])).number_format = _NUM_FMT
    ws.cell(row=row, column=3, value=round(ap_total['matched'])).number_format = _NUM_FMT
    ws.cell(row=row, column=4, value=round(ap_total['outstanding'])).number_format = _NUM_FMT
    ws.cell(row=row, column=5, value=ap_total['count'])
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _SUBTOTAL_FILL
        cell.font = _SUBTOTAL_FONT
        cell.border = _THIN_BORDER

    _auto_column_width(ws, len(headers))


def _build_expense_sheet(ws, db, year_month, display_month):
    """비용 내역서 시트 작성."""
    # 제목
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f"{display_month} 비용 내역서"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    row = 3
    headers = ["날짜", "카테고리", "세부항목", "금액", "비고", "등록자"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_style(ws, row, len(headers))

    # 비용 데이터
    expense_rows = db.query_expenses(month=year_month)
    row = 4
    cat_totals = defaultdict(float)
    grand_total = 0

    for exp in expense_rows:
        cat = exp.get('category', '기타')
        amt = float(exp.get('amount', 0))
        cat_totals[cat] += amt
        grand_total += amt

        ws.cell(row=row, column=1, value=exp.get('expense_date', ''))
        ws.cell(row=row, column=2, value=cat)
        ws.cell(row=row, column=3, value=exp.get('subcategory', ''))
        ws.cell(row=row, column=4, value=round(amt)).number_format = _NUM_FMT
        ws.cell(row=row, column=5, value=exp.get('memo', ''))
        ws.cell(row=row, column=6, value=exp.get('registered_by', ''))
        _apply_cell_border(ws, row, len(headers))
        row += 1

    if not expense_rows:
        ws.cell(row=row, column=1, value="(비용 데이터 없음)")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1

    # 카테고리별 소계
    row += 1
    ws.cell(row=row, column=1, value="[카테고리별 합계]")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2C3E50")
    row += 1

    summary_headers = ["카테고리", "금액", "비율(%)"]
    for i, h in enumerate(summary_headers, 1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_style(ws, row, len(summary_headers))
    row += 1

    for cat, amt in sorted(cat_totals.items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=round(amt)).number_format = _NUM_FMT
        ratio = amt / grand_total if grand_total else 0
        ws.cell(row=row, column=3, value=ratio).number_format = _PCT_FMT
        _apply_cell_border(ws, row, len(summary_headers))
        row += 1

    # 총합계
    ws.cell(row=row, column=1, value="총 합계")
    ws.cell(row=row, column=2, value=round(grand_total)).number_format = _NUM_FMT
    ws.cell(row=row, column=3, value=1.0).number_format = _PCT_FMT
    for col in range(1, len(summary_headers) + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _SUBTOTAL_FILL
        cell.font = _SUBTOTAL_FONT
        cell.border = _THIN_BORDER

    _auto_column_width(ws, len(headers))
