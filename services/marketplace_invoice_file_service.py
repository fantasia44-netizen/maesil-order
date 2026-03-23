"""
marketplace_invoice_file_service.py — 마켓별 송장등록 파일 생성 서비스.

각 마켓 관리자에서 수동 업로드 가능한 형식의 엑셀 파일 생성.
"""
import logging
import os
import tempfile

import openpyxl

from services.channel_config import PLATFORM_MAP

logger = logging.getLogger(__name__)


def generate_marketplace_invoice_file(db, channel: str) -> str:
    """마켓별 송장등록 엑셀 파일 생성.

    Args:
        db: SupabaseDB 인스턴스
        channel: 채널명

    Returns:
        str: 생성된 파일 경로

    Raises:
        ValueError: 데이터 없음 등
    """
    pending = db.query_pending_invoice_push(channel=channel)
    # API 매핑 있는 건만 (수동 주문 제외)
    mapped = [p for p in pending if p.get('api_order_id')]

    if not mapped:
        raise ValueError(f'{channel}: 송장등록 대기 건이 없습니다.')

    platform = PLATFORM_MAP.get(channel, '')

    if platform == 'naver':
        return _generate_naver_file(mapped, channel)
    elif platform == 'coupang':
        return _generate_coupang_file(mapped, channel)
    elif platform == 'cafe24':
        return _generate_cafe24_file(mapped, channel)
    else:
        raise ValueError(f'{channel}: 지원하지 않는 플랫폼 ({platform})')


def _generate_naver_file(orders: list, channel: str) -> str:
    """네이버 스마트스토어 발주확인 일괄발송 양식.

    컬럼: 상품주문번호, 택배사, 송장번호
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '일괄발송'

    ws.append(['상품주문번호', '택배사', '송장번호'])

    for o in orders:
        ws.append([
            o.get('api_line_id') or o.get('api_order_id', ''),
            'CJ대한통운',
            o.get('invoice_no', ''),
        ])

    return _save_workbook(wb, f'네이버_송장등록_{channel}')


def _generate_coupang_file(orders: list, channel: str) -> str:
    """쿠팡 송장업로드 양식.

    컬럼: 주문번호, 묶음배송번호(shipmentBoxId), 택배사코드, 송장번호
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '송장업로드'

    ws.append(['주문번호', '묶음배송번호', '택배사코드', '송장번호'])

    for o in orders:
        raw = o.get('raw_data') or {}
        shipment_box_id = ''
        # raw_data에서 shipmentBoxId 추출
        if isinstance(raw, dict):
            shipment_box_id = str(raw.get('shipmentBoxId', ''))

        ws.append([
            o.get('api_order_id', ''),
            shipment_box_id,
            'CJGLS',
            o.get('invoice_no', ''),
        ])

    return _save_workbook(wb, f'쿠팡_송장등록_{channel}')


def _generate_cafe24_file(orders: list, channel: str) -> str:
    """Cafe24 배송등록 양식.

    컬럼: 주문번호, 배송업체코드, 송장번호
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '배송등록'

    ws.append(['주문번호', '배송업체코드', '송장번호'])

    for o in orders:
        ws.append([
            o.get('api_order_id', ''),
            'cj',
            o.get('invoice_no', ''),
        ])

    return _save_workbook(wb, f'카페24_송장등록_{channel}')


def _save_workbook(wb, prefix: str) -> str:
    """Workbook을 임시 파일로 저장."""
    from services.tz_utils import now_kst
    ts = now_kst().strftime('%Y%m%d_%H%M%S')
    fd, path = tempfile.mkstemp(suffix='.xlsx', prefix=f'{prefix}_{ts}_')
    os.close(fd)
    wb.save(path)
    wb.close()
    return path
