"""
report_service.py -- 세무사 전달용 리포트 생성 서비스.

- 월간 요약 (매출/매입/미수금/미지급금/입출금)
- 세금계산서 엑셀 내보내기 (국세청 양식)
- 은행 거래내역 월간 요약 엑셀
"""
import io
import logging
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


# ── 공통 스타일 ──
_HEADER_FONT = Font(name='맑은 고딕', bold=True, size=11, color='FFFFFF')
_HEADER_FILL = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
_HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
_CELL_FONT = Font(name='맑은 고딕', size=10)
_MONEY_FORMAT = '#,##0'
_THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def _apply_header_style(ws, row_num, col_count):
    """헤더 행에 스타일 적용."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _apply_cell_style(cell, is_money=False):
    """일반 셀 스타일."""
    cell.font = _CELL_FONT
    cell.border = _THIN_BORDER
    if is_money:
        cell.number_format = _MONEY_FORMAT
        cell.alignment = Alignment(horizontal='right')


def _month_range(year_month):
    """'2026-03' → ('2026-03-01', '2026-03-31')."""
    y, m = year_month.split('-')
    m = int(m)
    last_day = calendar.monthrange(int(y), m)[1]
    return f'{y}-{int(m):02d}-01', f'{y}-{int(m):02d}-{last_day:02d}'


# ═══════════════════════════════════════════════════════════
#  1) 월간 요약 (JSON)
# ═══════════════════════════════════════════════════════════

def generate_monthly_summary(db, year_month):
    """월간 회계 요약 생성.

    Args:
        db: SupabaseDB 인스턴스
        year_month: 'YYYY-MM' 형식

    Returns:
        dict: {
            year_month, sales_invoices, purchase_invoices,
            receivable_balance, payable_balance, bank_summary
        }
    """
    date_from, date_to = _month_range(year_month)

    # ── 매출 세금계산서 합계 ──
    sales_invoices = db.query_tax_invoices(
        direction='sales', date_from=date_from, date_to=date_to)
    sales_supply = sum(i.get('supply_cost_total', 0) for i in sales_invoices)
    sales_tax = sum(i.get('tax_total', 0) for i in sales_invoices)
    sales_total = sum(i.get('total_amount', 0) for i in sales_invoices)

    # ── 매입 세금계산서 합계 ──
    purchase_invoices = db.query_tax_invoices(
        direction='purchase', date_from=date_from, date_to=date_to)
    purchase_supply = sum(i.get('supply_cost_total', 0) for i in purchase_invoices)
    purchase_tax = sum(i.get('tax_total', 0) for i in purchase_invoices)
    purchase_total = sum(i.get('total_amount', 0) for i in purchase_invoices)

    # ── 미수금 잔액 (매출 세금계산서 미매칭 합계) ──
    unmatched_sales = db.query_tax_invoices(
        direction='sales', unmatched_only=True)
    receivable_balance = sum(i.get('total_amount', 0) for i in unmatched_sales)

    # ── 미지급금 잔액 (매입 세금계산서 미매칭 합계) ──
    unmatched_purchase = db.query_tax_invoices(
        direction='purchase', unmatched_only=True)
    payable_balance = sum(i.get('total_amount', 0) for i in unmatched_purchase)

    # ── 은행 입출금 요약 ──
    from services.bank_service import get_transaction_summary
    bank_summary = get_transaction_summary(db, date_from=date_from, date_to=date_to)

    return {
        'year_month': year_month,
        'sales_invoices': {
            'count': len(sales_invoices),
            'supply_amount': sales_supply,
            'tax_amount': sales_tax,
            'total_amount': sales_total,
        },
        'purchase_invoices': {
            'count': len(purchase_invoices),
            'supply_amount': purchase_supply,
            'tax_amount': purchase_tax,
            'total_amount': purchase_total,
        },
        'receivable_balance': receivable_balance,
        'payable_balance': payable_balance,
        'bank_summary': {
            'total_in': bank_summary.get('total_in', 0),
            'total_out': bank_summary.get('total_out', 0),
            'net': bank_summary.get('net', 0),
            'transaction_count': bank_summary.get('count', 0),
            'by_category': bank_summary.get('by_category', {}),
        },
    }


# ═══════════════════════════════════════════════════════════
#  2) 세금계산서 엑셀 내보내기 (국세청 양식)
# ═══════════════════════════════════════════════════════════

def export_tax_invoices_excel(db, year_month, direction='sales'):
    """세금계산서 목록을 엑셀로 내보내기.

    Args:
        db: SupabaseDB 인스턴스
        year_month: 'YYYY-MM'
        direction: 'sales' (매출) 또는 'purchase' (매입)

    Returns:
        io.BytesIO: 엑셀 파일 바이트스트림
    """
    date_from, date_to = _month_range(year_month)
    invoices = db.query_tax_invoices(
        direction=direction, date_from=date_from, date_to=date_to)

    wb = Workbook()
    ws = wb.active

    direction_label = '매출' if direction == 'sales' else '매입'
    ws.title = f'{direction_label} 세금계산서'

    # ── 제목 ──
    ws.merge_cells('A1:L1')
    title_cell = ws['A1']
    title_cell.value = f'{year_month} {direction_label} 세금계산서 목록'
    title_cell.font = Font(name='맑은 고딕', bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # ── 헤더 (국세청 양식 참고) ──
    headers = [
        'No.', '작성일', '발행일', '종류',
        '공급자 사업자번호', '공급자 상호',
        '공급받는자 사업자번호', '공급받는자 상호',
        '공급가액', '세액', '합계금액',
        '비고',
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    _apply_header_style(ws, 3, len(headers))

    # ── 데이터 ──
    for idx, inv in enumerate(invoices, 1):
        row = idx + 3
        values = [
            idx,
            str(inv.get('write_date', '')),
            str(inv.get('issue_date', '') or inv.get('write_date', '')),
            inv.get('invoice_type', ''),
            inv.get('supplier_corp_num', ''),
            inv.get('supplier_corp_name', ''),
            inv.get('buyer_corp_num', ''),
            inv.get('buyer_corp_name', ''),
            inv.get('supply_cost_total', 0),
            inv.get('tax_total', 0),
            inv.get('total_amount', 0),
            inv.get('note', '') or '',
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            is_money = col in (9, 10, 11)
            _apply_cell_style(cell, is_money=is_money)

    # ── 합계 행 ──
    total_row = len(invoices) + 4
    ws.cell(row=total_row, column=1, value='합계')
    ws.cell(row=total_row, column=1).font = Font(name='맑은 고딕', bold=True, size=11)
    ws.merge_cells(start_row=total_row, start_column=1,
                   end_row=total_row, end_column=8)

    for col, field in [(9, 'supply_cost_total'), (10, 'tax_total'), (11, 'total_amount')]:
        total_val = sum(i.get(field, 0) for i in invoices)
        cell = ws.cell(row=total_row, column=col, value=total_val)
        cell.font = Font(name='맑은 고딕', bold=True, size=11)
        cell.number_format = _MONEY_FORMAT
        cell.alignment = Alignment(horizontal='right')
        cell.border = _THIN_BORDER

    # ── 컬럼 너비 자동 조정 ──
    col_widths = [6, 12, 12, 10, 16, 20, 16, 20, 14, 14, 14, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── 파일 출력 ──
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ═══════════════════════════════════════════════════════════
#  3) 은행 거래내역 월간 요약 엑셀
# ═══════════════════════════════════════════════════════════

def export_bank_summary_excel(db, year_month):
    """은행 거래내역 월간 요약 엑셀.

    시트 구성:
    - Sheet1: 전체 거래 목록
    - Sheet2: 카테고리별 요약

    Args:
        db: SupabaseDB 인스턴스
        year_month: 'YYYY-MM'

    Returns:
        io.BytesIO: 엑셀 파일 바이트스트림
    """
    date_from, date_to = _month_range(year_month)
    transactions = db.query_bank_transactions(
        date_from=date_from, date_to=date_to)

    wb = Workbook()

    # ═══ Sheet 1: 전체 거래 목록 ═══
    ws1 = wb.active
    ws1.title = '거래내역'

    ws1.merge_cells('A1:H1')
    title_cell = ws1['A1']
    title_cell.value = f'{year_month} 은행 거래내역'
    title_cell.font = Font(name='맑은 고딕', bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    headers = ['No.', '거래일', '구분', '금액', '잔액', '거래처', '카테고리', '적요']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=3, column=col, value=h)
    _apply_header_style(ws1, 3, len(headers))

    for idx, tx in enumerate(transactions, 1):
        row = idx + 3
        values = [
            idx,
            str(tx.get('transaction_date', '')),
            tx.get('transaction_type', ''),
            tx.get('amount', 0),
            tx.get('balance', 0) or '',
            tx.get('counterpart_name', '') or '',
            tx.get('category', '미분류') or '미분류',
            tx.get('description', '') or '',
        ]
        for col, val in enumerate(values, 1):
            cell = ws1.cell(row=row, column=col, value=val)
            is_money = col in (4, 5)
            _apply_cell_style(cell, is_money=is_money)

    # 합계 행
    total_row = len(transactions) + 4
    total_in = sum(t.get('amount', 0) for t in transactions if t.get('transaction_type') == '입금')
    total_out = sum(t.get('amount', 0) for t in transactions if t.get('transaction_type') == '출금')

    ws1.merge_cells(start_row=total_row, start_column=1,
                    end_row=total_row, end_column=2)
    ws1.cell(row=total_row, column=1, value='입금 합계')
    ws1.cell(row=total_row, column=1).font = Font(name='맑은 고딕', bold=True)
    cell_in = ws1.cell(row=total_row, column=4, value=total_in)
    cell_in.font = Font(name='맑은 고딕', bold=True, color='0000FF')
    cell_in.number_format = _MONEY_FORMAT
    cell_in.border = _THIN_BORDER

    total_row2 = total_row + 1
    ws1.merge_cells(start_row=total_row2, start_column=1,
                    end_row=total_row2, end_column=2)
    ws1.cell(row=total_row2, column=1, value='출금 합계')
    ws1.cell(row=total_row2, column=1).font = Font(name='맑은 고딕', bold=True)
    cell_out = ws1.cell(row=total_row2, column=4, value=total_out)
    cell_out.font = Font(name='맑은 고딕', bold=True, color='FF0000')
    cell_out.number_format = _MONEY_FORMAT
    cell_out.border = _THIN_BORDER

    col_widths = [6, 12, 8, 14, 14, 20, 12, 25]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ═══ Sheet 2: 카테고리별 요약 ═══
    ws2 = wb.create_sheet('카테고리별 요약')

    ws2.merge_cells('A1:D1')
    ws2['A1'].value = f'{year_month} 카테고리별 입출금 요약'
    ws2['A1'].font = Font(name='맑은 고딕', bold=True, size=14)
    ws2['A1'].alignment = Alignment(horizontal='center')

    cat_headers = ['카테고리', '입금', '출금', '건수']
    for col, h in enumerate(cat_headers, 1):
        ws2.cell(row=3, column=col, value=h)
    _apply_header_style(ws2, 3, len(cat_headers))

    # 카테고리별 집계
    by_cat = {}
    for tx in transactions:
        cat = tx.get('category', '미분류') or '미분류'
        if cat not in by_cat:
            by_cat[cat] = {'in': 0, 'out': 0, 'count': 0}
        by_cat[cat]['count'] += 1
        if tx.get('transaction_type') == '입금':
            by_cat[cat]['in'] += tx.get('amount', 0)
        else:
            by_cat[cat]['out'] += tx.get('amount', 0)

    row = 4
    for cat_name in sorted(by_cat.keys()):
        data = by_cat[cat_name]
        values = [cat_name, data['in'], data['out'], data['count']]
        for col, val in enumerate(values, 1):
            cell = ws2.cell(row=row, column=col, value=val)
            is_money = col in (2, 3)
            _apply_cell_style(cell, is_money=is_money)
        row += 1

    # 카테고리 합계
    ws2.cell(row=row, column=1, value='합계')
    ws2.cell(row=row, column=1).font = Font(name='맑은 고딕', bold=True)
    ws2.cell(row=row, column=1).border = _THIN_BORDER

    cell_cat_in = ws2.cell(row=row, column=2, value=total_in)
    cell_cat_in.font = Font(name='맑은 고딕', bold=True)
    cell_cat_in.number_format = _MONEY_FORMAT
    cell_cat_in.border = _THIN_BORDER

    cell_cat_out = ws2.cell(row=row, column=3, value=total_out)
    cell_cat_out.font = Font(name='맑은 고딕', bold=True)
    cell_cat_out.number_format = _MONEY_FORMAT
    cell_cat_out.border = _THIN_BORDER

    cell_cat_cnt = ws2.cell(row=row, column=4, value=len(transactions))
    cell_cat_cnt.font = Font(name='맑은 고딕', bold=True)
    cell_cat_cnt.border = _THIN_BORDER

    cat_col_widths = [16, 14, 14, 10]
    for i, w in enumerate(cat_col_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── 파일 출력 ──
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
