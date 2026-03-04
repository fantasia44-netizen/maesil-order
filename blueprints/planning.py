"""
blueprints/planning.py — 생산계획 Blueprint.
수량 기반 생산계획 엔진 (Production Planning v1).
"""
from flask import (
    Blueprint, render_template, request, current_app, jsonify,
)
from auth import role_required, _log_action
from services.tz_utils import today_kst

planning_bp = Blueprint('planning', __name__, url_prefix='/planning')


@planning_bp.route('/')
@role_required('admin', 'ceo', 'manager', 'production')
def index():
    """생산계획 메인 페이지"""
    return render_template('planning/index.html')


@planning_bp.route('/sales')
@role_required('admin', 'ceo', 'manager', 'production')
def sales_analysis():
    """월간 판매분석 페이지"""
    return render_template('planning/sales.html')


# ── API: 월간 판매분석 ──

@planning_bp.route('/api/sales-analysis')
@role_required('admin', 'ceo', 'manager', 'production')
def api_sales_analysis():
    """월간 판매분석 데이터"""
    try:
        from services.sales_analysis_service import get_monthly_sales_analysis

        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)

        result = get_monthly_sales_analysis(
            current_app.db, year=year, month=month,
        )
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': f'판매분석 오류: {e}'}), 500


# ── API: 생산계획 계산 실행 ──

@planning_bp.route('/api/calculate', methods=['POST'])
@role_required('admin', 'manager', 'production')
def api_calculate():
    """생산계획 계산 실행"""
    try:
        from services.planning_service import calculate_production_plan

        data = request.get_json(silent=True) or {}
        window = int(data.get('sales_window', 7))
        if window < 1 or window > 90:
            window = 7

        critical_days = data.get('critical_days')
        warning_days = data.get('warning_days')
        if critical_days is not None:
            critical_days = max(1, min(int(critical_days), 90))
        if warning_days is not None:
            warning_days = max(1, min(int(warning_days), 180))

        result = calculate_production_plan(
            current_app.db,
            sales_window=window,
            critical_days=critical_days,
            warning_days=warning_days,
            save=True,
        )

        _log_action('production_plan',
                     detail=f"생산계획 생성: {result['summary']['total_targets']}품목, "
                            f"생산필요 {result['summary']['need_production']}건")

        return jsonify(result)
    except Exception as e:
        return jsonify({'error': f'생산계획 계산 오류: {e}'}), 500


# ── API: 생산계획 이력 조회 ──

@planning_bp.route('/api/history')
@role_required('admin', 'ceo', 'manager', 'production')
def api_history():
    """생산계획 이력"""
    try:
        from services.planning_service import get_plan_history
        return jsonify(get_plan_history(current_app.db, limit=30))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── API: 특정 날짜 계획 조회 ──

@planning_bp.route('/api/plan/<plan_date>')
@role_required('admin', 'ceo', 'manager', 'production')
def api_plan_by_date(plan_date):
    """특정 날짜 생산계획 상세"""
    try:
        from services.planning_service import get_plan_by_date
        items = get_plan_by_date(current_app.db, plan_date)
        return jsonify(items)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── API: 품목별 생산계획 설정 수정 ──

@planning_bp.route('/api/config', methods=['POST'])
@role_required('admin', 'manager')
def api_update_config():
    """품목별 안전재고/리드타임/생산대상 설정"""
    try:
        from services.planning_service import update_product_planning_config

        data = request.get_json(silent=True) or {}
        product_name = data.get('product_name', '').strip()
        if not product_name:
            return jsonify({'error': '품목명을 입력하세요.'}), 400

        ok = update_product_planning_config(
            current_app.db,
            product_name=product_name,
            safety_stock=data.get('safety_stock'),
            lead_time_days=data.get('lead_time_days'),
            is_production_target=data.get('is_production_target'),
        )

        if ok:
            _log_action('update_planning_config', target=product_name,
                         detail=str(data))
            return jsonify({'success': True})
        return jsonify({'error': '수정 실패'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── API: 품목 sales_category 일괄 설정 ──

@planning_bp.route('/api/sales-category', methods=['POST'])
@role_required('admin', 'manager')
def api_update_sales_category():
    """품목별 판매분류(sales_category) 수정"""
    try:
        data = request.get_json(silent=True) or {}
        product_name = data.get('product_name', '').strip()
        category = data.get('sales_category', '').strip()
        if not product_name:
            return jsonify({'error': '품목명 필수'}), 400

        current_app.db.client.table('product_costs').update({
            'sales_category': category,
        }).eq('product_name', product_name).execute()

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── API: 품목 분류 엑셀 다운로드 ──

@planning_bp.route('/api/category-download')
@role_required('admin', 'manager')
def api_category_download():
    """product_costs 전체 → 분류 작업용 엑셀 다운로드."""
    from flask import send_file
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    try:
        raw_cost_map = current_app.db.query_product_costs()

        # ── 공백 정규화: 동일 품목 중복 병합 (공백 있는 버전 우선) ──
        cost_map = {}
        norm_lookup = {}   # norm_key → canonical_name
        for name, info in raw_cost_map.items():
            norm = name.replace(' ', '')
            if norm in norm_lookup:
                # 이미 존재 — 분류 정보가 있는 쪽 우선, 같으면 공백 있는 버전 유지
                existing = norm_lookup[norm]
                existing_info = cost_map[existing]
                has_class_new = bool(info.get('cost_type') or info.get('food_type'))
                has_class_old = bool(existing_info.get('cost_type') or existing_info.get('food_type'))
                if has_class_new and not has_class_old:
                    # 새 항목이 분류 있음 → 교체
                    del cost_map[existing]
                    cost_map[name] = info
                    norm_lookup[norm] = name
                elif not has_class_new and not has_class_old and len(name) > len(existing):
                    # 둘 다 분류 없으면 공백 있는(긴) 이름 유지
                    del cost_map[existing]
                    cost_map[name] = info
                    norm_lookup[norm] = name
                # else: 기존 유지
            else:
                cost_map[name] = info
                norm_lookup[norm] = name

        # 판매 데이터에 있는데 product_costs에 없는 품목도 포함
        from services.sales_analysis_service import _fetch_month_sales
        from services.tz_utils import today_kst
        from datetime import datetime
        today = datetime.strptime(today_kst(), '%Y-%m-%d')
        cur_sales = _fetch_month_sales(current_app.db, today.year, today.month)
        py, pm = (today.year, today.month - 1) if today.month > 1 else (today.year - 1, 12)
        prev_sales = _fetch_month_sales(current_app.db, py, pm)

        # 판매 품목 중 product_costs에 없는 것 추가 (공백 정규화 대응)
        all_sales_names = set(cur_sales.keys()) | set(prev_sales.keys())
        for pn in all_sales_names:
            norm = pn.replace(' ', '')
            if norm not in norm_lookup:
                cost_map[pn] = {
                    'cost_type': '', 'food_type': '', 'material_type': '완제품',
                    'cost_price': 0, 'unit': '', 'memo': '(판매데이터에서 자동추가)',
                }
                norm_lookup[norm] = pn

        wb = Workbook()
        ws = wb.active
        ws.title = '품목분류'

        # 헤더
        headers = ['품목명', 'cost_type', 'food_type', 'material_type', '단가', '단위', '재고관리', '비고']
        hfill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        hfont = Font(bold=True, color='FFFFFF', size=11)
        thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin

        # 데이터
        row = 2
        for name in sorted(cost_map.keys()):
            info = cost_map[name]
            is_managed = info.get('is_stock_managed')
            managed_val = 'N' if is_managed is False else 'Y'
            ws.cell(row, 1, name).border = thin
            ws.cell(row, 2, info.get('cost_type') or '').border = thin
            ws.cell(row, 3, info.get('food_type') or '').border = thin
            ws.cell(row, 4, info.get('material_type') or '').border = thin
            ws.cell(row, 5, info.get('cost_price', 0) or 0).border = thin
            ws.cell(row, 6, info.get('unit') or '').border = thin
            ws.cell(row, 7, managed_val).border = thin
            ws.cell(row, 8, info.get('memo') or '').border = thin
            row += 1

        # 열 너비
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 6
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 20

        # 안내 시트
        ws2 = wb.create_sheet('안내')
        ws2.cell(1, 1, '사용법').font = Font(bold=True, size=14)
        ws2.cell(3, 1, '1. B열(cost_type): 생산, OEM, 소분, 매입 중 하나 입력')
        ws2.cell(4, 1, '2. C열(food_type): 농산물, 수산물, 축산물 중 하나 입력')
        ws2.cell(5, 1, '3. G열(재고관리): Y=재고관리 대상, N=재고현황 제외 (아이스팩 등)')
        ws2.cell(6, 1, '4. 작성 후 "판매분석" 페이지에서 업로드하면 일괄 반영됩니다')
        ws2.cell(8, 1, 'A열(품목명)은 수정하지 마세요. 매칭 기준입니다.')
        ws2.column_dimensions['A'].width = 60

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='품목분류_작업용.xlsx',
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── API: 품목 분류 엑셀 업로드 (일괄 반영) ──

@planning_bp.route('/api/category-upload', methods=['POST'])
@role_required('admin', 'manager')
def api_category_upload():
    """엑셀 업로드 → cost_type + food_type 일괄 수정."""
    import pandas as pd

    try:
        f = request.files.get('file')
        if not f:
            return jsonify({'error': '파일을 선택하세요.'}), 400

        df = pd.read_excel(f, sheet_name=0)

        # 컬럼 확인
        if '품목명' not in df.columns:
            return jsonify({'error': "'품목명' 컬럼이 필요합니다."}), 400
        has_ct = 'cost_type' in df.columns
        has_ft = 'food_type' in df.columns
        has_sm = '재고관리' in df.columns
        # 하위호환: sales_category가 있으면 cost_type으로 취급
        has_sc = 'sales_category' in df.columns
        if not has_ct and not has_ft and not has_sc and not has_sm:
            return jsonify({'error': "'cost_type' 또는 'food_type' 또는 '재고관리' 컬럼이 필요합니다."}), 400

        updated = 0
        skipped = 0
        errors = []

        # DB 전체 품목명 로드 → 공백 정규화 매핑 (norm → [실제이름들])
        all_costs = current_app.db.query_product_costs()
        norm_to_names = {}
        for db_name in all_costs.keys():
            nk = db_name.replace(' ', '')
            norm_to_names.setdefault(nk, []).append(db_name)

        for _, row in df.iterrows():
            name = str(row.get('품목명', '')).strip()
            if not name:
                skipped += 1
                continue

            update_data = {}
            if has_ct:
                update_data['cost_type'] = str(row.get('cost_type', '')).strip()
            elif has_sc:
                update_data['cost_type'] = str(row.get('sales_category', '')).strip()
            if has_ft:
                update_data['food_type'] = str(row.get('food_type', '')).strip()
            if has_sm:
                sm_val = str(row.get('재고관리', '')).strip().upper()
                if sm_val in ('N', 'FALSE', '0', 'NO'):
                    update_data['is_stock_managed'] = False
                elif sm_val in ('Y', 'TRUE', '1', 'YES', ''):
                    update_data['is_stock_managed'] = True

            if not update_data:
                skipped += 1
                continue

            try:
                # 공백 정규화로 DB의 모든 변형 이름을 찾아서 일괄 업데이트
                norm = name.replace(' ', '')
                db_names = norm_to_names.get(norm, [name])
                for db_name in db_names:
                    current_app.db.client.table('product_costs').update(
                        update_data
                    ).eq('product_name', db_name).execute()
                updated += 1
            except Exception as e:
                errors.append(f'{name}: {e}')
                if len(errors) > 10:
                    break

        _log_action('bulk_update_product_category',
                     detail=f'일괄 분류 업데이트: {updated}건 반영, {skipped}건 스킵')

        return jsonify({
            'success': True,
            'updated': updated,
            'skipped': skipped,
            'errors': errors[:10],
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── API: 생산대상 품목 목록 + 현재 설정 ──

@planning_bp.route('/api/targets')
@role_required('admin', 'ceo', 'manager', 'production')
def api_targets():
    """생산대상 품목 + 설정 조회"""
    try:
        cost_map = current_app.db.query_product_costs()
        targets = []
        for name, info in cost_map.items():
            cost_type = info.get('cost_type', '')
            material_type = info.get('material_type', '')
            is_target = info.get('is_production_target')

            if is_target is False:
                continue
            if is_target is not True:
                if cost_type != '생산' and material_type not in ('완제품', '반제품'):
                    continue

            targets.append({
                'product_name': name,
                'safety_stock': int(info.get('safety_stock', 0) or 0),
                'lead_time_days': int(info.get('lead_time_days', 0) or 3),
                'is_production_target': is_target if is_target is not None else True,
                'material_type': material_type,
                'unit': info.get('unit', '개') or '개',
            })

        targets.sort(key=lambda x: x['product_name'])
        return jsonify(targets)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
