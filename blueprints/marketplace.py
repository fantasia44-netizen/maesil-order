"""marketplace.py -- 마켓플레이스 API 연동 Blueprint.

네이버 스마트스토어 / 쿠팡 / Cafe24 API 설정, 동기화, 교차검증.
쿠팡 로켓배송 정산 엑셀 업로드 + 광고비 수동 기입.
API 매출 대시보드 (총무/경리용).
"""
import io
import logging
import uuid
from collections import defaultdict
from flask import (Blueprint, render_template, request, current_app,
                   jsonify, flash, redirect, url_for, g)
from flask_login import login_required, current_user
from auth import role_required, _log_action
from services.tz_utils import today_kst, days_ago_kst
from db_utils import get_db

logger = logging.getLogger(__name__)

# 교차검증 결과 서버 메모리 캐시 (세션 쿠키는 용량 부족)
_validation_cache = {}


def _utc_to_kst_str(ts_str: str) -> str:
    """UTC 타임스탬프 문자열 → KST 'MM-DD HH:MM' 형식."""
    from datetime import datetime, timedelta, timezone
    try:
        # '2026-03-10T08:36:00+00:00' or '2026-03-10T08:36:00'
        s = ts_str.replace('Z', '+00:00')
        if '+' in s[10:] or s.count('-') > 2:
            dt = datetime.fromisoformat(s)
        else:
            dt = datetime.fromisoformat(s).replace(tzinfo=timezone.utc)
        kst = dt.astimezone(timezone(timedelta(hours=9)))
        return kst.strftime('%Y-%m-%dT%H:%M')
    except Exception:
        return ts_str[:16] if ts_str else '-'


def _cafe24_callback_url() -> str:
    """Cafe24 OAuth redirect_uri — 항상 HTTPS."""
    u = url_for('cafe24_oauth_redirect', _external=True)
    return u.replace('http://', 'https://', 1)

marketplace_bp = Blueprint('marketplace', __name__, url_prefix='/marketplace')


@marketplace_bp.route('/')
@role_required('admin', 'general')
def index():
    """API 연동 설정 대시보드."""
    db = get_db()
    mgr = g.marketplace

    channels = mgr.get_all_channels()

    # 각 채널의 DB config 로드
    configs = {}
    for ch in channels:
        ch_name = ch['channel']
        rows = db.query_marketplace_api_configs(channel=ch_name)
        configs[ch_name] = rows[0] if rows else {}

    # 최근 동기화 로그 (UTC→KST 변환)
    recent_logs = db.query_api_sync_logs(limit=10)
    for log in recent_logs:
        if log.get('started_at'):
            log['started_at_kst'] = _utc_to_kst_str(log['started_at'])

    return render_template('marketplace/index.html',
                           channels=channels,
                           configs=configs,
                           recent_logs=recent_logs)


@marketplace_bp.route('/config/<channel>', methods=['POST'])
@role_required('admin')
def save_config(channel):
    """채널 API 설정 저장."""
    db = get_db()
    mgr = g.marketplace

    payload = {
        'channel': channel,
        'client_id': request.form.get('client_id', '').strip(),
        'client_secret': request.form.get('client_secret', '').strip(),
        'is_active': request.form.get('is_active') == 'on',
    }

    # 채널별 추가 필드
    extra = {}
    if channel == '쿠팡':
        payload['vendor_id'] = request.form.get('vendor_id', '').strip()
    elif channel == '자사몰':
        payload['mall_id'] = request.form.get('mall_id', '').strip()
    elif channel == '스마트스토어':
        # 네이버 검색광고 API 키 (extra_config에 저장)
        ad_cid = request.form.get('ad_customer_id', '').strip()
        ad_key = request.form.get('ad_api_key', '').strip()
        ad_secret = request.form.get('ad_secret_key', '').strip()
        if ad_cid or ad_key or ad_secret:
            extra['ad_customer_id'] = ad_cid
            extra['ad_api_key'] = ad_key
            extra['ad_secret_key'] = ad_secret

    if extra:
        # 기존 extra_config와 병합
        existing_list = db.query_marketplace_api_configs(channel=channel)
        existing = existing_list[0] if existing_list else None
        old_extra = (existing.get('extra_config') or {}) if existing else {}
        old_extra.update(extra)
        payload['extra_config'] = old_extra

    db.upsert_marketplace_api_config(payload)
    _log_action(f'마켓플레이스 API 설정 저장: {channel}')

    # 클라이언트 재로드
    mgr._load_configs(db)

    # 네이버 광고 클라이언트 재로드
    if channel == '스마트스토어' and extra.get('ad_customer_id'):
        try:
            from services.marketplace.naver_ad_client import NaverAdClient
            current_app.naver_ad = NaverAdClient(extra)
        except Exception as e:
            logger.warning(f'NaverAdClient 재로드 실패: {e}')

    flash(f'{channel} API 설정이 저장되었습니다.', 'success')
    return redirect(url_for('marketplace.index'))


@marketplace_bp.route('/test/<channel>', methods=['POST'])
@role_required('admin')
def test_connection(channel):
    """API 연결 테스트."""
    db = get_db()
    mgr = g.marketplace
    client = mgr.get_client(channel)

    if not client:
        return jsonify({'success': False, 'message': f'{channel} 클라이언트 없음'})

    result = client.test_connection(db)

    # Cafe24 OAuth 인증이 필요한 경우 올바른 redirect_uri로 auth_url 재생성
    if result.get('auth_url'):
        callback_url = _cafe24_callback_url()
        result['auth_url'] = client.get_auth_url(callback_url, state='connect')

    return jsonify(result)


@marketplace_bp.route('/oauth/authorize/<channel>')
@role_required('admin')
def oauth_authorize(channel):
    """OAuth2 인증 시작 (Cafe24 등)."""
    mgr = g.marketplace
    client = mgr.get_client(channel)
    if not client:
        flash(f'{channel} 클라이언트 없음', 'danger')
        return redirect(url_for('marketplace.index'))

    callback_url = _cafe24_callback_url()
    auth_url = client.get_auth_url(callback_url, state='connect')
    return redirect(auth_url)


@marketplace_bp.route('/oauth/callback/<channel>')
@role_required('admin')
def oauth_callback(channel):
    """OAuth2 콜백 — 인가 코드를 토큰으로 교환."""
    db = get_db()
    mgr = g.marketplace
    client = mgr.get_client(channel)

    code = request.args.get('code', '')
    error = request.args.get('error', '')

    if error:
        flash(f'{channel} OAuth 인증 거부: {error}', 'danger')
        return redirect(url_for('marketplace.index'))

    if not code:
        flash(f'{channel} 인가 코드가 없습니다', 'danger')
        return redirect(url_for('marketplace.index'))

    callback_url = _cafe24_callback_url()
    result = client.exchange_code(db, code, callback_url)

    if result is True:
        mgr._load_configs(db)
        _log_action(f'{channel} OAuth 인증 완료')
        flash(f'{channel} OAuth 인증 성공!', 'success')
    else:
        flash(f'{channel} 토큰 교환 실패: {result}', 'danger')

    return redirect(url_for('marketplace.index'))


@marketplace_bp.route('/sync', methods=['POST'])
@role_required('admin', 'general')
def sync():
    """수동 동기화 실행."""
    try:
        db = get_db()
        mgr = g.marketplace

        channel = request.form.get('channel', '')
        sync_type = request.form.get('sync_type', 'orders')
        date_from = request.form.get('date_from', days_ago_kst(7))
        date_to = request.form.get('date_to', today_kst())

        from services.marketplace_sync_service import (
            sync_orders, sync_settlements, sync_revenue_fees, sync_ad_costs)

        results = {}

        if sync_type in ('orders', 'all'):
            results['orders'] = sync_orders(
                db, mgr, channel, date_from, date_to,
                triggered_by=current_user.username
            )

        if sync_type in ('settlements', 'all'):
            results['settlements'] = sync_settlements(
                db, mgr, channel, date_from, date_to,
                triggered_by=current_user.username
            )

        # 쿠팡: 매출내역(revenue-history)으로 수수료/정산 업데이트
        if channel == '쿠팡' and sync_type in ('settlements', 'all'):
            results['revenue_fees'] = sync_revenue_fees(
                db, mgr, channel, date_from, date_to,
                triggered_by=current_user.username
            )

        # 네이버 검색광고 비용 동기화
        if channel == '스마트스토어' and sync_type in ('settlements', 'all'):
            ad_client = getattr(current_app, 'naver_ad', None)
            # 앱 시작 시 초기화 안 됐으면 DB에서 다시 로드 시도
            if not ad_client or not ad_client.is_ready:
                try:
                    from services.marketplace.naver_ad_client import NaverAdClient
                    cfgs = db.query_marketplace_api_configs(channel='스마트스토어')
                    if cfgs:
                        ec = (cfgs[0].get('extra_config') or {})
                        if ec.get('ad_customer_id'):
                            ad_client = NaverAdClient(ec)
                            current_app.naver_ad = ad_client
                except Exception:
                    pass
            if ad_client and ad_client.is_ready:
                results['ad_costs'] = sync_ad_costs(
                    db, ad_client, date_from, date_to,
                    triggered_by=current_user.username
                )

        _log_action(f'마켓플레이스 동기화: {channel} {sync_type} {date_from}~{date_to}')
        return jsonify(results)

    except Exception as e:
        logger.error(f'[Sync] 동기화 오류: {e}', exc_info=True)
        return jsonify({'error': str(e)}), 500


@marketplace_bp.route('/api/cj-tracking-upload', methods=['POST'])
@role_required('admin', 'general')
def cj_tracking_upload():
    """CJ 송장결과 파일 업로드 → 이름+전화뒷4자리로 매칭 → order_shipping 반영."""
    import io, re
    file = request.files.get('file')
    if not file:
        return jsonify({'ok': False, 'error': '파일이 없습니다.'})

    filename = file.filename.lower()
    if not filename.endswith(('.xlsx', '.xls')):
        return jsonify({'ok': False, 'error': 'xlsx 또는 xls 파일만 지원합니다.'})

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), read_only=False)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 2:
            return jsonify({'ok': False, 'error': '데이터가 없습니다.'})

        headers = [str(c or '').replace('\r\n', '').replace(' ', '') for c in rows[0]]

        # CJ 표준 형식: 컬럼 인덱스로 감지 (40컬럼)
        # 또는 컬럼명으로 감지
        col_invoice = col_name = col_phone = None

        for i, h in enumerate(headers):
            if '운송장번호' in h:
                col_invoice = i
            elif h == '받는분' or (h == '받는분' and col_name is None):
                col_name = i
            elif '받는분전화' in h or '받는분휴대' in h:
                col_phone = i

        # 컬럼명으로 못 찾으면 CJ 표준 인덱스 사용 (40컬럼)
        if col_invoice is None and len(headers) >= 22:
            col_invoice = 7   # 운송장번호
        if col_name is None and len(headers) >= 22:
            col_name = 20     # 받는분
        if col_phone is None and len(headers) >= 22:
            col_phone = 21    # 받는분전화번호

        if col_invoice is None or col_name is None or col_phone is None:
            return jsonify({'ok': False,
                            'error': '필수 컬럼을 찾을 수 없습니다. '
                                     '"운송장번호", "받는분", "받는분전화번호" 컬럼이 필요합니다.'})

        # CJ 파일 파싱: {이름+전화뒷4자리 → 운송장번호}
        cj_map = {}
        for row in rows[1:]:
            row_list = list(row)
            inv_no = str(row_list[col_invoice] or '').strip()
            name = str(row_list[col_name] or '').strip()
            phone = re.sub(r'[^0-9]', '', str(row_list[col_phone] or ''))

            if not inv_no or not name or not phone:
                continue
            key = f"{name}_{phone[-4:]}" if len(phone) >= 4 else f"{name}_{phone}"
            cj_map[key] = inv_no

        if not cj_map:
            return jsonify({'ok': False, 'error': '유효한 송장 데이터가 없습니다.'})

        # api_orders에서 수취인 정보 추출 → CJ 파일과 매칭
        db = get_db()
        from datetime import date, timedelta
        today = date.today().strftime('%Y-%m-%d')
        week_ago = (date.today() - timedelta(days=7)).strftime('%Y-%m-%d')

        api_rows = db.query_api_orders(date_from=week_ago, date_to=today)

        # raw_data에서 채널별 수취인 이름+전화번호 추출
        collected = []
        for row in api_rows:
            ch = row.get('channel', '')
            raw = row.get('raw_data') or {}
            order_no = row.get('api_line_id') or row.get('api_order_id', '')
            name = phone = ''

            if ch in ('스마트스토어', '해미애찬'):
                po = raw.get('productOrder', {})
                sa = po.get('shippingAddress', {})
                name = sa.get('name', '')
                phone = sa.get('tel1', '') or sa.get('tel2', '')
            elif ch == '쿠팡':
                # 쿠팡 ordersheets: receiver.{name, safeNumber, receiverNumber}
                rcv = raw.get('receiver', {})
                name = rcv.get('name', '')
                phone = rcv.get('safeNumber', '') or rcv.get('receiverNumber', '')
            elif ch == '자사몰':
                name = raw.get('shipping_name', '') or raw.get('buyer_name', '')
                phone = raw.get('shipping_phone', '') or raw.get('buyer_cellphone', '')

            if name and order_no:
                collected.append({
                    'channel': ch,
                    'order_no': order_no,
                    'name': name,
                    'phone': re.sub(r'[^0-9]', '', str(phone)),
                })

        # 이름+전화뒷4자리로 매칭
        updates = []
        matched = 0
        seen = set()
        for item in collected:
            key = f"{item['name']}_{item['phone'][-4:]}" if len(item['phone']) >= 4 else f"{item['name']}_{item['phone']}"
            if key in cj_map and item['order_no'] not in seen:
                updates.append({
                    'channel': item['channel'],
                    'order_no': item['order_no'],
                    'invoice_no': cj_map[key],
                    'courier': 'CJ대한통운',
                })
                matched += 1
                seen.add(item['order_no'])

        if not updates:
            return jsonify({'ok': False,
                            'error': f'매칭 건이 없습니다. CJ 파일 {len(cj_map)}건, '
                                     f'수집된 주문 {len(collected)}건'})

        success_count = db.bulk_update_shipping_invoices(updates)

        _log_action(f'CJ 송장파일 업로드: {file.filename} — '
                    f'매칭 {matched}건, DB반영 {success_count}건')

        return jsonify({
            'ok': True,
            'total': len(cj_map),
            'matched': matched,
            'success': success_count,
        })
    except Exception as e:
        logger.error(f'[CJ Upload] 오류: {e}', exc_info=True)
        return jsonify({'ok': False, 'error': str(e)})


@marketplace_bp.route('/push-invoices', methods=['POST'])
@role_required('admin', 'general')
def push_invoices_route():
    """마켓플레이스에 송장번호 일괄 전송 (발송처리)."""
    try:
        db = get_db()
        mgr = g.marketplace
        channel = request.form.get('channel', '')

        if not channel:
            return jsonify({'error': '채널을 선택하세요.'}), 400

        from services.marketplace_sync_service import push_invoices
        result = push_invoices(
            db, mgr, channel,
            triggered_by=current_user.username,
        )

        _log_action(f'송장 전송: {channel} — '
                     f'{result.get("success", 0)}/{result.get("total", 0)}건 성공')
        return jsonify(result)

    except Exception as e:
        logger.error(f'[Push] 송장 전송 오류: {e}', exc_info=True)
        return jsonify({'error': str(e)}), 500


@marketplace_bp.route('/api/pending-invoices')
@role_required('admin', 'general')
def api_pending_invoices():
    """채널별 송장 push 대기건 카운트."""
    db = get_db()
    mgr = g.marketplace
    counts = {}
    for channel in mgr.get_active_channels():
        pending = db.query_pending_invoice_push(channel=channel)
        pushable = [p for p in pending if p.get('api_order_id')]
        counts[channel] = len(pushable)
    return jsonify(counts)


@marketplace_bp.route('/api/pending-invoices/list')
@role_required('admin', 'general')
def api_pending_invoices_list():
    """채널별 송장 push 대기건 상세 목록."""
    db = get_db()
    channel = request.args.get('channel', '')
    if not channel:
        return jsonify({'error': '채널 필수'}), 400

    pending = db.query_pending_invoice_push(channel=channel)
    items = []
    for p in pending:
        # raw_data에서 상품명 추출
        raw = p.get('raw_data') or {}
        product_name = raw.get('product_name', '') or raw.get('productOrder', {}).get('productName', '')
        items.append({
            'channel': p.get('channel', ''),
            'order_no': p.get('order_no', ''),
            'invoice_no': p.get('invoice_no', ''),
            'courier': p.get('courier', ''),
            'api_order_id': p.get('api_order_id', ''),
            'api_line_id': p.get('api_line_id', ''),
            'product_name': product_name,
        })
    return jsonify(items)


@marketplace_bp.route('/api/push-invoices/selective', methods=['POST'])
@role_required('admin', 'general')
def push_invoices_selective():
    """선택적 송장 push (테스트용, 최대 10건)."""
    import json
    db = get_db()
    mgr = g.marketplace
    channel = request.form.get('channel', '')
    order_nos_raw = request.form.get('order_nos', '[]')

    if not channel:
        return jsonify({'error': '채널을 선택하세요.'}), 400

    try:
        order_nos = json.loads(order_nos_raw)
    except (json.JSONDecodeError, TypeError):
        return jsonify({'error': 'order_nos 형식 오류'}), 400

    if not order_nos or len(order_nos) > 10:
        return jsonify({'error': f'1~10건 선택 필요 (현재 {len(order_nos)}건)'}), 400

    from services.marketplace_sync_service import push_invoices
    result = push_invoices(
        db, mgr, channel,
        triggered_by=current_user.username,
        order_nos=order_nos,
        max_batch=10,
    )

    _log_action(f'선택적 송장 전송: {channel} — '
                f'{result.get("success", 0)}/{result.get("total", 0)}건 성공')
    return jsonify(result)


@marketplace_bp.route('/diag/<channel>')
@role_required('admin')
def diag(channel):
    """API 진단 — 실제 API 응답 확인."""
    db = get_db()
    mgr = g.marketplace
    client = mgr.get_client(channel)

    if not client:
        return jsonify({'error': f'{channel} 클라이언트 없음'})

    info = {
        'channel': channel,
        'is_ready': client.is_ready,
        'config_keys': list(client.config.keys()),
        'has_token': bool(client.config.get('access_token')),
        'token_expires': client.config.get('token_expires_at', ''),
    }

    # 토큰 갱신 시도
    try:
        refresh_ok = client.refresh_token(db)
        info['refresh_ok'] = refresh_ok
    except Exception as e:
        info['refresh_error'] = str(e)

    # 주문 조회 (최근 1일, 최대 3건)
    try:
        from services.tz_utils import today_kst, days_ago_kst
        orders = client.fetch_orders(days_ago_kst(1), today_kst())
        info['orders_count'] = len(orders)
        if orders:
            info['sample_order'] = {k: str(v)[:50] for k, v in orders[0].items()}
    except Exception as e:
        info['orders_error'] = str(e)

    return jsonify(info)


@marketplace_bp.route('/sync/log')
@role_required('admin', 'general')
def sync_log():
    """동기화 이력."""
    db = get_db()
    channel = request.args.get('channel', '')
    logs = db.query_api_sync_logs(channel=channel or None, limit=100)
    for log in logs:
        if log.get('started_at'):
            log['started_at_kst'] = _utc_to_kst_str(log['started_at'])
        if log.get('finished_at'):
            log['finished_at_kst'] = _utc_to_kst_str(log['finished_at'])
    return render_template('marketplace/sync_log.html', logs=logs, channel=channel)


@marketplace_bp.route('/validation')
@role_required('admin', 'general')
def validation():
    """교차검증 대시보드."""
    channel = request.args.get('channel', '')
    date_from = request.args.get('date_from', days_ago_kst(7))
    date_to = request.args.get('date_to', today_kst())

    # AJAX run_validation()에서 캐시한 결과 사용 (중복 실행 방지)
    cache_key = request.args.get('_vc')
    validation_result = None
    if cache_key:
        validation_result = _validation_cache.pop(cache_key, None)

    return render_template('marketplace/validation.html',
                           channel=channel,
                           date_from=date_from,
                           date_to=date_to,
                           result=validation_result)


@marketplace_bp.route('/api/validation/run', methods=['POST'])
@role_required('admin', 'general')
def run_validation():
    """교차검증 실행 (AJAX)."""
    db = get_db()
    channel = request.form.get('channel', '')
    date_from = request.form.get('date_from', days_ago_kst(7))
    date_to = request.form.get('date_to', today_kst())

    if not channel:
        return jsonify({'error': '채널을 선택하세요'}), 400

    from services.marketplace_validation_service import validate_orders, validate_settlements

    orders_result = validate_orders(db, channel, date_from, date_to)
    settlements_result = validate_settlements(db, channel, date_from, date_to)

    # 서버 메모리 캐시 → GET validation()에서 재사용 (중복 실행 방지)
    cache_key = uuid.uuid4().hex[:12]
    _validation_cache[cache_key] = orders_result
    # 오래된 캐시 정리 (10개 초과 시 가장 오래된 것 삭제)
    while len(_validation_cache) > 10:
        _validation_cache.pop(next(iter(_validation_cache)))

    result = {
        'orders': orders_result,
        'settlements': settlements_result,
        '_vc': cache_key,
    }

    _log_action(f'마켓플레이스 교차검증: {channel} {date_from}~{date_to}')
    return jsonify(result)


# ── API 주문수집 테스트 (샌드박스 — DB 미반영) ──

@marketplace_bp.route('/api/test-collect', methods=['POST'])
@role_required('admin', 'general')
def test_collect():
    """API 주문수집 테스트 — fetch_orders()만 호출, DB 저장 안 함."""
    try:
        mgr = g.marketplace
        channel = request.form.get('channel', '')
        date_from = request.form.get('date_from', days_ago_kst(7))
        date_to = request.form.get('date_to', today_kst())

        channels = [channel] if channel != 'all' else mgr.get_active_channels()
        all_orders = []
        channel_status = {}

        for ch in channels:
            client = mgr.get_client(ch)
            if not client:
                channel_status[ch] = '클라이언트 없음 (API 설정 확인)'
                continue
            if not client.is_ready:
                channel_status[ch] = '인증 미완료 (토큰 갱신 필요)'
                # 토큰 갱신 시도
                try:
                    db = get_db()
                    client.refresh_token(db)
                    if not client.is_ready:
                        continue
                except Exception as e:
                    channel_status[ch] = f'토큰 갱신 실패: {e}'
                    continue
            try:
                orders = client.fetch_orders(date_from, date_to,
                                             status_filter='invoice_target')
                for o in orders:
                    o['channel'] = ch
                all_orders.extend(orders)
                channel_status[ch] = f'{len(orders)}건 수집'
            except Exception as e:
                logger.warning(f'[TestCollect] {ch} 수집 오류: {e}')
                channel_status[ch] = f'오류: {e}'
                all_orders.append({'channel': ch, '_error': str(e)})

        # 날짜순 정렬
        all_orders.sort(key=lambda x: x.get('order_date', ''), reverse=True)

        # 필요한 필드만 추출 (raw_data 제외 — 용량 절감)
        slim = []
        for o in all_orders:
            if '_error' in o:
                slim.append(o)
                continue
            slim.append({
                'channel': o.get('channel', ''),
                'order_date': o.get('order_date', ''),
                'api_order_id': o.get('api_order_id', ''),
                'api_line_id': o.get('api_line_id', ''),
                'product_name': o.get('product_name', ''),
                'option_name': o.get('option_name', ''),
                'qty': o.get('qty', 0),
                'unit_price': o.get('unit_price', 0),
                'total_amount': o.get('total_amount', 0),
                'settlement_amount': o.get('settlement_amount', 0),
                'commission': o.get('commission', 0),
                'shipping_fee': o.get('shipping_fee', 0),
                'order_status': o.get('order_status', ''),
            })

        return jsonify({'count': len(slim), 'orders': slim, 'channel_status': channel_status})

    except Exception as e:
        logger.error(f'[TestCollect] 오류: {e}', exc_info=True)
        return jsonify({'error': str(e)}), 500


@marketplace_bp.route('/api/test-collect/download', methods=['POST'])
@role_required('admin', 'general')
def test_collect_download():
    """API 주문수집 테스트 결과 엑셀 다운로드 — DB 미반영."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        mgr = g.marketplace
        channel = request.form.get('channel', '')
        date_from = request.form.get('date_from', days_ago_kst(7))
        date_to = request.form.get('date_to', today_kst())

        channels = [channel] if channel != 'all' else mgr.get_active_channels()
        all_orders = []

        for ch in channels:
            client = mgr.get_client(ch)
            if not client or not client.is_ready:
                continue
            try:
                orders = client.fetch_orders(date_from, date_to,
                                             status_filter='invoice_target')
                for o in orders:
                    o['channel'] = ch
                all_orders.extend(orders)
            except Exception as e:
                logger.warning(f'[TestCollect DL] {ch} 수집 오류: {e}')

        all_orders.sort(key=lambda x: x.get('order_date', ''))

        # 엑셀 생성
        wb = Workbook()
        ws = wb.active
        ws.title = 'API 주문수집 테스트'

        headers = ['채널', '주문일자', '주문번호', '상품주문번호', '상품명',
                    '옵션명', '수량', '단가', '결제금액', '정산예정액',
                    '수수료', '배송비', '주문상태']
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=10)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for i, o in enumerate(all_orders, 2):
            ws.cell(row=i, column=1, value=o.get('channel', ''))
            ws.cell(row=i, column=2, value=o.get('order_date', ''))
            ws.cell(row=i, column=3, value=o.get('api_order_id', ''))
            ws.cell(row=i, column=4, value=o.get('api_line_id', ''))
            ws.cell(row=i, column=5, value=o.get('product_name', ''))
            ws.cell(row=i, column=6, value=o.get('option_name', ''))
            ws.cell(row=i, column=7, value=o.get('qty', 0))
            ws.cell(row=i, column=8, value=o.get('unit_price', 0))
            ws.cell(row=i, column=9, value=o.get('total_amount', 0))
            ws.cell(row=i, column=10, value=o.get('settlement_amount', 0))
            ws.cell(row=i, column=11, value=o.get('commission', 0))
            ws.cell(row=i, column=12, value=o.get('shipping_fee', 0))
            ws.cell(row=i, column=13, value=o.get('order_status', ''))

        # 컬럼 너비 자동 조정
        col_widths = [10, 12, 18, 18, 30, 20, 6, 10, 12, 12, 10, 10, 10]
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        from flask import send_file
        ch_label = channel if channel != 'all' else '전체'
        filename = f'API주문수집_테스트_{ch_label}_{date_from}_{date_to}.xlsx'
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)

    except Exception as e:
        logger.error(f'[TestCollect DL] 오류: {e}', exc_info=True)
        return jsonify({'error': str(e)}), 500


# ── API 주문 → 송장 생성 (테스트 — DB 미반영) ──

def _extract_customer(channel, order):
    """raw_data에서 고객 배송정보 추출 (채널별)."""
    raw = order.get('raw_data', {})
    info = {'name': '', 'addr': '', 'addr2': '', 'phone': '', 'phone2': '', 'memo': ''}

    if channel in ('스마트스토어', '해미애찬'):
        po = raw.get('productOrder', {})
        sa = po.get('shippingAddress', {})
        info['name'] = sa.get('name', '')
        info['addr'] = sa.get('baseAddress', '')
        info['addr2'] = sa.get('detailedAddress', '')
        info['phone'] = sa.get('tel1', '')
        info['phone2'] = sa.get('tel2', '')
        info['memo'] = po.get('shippingMemo', '')
    elif channel == '쿠팡':
        rcv = raw.get('receiver', {})
        info['name'] = rcv.get('name', '')
        info['addr'] = rcv.get('addr1', '')
        info['addr2'] = rcv.get('addr2', '')
        info['phone'] = rcv.get('receiverTel1', '')
        info['phone2'] = rcv.get('receiverTel2', '')
        info['memo'] = raw.get('parcelPrintMessage', '')
    elif channel == '자사몰':
        order = raw.get('order', raw)
        receivers = order.get('receivers', [])
        rcv = receivers[0] if receivers else {}
        info['name'] = rcv.get('name', rcv.get('receiver_name',
                       order.get('shipping_name', order.get('receiver_name', ''))))
        info['addr'] = (rcv.get('address1', '') + ' ' + rcv.get('address2', '')).strip() \
                       or order.get('shipping_address', order.get('receiver_address', ''))
        info['phone'] = rcv.get('cellphone', rcv.get('receiver_cellphone',
                        order.get('shipping_phone', order.get('receiver_phone', ''))))
        info['phone2'] = rcv.get('phone', rcv.get('receiver_phone', ''))
        info['memo'] = rcv.get('shipping_message', order.get('shipping_memo', ''))

    return info


def _sanitize_receiver_name(name, fallback_name=''):
    """수취인명 검증 — CJ 택배 거부 방지.

    '집', '집.', '회사', 1글자 등 비정상 이름은 주문자명으로 대체.
    """
    _INVALID = {'집', '집.', '회사', '회사.', '사무실', '사무실.', '경비실',
                '문앞', '현관', '-', '.', '..', '본인', '자택'}
    n = str(name or '').strip()
    fb = str(fallback_name or '').strip()
    if not n or n in _INVALID or len(n) == 1:
        return fb or n  # fallback도 없으면 원본 유지
    return n


def _api_orders_to_excel_df(orders, channel):
    """API raw_data → 채널별 엑셀 컬럼 형식 DataFrame.

    OrderProcessor가 인식하는 엑셀 컬럼명으로 변환하여
    채널별 모든 차이점(옵션/단일상품/공백 등)이 동일하게 적용되도록 함.
    """
    import pandas as pd
    rows = []
    _coupang_logged = False

    for o in orders:
        raw = o.get('raw_data', {})

        if channel in ('스마트스토어', '해미애찬'):
            po = raw.get('productOrder', {})
            # N배송(네이버 풀필먼트) 제외 — 직접 배송 대상 아님
            if po.get('deliveryAttributeType') == 'ARRIVAL_GUARANTEE':
                continue
            sa = po.get('shippingAddress', {})
            rows.append({
                '상품주문번호': str(po.get('productOrderId', '')),
                '주문번호': str(raw.get('order', {}).get('orderId', '')),
                '주문일시': str(raw.get('order', {}).get('orderDate', ''))[:16],
                '상품명': po.get('productName', ''),
                '옵션정보': po.get('productOption', ''),
                '수량': int(po.get('quantity', 0)),
                '수취인명': _sanitize_receiver_name(
                    sa.get('name', ''),
                    raw.get('order', {}).get('ordererName', '')),
                '수취인연락처1': sa.get('tel1', ''),
                '수취인연락처2': sa.get('tel2', ''),
                '기본배송지': sa.get('baseAddress', ''),
                '상세배송지': sa.get('detailedAddress', ''),
                '배송메세지': po.get('shippingMemo', ''),
                '주문상태': po.get('productOrderStatus', ''),
                '상품가격': int(po.get('unitPrice', 0)),
                '최종 상품별 총 주문금액': int(po.get('totalPaymentAmount', 0)),
                '정산예정금액': int(po.get('expectedSettlementAmount', 0)),
                '배송비 합계': int(po.get('deliveryFeeAmount', 0)),
                '배송속성': po.get('deliveryAttributeType', ''),
            })

        elif channel == '쿠팡':
            items = raw.get('orderItems', [])
            recv = raw.get('receiver', {})
            # api_orders는 라인별 1행 저장, raw_data에는 전체 orderItems 포함
            # → api_line_id(vendorItemId)와 매칭되는 아이템만 추출
            # 주의: line_id 없거나 매칭 실패 시 반드시 1건만 (7배 뻥튀기 방지)
            line_id = str(o.get('api_line_id', ''))
            matched_items = [it for it in items if str(it.get('vendorItemId', '')) == line_id] if line_id else []
            is_first_line = (items and str(items[0].get('vendorItemId', '')) == line_id)
            for item in (matched_items[:1] or items[:1]):
                rows.append({
                    '주문번호': str(raw.get('orderId', '')),
                    '묶음배송번호': str(raw.get('shipmentBoxId', '')),
                    '주문일': str(raw.get('orderedAt', ''))[:10],
                    '등록상품명': item.get('sellerProductName', item.get('vendorItemName', '')),
                    '등록옵션명': item.get('sellerProductItemName', ''),
                    '노출상품명': item.get('vendorItemName', ''),
                    '구매수(수량)': int(item.get('shippingCount', 0)),
                    '수취인이름': _sanitize_receiver_name(
                        recv.get('name', ''),
                        raw.get('orderer', {}).get('name', '')),
                    '수취인전화번호': recv.get('safeNumber', recv.get('receiverNumber', '')),
                    '수취인 주소': f"{recv.get('addr1', '')} {recv.get('addr2', '')}".strip(),
                    '배송메세지': raw.get('parcelPrintMessage', ''),
                    '주문상태명': raw.get('status', ''),
                    '옵션판매가(판매단가)': int(item.get('salesPrice', 0)),
                    '결제액': int(item.get('orderPrice', 0)),
                    '배송비': int(raw.get('shippingPrice', 0)) if is_first_line else 0,
                })

        elif channel == '자사몰':
            item = raw.get('item', raw)
            order = raw.get('order', raw)
            # Cafe24: receivers 배열에서 배송 정보 추출
            receivers = order.get('receivers', [])
            rcv = receivers[0] if receivers else {}
            rows.append({
                '쇼핑몰번호': '1',
                '주문번호': str(order.get('order_id', '')),
                '발주일': str(order.get('order_date', ''))[:10],
                '주문상품명': item.get('product_name', ''),
                '옵션정보': item.get('option_value', '') or o.get('option_name', ''),
                '수량': int(item.get('quantity', item.get('qty', 1)) or 1),
                '수령인': _sanitize_receiver_name(
                    rcv.get('name', rcv.get('receiver_name',
                             order.get('shipping_name', order.get('receiver_name', '')))),
                    order.get('buyer_name', order.get('member_name', ''))),
                '핸드폰': rcv.get('cellphone', rcv.get('receiver_cellphone',
                         order.get('shipping_phone', order.get('receiver_phone', '')))),
                '수령지전화': rcv.get('phone', rcv.get('receiver_phone',
                             order.get('shipping_cellphone', ''))),
                '주소': (rcv.get('address1', '') + ' ' + rcv.get('address2', '')).strip()
                       or order.get('shipping_address', order.get('receiver_address', '')),
                '비고': rcv.get('shipping_message', order.get('shipping_memo', '')),
                '판매가': int(float(item.get('product_price', 0) or 0)),
                '결제금액': int(float(item.get('payment_amount',
                               item.get('actual_payment_amount', 0)) or 0)),
                '배송비': int(float(order.get('shipping_fee', 0) or 0)),
            })

    return pd.DataFrame(rows) if rows else pd.DataFrame()


@marketplace_bp.route('/api/test-collect/invoice', methods=['POST'])
@role_required('admin', 'general')
def test_collect_invoice():
    """API 주문 → OrderProcessor 송장 생성 (테스트) — DB 미반영.

    API raw_data를 채널별 엑셀 형식으로 변환 후
    기존 OrderProcessor.run()을 save_to_db=False로 호출.
    채널별 옵션매칭/필터링/합포장 로직이 100% 동일하게 적용됩니다.
    """
    import os, tempfile, zipfile

    try:
        mgr = g.marketplace
        db = get_db()
        channel = request.form.get('channel', '')
        date_from = request.form.get('date_from', days_ago_kst(7))
        date_to = request.form.get('date_to', today_kst())

        channels = [channel] if channel != 'all' else mgr.get_active_channels()
        channel_status = {}
        all_files = []
        all_logs = []
        total_unmatched = []

        output_dir = tempfile.mkdtemp(prefix='api_invoice_')

        for ch in channels:
            client = mgr.get_client(ch)
            if not client:
                channel_status[ch] = '클라이언트 없음'
                continue
            if not client.is_ready:
                try:
                    client.refresh_token(db)
                    if not client.is_ready:
                        channel_status[ch] = '인증 미완료'
                        continue
                except Exception as e:
                    channel_status[ch] = f'토큰 갱신 실패: {e}'
                    continue

            try:
                orders = client.fetch_orders(date_from, date_to,
                                             status_filter='invoice_target')
                if not orders:
                    channel_status[ch] = '0건'
                    continue

                # API raw_data → 엑셀 형식 DataFrame
                df = _api_orders_to_excel_df(orders, ch)
                if df.empty:
                    channel_status[ch] = 'DataFrame 변환 실패'
                    continue

                logger.info(f'[TestInvoice] {ch}: {len(orders)}건 수집 → {len(df)}행 DataFrame')
                logger.info(f'[TestInvoice] {ch} 컬럼: {list(df.columns)}')

                # DataFrame → BytesIO 엑셀
                excel_buf = io.BytesIO()
                df.to_excel(excel_buf, index=False, engine='openpyxl')
                excel_buf.seek(0)
                excel_buf.name = f'{ch}_api_orders.xlsx'

                # OrderProcessor 실행 (save_to_db=False!)
                from services.order_processor import OrderProcessor
                proc = OrderProcessor()
                result = proc.run(
                    mode=ch,
                    order_file=excel_buf,
                    option_file=None,
                    invoice_file=None,
                    target_type='송장',
                    output_dir=output_dir,
                    db=db,
                    option_source='db',
                    save_to_db=False,   # DB 미반영!
                    uploaded_by='(API테스트)',
                )

                if result.get('success'):
                    all_files.extend(result.get('files', []))
                    channel_status[ch] = f'{len(orders)}건 수집 → 송장 생성 성공'
                elif result.get('unmatched'):
                    total_unmatched.extend(result['unmatched'])
                    channel_status[ch] = f"미매칭 {len(result['unmatched'])}건: {result.get('error', '')[:80]}"
                else:
                    channel_status[ch] = f"처리 실패: {result.get('error', '')[:100]}"

                all_logs.extend(result.get('logs', []))

            except Exception as e:
                logger.error(f'[TestInvoice] {ch} 오류: {e}', exc_info=True)
                channel_status[ch] = f'오류: {e}'

        if not all_files:
            return jsonify({
                'error': '생성된 송장 파일 없음',
                'channel_status': channel_status,
                'logs': all_logs[-20:],
                'unmatched': total_unmatched[:30],
            })

        # 파일이 1개면 그대로, 여러 개면 ZIP으로 묶기
        from flask import send_file
        from services.tz_utils import now_kst
        ts = now_kst().strftime('%Y%m%d_%H%M%S')
        ch_label = channel if channel != 'all' else '전체'

        # 요약 헤더 생성
        summary_parts = []
        summary_parts.append(f"files={len(all_files)}")
        summary_parts.append(f"unmatched={len(total_unmatched)}")
        for ch, st in channel_status.items():
            summary_parts.append(f"{ch}={st}")
        from urllib.parse import quote
        summary_header = quote(','.join(summary_parts))
        unmatched_header = quote('|'.join(total_unmatched[:20]))

        def _add_headers(resp):
            resp.headers['X-Invoice-Summary'] = summary_header
            resp.headers['X-Invoice-Unmatched'] = unmatched_header
            resp.headers['Access-Control-Expose-Headers'] = 'X-Invoice-Summary, X-Invoice-Unmatched'
            return resp

        if len(all_files) == 1:
            fp = all_files[0]
            mime = ('application/vnd.ms-excel' if fp.endswith('.xls')
                    else 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            resp = send_file(
                fp,
                mimetype=mime,
                as_attachment=True,
                download_name=os.path.basename(fp)
            )
            return _add_headers(resp)
        else:
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                for fp in all_files:
                    zf.write(fp, os.path.basename(fp))
            zip_buf.seek(0)
            resp = send_file(
                zip_buf,
                mimetype='application/zip',
                as_attachment=True,
                download_name=f'API테스트_송장_{ch_label}_{ts}.zip'
            )
            return _add_headers(resp)

    except Exception as e:
        logger.error(f'[TestInvoice] 오류: {e}', exc_info=True)
        return jsonify({'error': str(e)}), 500


@marketplace_bp.route('/api/test-collect/invoice-preview', methods=['POST'])
@role_required('admin', 'general')
def test_collect_invoice_preview():
    """송장 생성 미리보기 — option_matcher 기반 매칭 결과 JSON 반환.

    OrderProcessor와 동일한 option_matcher 공통 모듈로 매칭 수/미매칭 목록을 계산.
    OrderProcessor는 미매칭 1건이라도 있으면 중단하지만,
    미리보기는 전체 매칭/미매칭 수를 정확히 보여줍니다.
    """
    try:
        mgr = g.marketplace
        db = get_db()
        channel = request.form.get('channel', '')
        date_from = request.form.get('date_from', days_ago_kst(7))
        date_to = request.form.get('date_to', today_kst())

        channels = [channel] if channel != 'all' else mgr.get_active_channels()
        channel_status = {}
        total_count = 0
        total_matched = 0
        total_unmatched = []
        total_excluded = 0
        status_counts = defaultdict(int)

        # 옵션마스터 로드 (한 번만)
        from services.option_matcher import check_option_registration
        opt_list = db.query_option_master_as_list() or []
        _header_vals = {'standard_name', 'product_name', '품목명', 'original_name', '원문명'}
        opt_list = [o for o in opt_list
                    if str(o.get('원문명', '')).strip()
                    and str(o.get('품목명', '')).strip().lower() not in _header_vals]

        for ch in channels:
            client = mgr.get_client(ch)
            if not client:
                channel_status[ch] = '클라이언트 없음'
                continue
            if not client.is_ready:
                try:
                    client.refresh_token(db)
                    if not client.is_ready:
                        channel_status[ch] = '인증 미완료'
                        continue
                except Exception:
                    channel_status[ch] = '인증 미완료'
                    continue

            try:
                orders = client.fetch_orders(date_from, date_to,
                                             status_filter='invoice_target')
                if not orders:
                    channel_status[ch] = '0건'
                    continue

                # 취소/환불 필터링
                filtered = [o for o in orders
                            if not any(ex in str(o.get('order_status', '')).upper()
                                       for ex in ('취소', '환불', 'CANCEL', 'REFUND'))]
                # N배송 필터링 (스마트스토어/해미애찬)
                if ch in ('스마트스토어', '해미애찬'):
                    filtered = [o for o in filtered
                                if (o.get('fee_detail') or {}).get('delivery_type') != 'ARRIVAL_GUARANTEE']
                excluded = len(orders) - len(filtered)
                total_count += len(orders)
                total_excluded += excluded

                # 상태 분포 집계
                for o in orders:
                    status_counts[o.get('order_status', '(없음)')] += 1

                # 쿠팡: raw_data에서 sellerProductName 사용
                orders_for_check = filtered
                if ch == '쿠팡':
                    orders_for_check = []
                    for o in filtered:
                        raw = o.get('raw_data', {})
                        items = raw.get('orderItems', [{}])
                        item = items[0] if items else {}
                        orders_for_check.append({
                            'product_name': item.get('sellerProductName',
                                            item.get('vendorItemName', o.get('product_name', ''))),
                            'option_name': item.get('sellerProductItemName',
                                          o.get('option_name', '')),
                            'order_status': o.get('order_status', ''),
                        })

                # option_matcher로 매칭 검사
                chk = check_option_registration(orders_for_check, ch, opt_list)
                total_matched += chk['registered']
                total_unmatched.extend(chk['unregistered_items'])
                channel_status[ch] = f"{len(orders)}건 → 매칭 {chk['registered']}건, 미매칭 {chk['unregistered']}건"

            except Exception as e:
                logger.error(f'[InvoicePreview] {ch} 오류: {e}', exc_info=True)
                channel_status[ch] = f'오류: {e}'

        return jsonify({
            'total': total_count,
            'filtered': total_count - total_excluded,
            'excluded': total_excluded,
            'matched': total_matched,
            'unmatched_count': len(total_unmatched),
            'unmatched': total_unmatched[:30],
            'status_counts': dict(status_counts),
            'channel_status': channel_status,
        })

    except Exception as e:
        logger.error(f'[InvoicePreview] 오류: {e}', exc_info=True)
        return jsonify({'error': str(e)}), 500


# ── 쿠팡 로켓배송 정산 엑셀 업로드 ──

def _parse_coupang_billing_xlsx(file_bytes):
    """쿠팡 Supplier Hub Billing Statement xlsx 파싱.

    쿠팡 엑셀은 inline string 사용 + 인코딩 깨짐 → XML 직접 파싱.
    컬럼 고정 순서: 계산서번호(A) 거래처명(B) 작성일자(C) 지급일자(D)
    과세유형(E) 발주유형(F) 정산유형(G) 공급가액(H) VAT(I) 지급예정금액(J)
    세금계산서확정일(K) 1차지급일(L) 1차지급액(M) 2차지급액(N) 발주정보(O) 반출정보(P)
    """
    import zipfile
    import xml.etree.ElementTree as ET

    zf = zipfile.ZipFile(io.BytesIO(file_bytes))
    ns = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
    sheet_xml = zf.read('xl/worksheets/sheet1.xml').decode('utf-8')
    root = ET.fromstring(sheet_xml)

    # 컬럼 인덱스 매핑 (A=0, B=1, ...)
    def _col_idx(ref):
        col = ''.join(c for c in ref if c.isalpha())
        return ord(col) - ord('A')

    def _cell_value(cell_el):
        is_el = cell_el.find(f'{ns}is')
        if is_el is not None:
            t_el = is_el.find(f'{ns}t')
            return t_el.text if t_el is not None else ''
        v_el = cell_el.find(f'{ns}v')
        return v_el.text if v_el is not None else ''

    rows_data = []
    for row_el in root.findall(f'.//{ns}row'):
        cells = {}
        for c in row_el.findall(f'{ns}c'):
            ref = c.get('r', '')
            idx = _col_idx(ref)
            cells[idx] = _cell_value(c)
        rows_data.append(cells)

    return rows_data  # [{col_idx: value, ...}, ...]


@marketplace_bp.route('/rocket/upload', methods=['POST'])
@role_required('admin', 'general')
def upload_rocket_settlement():
    """쿠팡 로켓배송(Supplier Hub) Billing Statement 업로드 → api_settlements 저장.

    과세/면세 구분하여 작성일자 기준으로 일별 집계.
    """
    db = get_db()

    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'error': '파일을 선택하세요'}), 400

    try:
        file_bytes = f.read()
        rows = _parse_coupang_billing_xlsx(file_bytes)

        if len(rows) < 2:
            return jsonify({'error': '데이터가 없습니다'}), 400

        # 헤더(row 0) 스킵, 데이터(row 1~)
        # 고정 컬럼: C=작성일자(2), E=과세유형(4), H=공급가액(7), I=VAT(8), J=지급예정금액(9)
        from collections import defaultdict as _dd
        daily_agg = _dd(lambda: {
            'supply_taxable': 0, 'vat_taxable': 0,
            'supply_exempt': 0,
            'total_payment': 0, 'count': 0,
        })

        def _to_int(v):
            if not v:
                return 0
            try:
                return int(float(str(v).replace(',', '')))
            except (ValueError, TypeError):
                return 0

        for row in rows[1:]:  # 데이터 행
            date_str = (row.get(2) or '').strip()[:10]  # C: 작성일자
            if not date_str or len(date_str) < 8:
                continue

            tax_type = row.get(4, '')  # E: 과세유형
            supply = _to_int(row.get(7))  # H: 공급가액
            vat = _to_int(row.get(8))  # I: VAT
            payment = _to_int(row.get(9))  # J: 지급예정금액

            d = daily_agg[date_str]
            if '면세' in tax_type:
                d['supply_exempt'] += supply
            else:
                d['supply_taxable'] += supply
                d['vat_taxable'] += vat
            d['total_payment'] += payment
            d['count'] += 1

        settlements = []
        for date_str in sorted(daily_agg):
            d = daily_agg[date_str]
            gross = d['supply_taxable'] + d['supply_exempt']  # 공급가액 합계 = 매출
            settlements.append({
                'channel': '쿠팡로켓',
                'settlement_date': date_str,
                'settlement_id': f'rocket_{date_str}',
                'gross_sales': gross,
                'total_commission': 0,  # 로켓은 매입 구조 → 수수료 없음
                'shipping_fee_income': 0,
                'shipping_fee_cost': 0,
                'coupon_discount': 0,
                'point_discount': 0,
                'other_deductions': 0,
                'net_settlement': d['total_payment'],  # 지급예정금액 (공급가액+VAT)
                'fee_breakdown': {
                    'source': 'billing_statement',
                    'supply_taxable': d['supply_taxable'],
                    'vat_taxable': d['vat_taxable'],
                    'supply_exempt': d['supply_exempt'],
                    'total_payment': d['total_payment'],
                    'invoice_count': d['count'],
                    'filename': f.filename,
                },
            })

        if not settlements:
            return jsonify({'error': '유효한 정산 데이터가 없습니다'}), 400

        db.upsert_api_settlements_batch(settlements)
        total_count = sum(d['count'] for d in daily_agg.values())
        _log_action(f'쿠팡로켓 정산 업로드: {len(settlements)}일, {total_count}건')

        return jsonify({
            'success': True,
            'count': total_count,
            'date_range': f'{settlements[0]["settlement_date"]} ~ {settlements[-1]["settlement_date"]}',
        })

    except Exception as e:
        logger.error(f'[로켓정산] 업로드 오류: {e}', exc_info=True)
        return jsonify({'error': str(e)}), 500


# ── 광고비 수동 기입 ──

@marketplace_bp.route('/ad-cost', methods=['POST'])
@role_required('admin', 'general')
def save_ad_cost():
    """채널별 광고비 수동 기입 → api_settlements 저장.
    input_mode: 'daily' (일별, date 필요) / 'monthly' (월별, month 필요)
    """
    db = get_db()

    channel = request.form.get('channel', '').strip()
    input_mode = request.form.get('input_mode', 'daily').strip()
    save_mode = request.form.get('save_mode', 'replace').strip()
    cost = request.form.get('cost', '0').strip().replace(',', '')
    memo = request.form.get('memo', '').strip()

    if input_mode == 'monthly':
        month_str = request.form.get('month', '').strip()
        if not channel or not month_str:
            return jsonify({'error': '채널과 월을 입력하세요'}), 400
        date_str = f'{month_str}-01'
        settle_id = f'ad_cost_monthly_{channel}_{month_str}'
        label = month_str
    else:
        date_str = request.form.get('date', '').strip()
        if not channel or not date_str:
            return jsonify({'error': '채널과 날짜를 입력하세요'}), 400
        settle_id = f'ad_cost_{date_str}'
        label = date_str

    try:
        cost_int = int(float(cost))
    except (ValueError, TypeError):
        return jsonify({'error': '광고비 금액이 올바르지 않습니다'}), 400

    # 추가 모드: 기존 값에 합산
    if save_mode == 'add':
        existing = db.query_api_settlements(channel, date_str[:7])
        for s in existing:
            if s.get('settlement_id') == settle_id:
                cost_int += int(s.get('other_deductions') or 0)
                break

    settlement = {
        'channel': channel,
        'settlement_date': date_str,
        'settlement_id': settle_id,
        'gross_sales': 0,
        'total_commission': 0,
        'shipping_fee_income': 0,
        'shipping_fee_cost': 0,
        'coupon_discount': 0,
        'point_discount': 0,
        'other_deductions': cost_int,
        'net_settlement': -cost_int,
        'fee_breakdown': {
            'source': 'manual_input',
            'input_mode': input_mode,
            'memo': memo,
        },
    }

    db.upsert_api_settlements_batch([settlement])
    _log_action(f'광고비 기입: {channel} {label} {cost_int:,}원 ({input_mode})')

    return jsonify({
        'success': True,
        'channel': channel,
        'date': label,
        'cost': cost_int,
    })


# ── 광고비 일괄 업로드 (엑셀) ──

@marketplace_bp.route('/ad-cost/upload', methods=['POST'])
@role_required('admin', 'general')
def upload_ad_cost():
    """광고비 엑셀 업로드 → api_settlements 저장.

    엑셀 형식: 날짜 | 채널 | 광고비 | (메모)
    """
    import openpyxl
    db = get_db()

    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'error': '파일을 선택하세요'}), 400

    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True)
        ws = wb.active

        headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]

        col_map = {}
        for i, h in enumerate(headers):
            hl = h.replace(' ', '')
            if '날짜' in hl or '일자' in hl or 'date' in hl.lower():
                col_map['date'] = i
            elif '채널' in hl or '매체' in hl or '플랫폼' in hl:
                col_map['channel'] = i
            elif '광고비' in hl or '비용' in hl or '금액' in hl or 'cost' in hl.lower():
                col_map['cost'] = i
            elif '메모' in hl or '비고' in hl:
                col_map['memo'] = i

        if 'cost' not in col_map:
            wb.close()
            return jsonify({'error': '광고비/비용/금액 컬럼을 찾을 수 없습니다'}), 400

        settlements = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row = list(row)
            if not row or all(v is None for v in row):
                continue

            date_val = row[col_map.get('date', 0)]
            if not date_val:
                continue
            if hasattr(date_val, 'strftime'):
                date_str = date_val.strftime('%Y-%m-%d')
            else:
                date_str = str(date_val).strip()[:10]

            raw_channel = str(row[col_map['channel']]).strip() if 'channel' in col_map else '쿠팡'
            # 채널명 표준화 (엑셀 입력값 → DB 표준 채널명)
            ch_lower = raw_channel.replace(' ', '')
            if '쿠팡' in ch_lower:
                channel = '쿠팡'
            elif '네이버' in ch_lower or '스마트스토어' in ch_lower or 'naver' in ch_lower.lower():
                channel = '스마트스토어'
            elif '자사몰' in ch_lower or 'cafe24' in ch_lower.lower():
                channel = '자사몰'
            else:
                channel = raw_channel
            cost_val = row[col_map['cost']]
            try:
                cost_int = int(float(str(cost_val).replace(',', '')))
            except (ValueError, TypeError):
                continue
            if cost_int <= 0:
                continue

            memo = str(row[col_map.get('memo', -1)] or '') if 'memo' in col_map else ''

            settlements.append({
                'channel': channel,
                'settlement_date': date_str,
                'settlement_id': f'ad_cost_{date_str}',
                'gross_sales': 0,
                'total_commission': 0,
                'shipping_fee_income': 0,
                'shipping_fee_cost': 0,
                'coupon_discount': 0,
                'point_discount': 0,
                'other_deductions': cost_int,
                'net_settlement': -cost_int,
                'fee_breakdown': {
                    'source': 'excel_upload',
                    'memo': memo,
                    'filename': f.filename,
                },
            })

        wb.close()

        if not settlements:
            return jsonify({'error': '유효한 광고비 데이터가 없습니다'}), 400

        db.upsert_api_settlements_batch(settlements)
        _log_action(f'광고비 엑셀 업로드: {len(settlements)}건')

        return jsonify({
            'success': True,
            'count': len(settlements),
        })

    except Exception as e:
        logger.error(f'[광고비] 업로드 오류: {e}', exc_info=True)
        return jsonify({'error': str(e)}), 500


# ── 네이버 차감항목 (쿠폰차감/바우처적립) 월별 입력 ──

@marketplace_bp.route('/deductions', methods=['POST'])
@role_required('admin', 'general')
def save_deductions():
    """스마트스토어 월별 차감항목 저장 (쿠폰차감, 바우처적립)."""
    db = get_db()
    month = request.form.get('month', '').strip()  # YYYY-MM
    coupon = request.form.get('coupon_deduct', '0').strip().replace(',', '')
    voucher = request.form.get('voucher_deduct', '0').strip().replace(',', '')

    if not month or len(month) != 7:
        return jsonify({'error': '월(YYYY-MM)을 입력하세요'}), 400

    try:
        coupon_int = int(float(coupon))
        voucher_int = int(float(voucher))
    except (ValueError, TypeError):
        return jsonify({'error': '금액이 올바르지 않습니다'}), 400

    settlements = []
    if coupon_int:
        settlements.append({
            'channel': '스마트스토어',
            'settlement_date': f'{month}-01',
            'settlement_id': f'deduct_coupon_{month}',
            'gross_sales': 0, 'total_commission': 0,
            'shipping_fee_income': 0, 'shipping_fee_cost': 0,
            'coupon_discount': coupon_int, 'point_discount': 0,
            'other_deductions': coupon_int,
            'net_settlement': -coupon_int,
            'fee_breakdown': {'source': 'manual_deduction', 'type': 'coupon',
                              'label': '쿠폰차감'},
        })
    if voucher_int:
        settlements.append({
            'channel': '스마트스토어',
            'settlement_date': f'{month}-01',
            'settlement_id': f'deduct_voucher_{month}',
            'gross_sales': 0, 'total_commission': 0,
            'shipping_fee_income': 0, 'shipping_fee_cost': 0,
            'coupon_discount': 0, 'point_discount': voucher_int,
            'other_deductions': voucher_int,
            'net_settlement': -voucher_int,
            'fee_breakdown': {'source': 'manual_deduction', 'type': 'voucher',
                              'label': '바우처적립'},
        })

    if settlements:
        db.upsert_api_settlements_batch(settlements)
        _log_action(f'차감항목 입력: {month} 쿠폰={coupon_int:,} 바우처={voucher_int:,}')

    return jsonify({
        'success': True,
        'month': month,
        'coupon_deduct': coupon_int,
        'voucher_deduct': voucher_int,
    })


# ── 정산서 업로드 (채널별 엑셀 파싱) ──

@marketplace_bp.route('/settlement/upload', methods=['POST'])
@role_required('admin', 'general')
def upload_settlement():
    """채널별 정산서 엑셀 업로드 → api_settlements 저장. 복수 파일 지원.
    월별 집계 채널(쿠팡Wing, 11번가, 오아시스)은 복수 파일을 먼저 합산 후 upsert.
    """
    import io
    db = get_db()
    channel = request.form.get('channel', '').strip()
    files = request.files.getlist('file')

    if not channel:
        return jsonify({'error': '채널을 선택하세요'}), 400
    if not files or not files[0].filename:
        return jsonify({'error': '파일을 선택하세요'}), 400

    # 월별 집계 채널: 복수 파일 합산이 필요한 채널
    MERGE_CHANNELS = ('쿠팡', '11번가', '오아시스')

    results = []
    errors = []
    total_count = 0

    # ── 월별 집계 채널: 전체 파일 파싱 → 합산 → 한번에 upsert ──
    if channel in MERGE_CHANNELS:
        all_settlements = []
        seen_orders = set()  # 쿠팡Wing: 파일 간 주문 중복 제거용
        for f in files:
            fname = f.filename
            try:
                file_bytes = f.read()
                if channel == '쿠팡':
                    result = _parse_coupang_wing_settlement(file_bytes, seen_orders)
                elif channel == '11번가':
                    result = _parse_11st_settlement(file_bytes)
                elif channel == '오아시스':
                    result = _parse_oasis_settlement(file_bytes)

                if not result['settlements']:
                    errors.append(f'{fname}: 파싱된 데이터 없음')
                    continue
                all_settlements.extend(result['settlements'])
                results.append(f'{fname}: {result["summary"]}')
            except Exception as e:
                logger.error(f'정산서 업로드 오류 ({channel}, {fname}): {e}')
                errors.append(f'{fname}: {str(e)}')

        # settlement_id 기준 합산 (같은 월 데이터 병합)
        if all_settlements:
            merged = {}
            for s in all_settlements:
                sid = s['settlement_id']
                if sid not in merged:
                    merged[sid] = s.copy()
                    # fee_breakdown 딕셔너리 복사 (원본 변형 방지)
                    merged[sid]['fee_breakdown'] = dict(s.get('fee_breakdown') or {})
                else:
                    m = merged[sid]
                    m['gross_sales'] += s.get('gross_sales', 0)
                    m['total_commission'] += s.get('total_commission', 0)
                    m['coupon_discount'] += s.get('coupon_discount', 0)
                    m['other_deductions'] += s.get('other_deductions', 0)
                    m['net_settlement'] += s.get('net_settlement', 0)
                    m['shipping_fee_income'] += s.get('shipping_fee_income', 0)
                    m['shipping_fee_cost'] += s.get('shipping_fee_cost', 0)
                    m['point_discount'] += s.get('point_discount', 0)
                    # order_count 합산
                    fb = m.get('fee_breakdown') or {}
                    sfb = s.get('fee_breakdown') or {}
                    fb['order_count'] = fb.get('order_count', 0) + sfb.get('order_count', 0)
                    m['fee_breakdown'] = fb

            final = list(merged.values())
            db.upsert_api_settlements_batch(final)
            total_count = len(final)
            summary_parts = [f'{s["settlement_id"]}: {s["gross_sales"]:,}→{s["net_settlement"]:,}'
                             for s in final]
            _log_action(f'정산서 업로드: {channel} (파일{len(files)}개 합산) '
                        + ', '.join(summary_parts))

    # ── 기타 채널: 파일별 개별 upsert ──
    else:
        for f in files:
            fname = f.filename
            try:
                file_bytes = f.read()

                if channel in ('스마트스토어', '해미애찬'):
                    result = _parse_naver_settlement(file_bytes, channel)
                elif channel == '자사몰':
                    result = _parse_toss_settlement(file_bytes)
                elif channel in ('옥션', '지마켓'):
                    result = _parse_auction_settlement(file_bytes, channel)
                else:
                    errors.append(f'{fname}: 지원하지 않는 채널')
                    continue

                if not result['settlements']:
                    errors.append(f'{fname}: 파싱된 데이터 없음')
                    continue

                db.upsert_api_settlements_batch(result['settlements'])
                _log_action(f'정산서 업로드: {channel} {result["summary"]}')
                total_count += len(result['settlements'])
                results.append(f'{fname}: {result["summary"]}')

            except Exception as e:
                logger.error(f'정산서 업로드 오류 ({channel}, {fname}): {e}')
                errors.append(f'{fname}: {str(e)}')

    if not results and errors:
        return jsonify({'error': ' | '.join(errors)}), 400

    return jsonify({
        'success': True,
        'channel': channel,
        'count': total_count,
        'summary': ' | '.join(results),
        'errors': errors if errors else None,
    })


@marketplace_bp.route('/settlement/history')
@role_required('admin', 'general')
def settlement_history():
    """정산서 업로드 이력 조회 (최근 50건, 업로드일 내림차순)."""
    db = get_db()
    try:
        res = db.client.table('api_settlements').select('*') \
            .order('synced_at', desc=True).limit(50).execute()
        rows = res.data or []
    except Exception:
        rows = db.query_api_settlements(limit=50)
    # settlement_id에서 source 정보 추출
    history = []
    for r in rows:
        fb = r.get('fee_breakdown') or {}
        if isinstance(fb, str):
            import json
            try:
                fb = json.loads(fb)
            except Exception:
                fb = {}
        history.append({
            'channel': r.get('channel', ''),
            'settlement_date': r.get('settlement_date', ''),
            'settlement_id': r.get('settlement_id', ''),
            'gross_sales': r.get('gross_sales', 0),
            'total_commission': r.get('total_commission', 0),
            'net_settlement': r.get('net_settlement', 0),
            'source': fb.get('source', ''),
            'order_count': fb.get('order_count', 0),
            'updated_at': r.get('updated_at', ''),
        })
    return jsonify(history)


@marketplace_bp.route('/settlement/delete', methods=['POST'])
@role_required('admin')
def settlement_delete():
    """정산서 데이터 삭제. settlement_id 또는 channel+settlement_id 기준."""
    db = get_db()
    data = request.get_json(silent=True) or {}
    settlement_id = data.get('settlement_id', '').strip()
    channel = data.get('channel', '').strip()

    if not settlement_id and not channel:
        return jsonify({'error': '삭제할 항목을 지정하세요'}), 400

    try:
        q = db.client.table('api_settlements').delete()
        if settlement_id:
            q = q.eq('settlement_id', settlement_id)
        if channel:
            q = q.eq('channel', channel)
        res = q.execute()
        deleted = len(res.data) if res.data else 0
        _log_action(f'정산서 삭제: ch={channel} sid={settlement_id} → {deleted}건')
        return jsonify({'success': True, 'deleted': deleted})
    except Exception as e:
        logger.error(f'정산서 삭제 오류: {e}')
        return jsonify({'error': str(e)}), 500


def _parse_naver_settlement(file_bytes, channel):
    """네이버 항목별정산 (SellerDailySettle.xlsx) 파싱.
    Row 1 헤더: 정산예정일,정산완료일,정산금액,...,정산기준금액[6],수수료합계[7],혜택정산[8],...
    """
    import openpyxl, io
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb[wb.sheetnames[0]]

    settlements = []
    total_sales = 0
    total_settle = 0
    month_set = set()

    for r in range(2, ws.max_row + 1):
        date_str = str(ws.cell(r, 1).value or '').strip()
        if not date_str or len(date_str) < 8:
            continue
        # 2026.03.03 → 2026-03-03
        date_str = date_str.replace('.', '-')
        if len(date_str) == 10:
            pass
        else:
            continue

        settle_amt = int(ws.cell(r, 3).value or 0)       # 정산금액
        base_amt = int(ws.cell(r, 6).value or 0)          # 정산기준금액
        fee_total = int(ws.cell(r, 7).value or 0)         # 수수료합계 (음수)
        benefit = int(ws.cell(r, 8).value or 0)            # 혜택정산 (음수)
        deduct = int(ws.cell(r, 9).value or 0)             # 일별 공제/환급

        settlements.append({
            'channel': channel,
            'settlement_date': date_str,
            'settlement_id': f'nsettle_{date_str}',
            'gross_sales': base_amt,
            'total_commission': abs(fee_total),
            'shipping_fee_income': 0,
            'shipping_fee_cost': 0,
            'coupon_discount': abs(benefit),
            'point_discount': 0,
            'other_deductions': abs(deduct),
            'net_settlement': settle_amt,
            'fee_breakdown': {
                'source': 'naver_settlement',
                'raw_fee': fee_total,
                'raw_benefit': benefit,
                'raw_deduct': deduct,
            },
        })
        total_sales += base_amt
        total_settle += settle_amt
        month_set.add(date_str[:7])

    wb.close()
    months = ','.join(sorted(month_set))
    return {
        'settlements': settlements,
        'summary': f'{months} {len(settlements)}일, 매출 {total_sales:,}원 → 정산 {total_settle:,}원',
    }


def _parse_coupang_wing_settlement(file_bytes, seen_orders=None):
    """쿠팡Wing 정산내역 (MSF_PAYMENT_REVENUE_DETAIL.xlsx) 파싱.
    Row 1 헤더: 주문번호[1],...,판매액[10],판매자할인쿠폰[11],...,판매수수료[14],...,정산금액[18],...,정산예정일[25]
    값이 문자열이므로 int 변환 필요. 주문일(결제일) 기준 월별 집계.
    seen_orders: 복수 파일 업로드 시 주문 중복 방지용 set (주문번호+옵션ID).
    """
    import openpyxl, io
    from datetime import datetime as _dt
    from collections import defaultdict

    if seen_orders is None:
        seen_orders = set()

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb[wb.sheetnames[0]]

    def to_int(v):
        if v is None or v == '':
            return 0
        try:
            return int(float(str(v).replace(',', '')))
        except (ValueError, TypeError):
            return 0

    # 헤더에서 결제완료일/주문일 컬럼 찾기 (fallback: 정산예정일[25])
    # 쿠팡Wing 실제 헤더: [20]결제완료일, [25]정산예정일
    date_col = 25
    for c in range(1, min(ws.max_column + 1, 30)):
        h = str(ws.cell(1, c).value or '').strip()
        if '결제완료일' in h:
            date_col = c
            break
        if '주문' in h and '일' in h:
            date_col = c
            break
        if '결제' in h and '일' in h and '정산' not in h:
            date_col = c
            break

    # 월별 집계 (주문일/결제일 기준) — 주문번호+옵션ID로 중복 제거
    monthly = defaultdict(lambda: {'sales': 0, 'fee': 0, 'coupon': 0, 'settle': 0, 'cnt': 0})
    dup_count = 0
    for r in range(2, ws.max_row + 1):
        sales = to_int(ws.cell(r, 10).value)  # 판매액
        if not sales:
            continue

        # 중복 제거: 주문번호[1] + 옵션ID[5]
        order_no = str(ws.cell(r, 1).value or '').strip()
        option_id = str(ws.cell(r, 5).value or '').strip()
        dedup_key = (order_no, option_id)
        if dedup_key in seen_orders:
            dup_count += 1
            continue
        seen_orders.add(dedup_key)

        coupon = to_int(ws.cell(r, 11).value)     # 판매자 할인쿠폰(A+B)
        fee = to_int(ws.cell(r, 14).value)         # 판매수수료
        settle = to_int(ws.cell(r, 18).value)      # 정산금액
        date_val = ws.cell(r, date_col).value
        if isinstance(date_val, _dt):
            order_date = date_val.strftime('%Y-%m-%d')
        else:
            order_date = str(date_val or '')[:10]

        month_key = order_date[:7] if len(order_date) >= 7 else 'unknown'
        m = monthly[month_key]
        m['sales'] += sales
        m['fee'] += fee
        m['coupon'] += coupon
        m['settle'] += settle
        m['cnt'] += 1

    wb.close()

    settlements = []
    summaries = []
    for mk in sorted(monthly):
        m = monthly[mk]
        if mk == 'unknown':
            continue
        settlements.append({
            'channel': '쿠팡',
            'settlement_date': f'{mk}-01',
            'settlement_id': f'wsettle_{mk}',
            'gross_sales': m['sales'],
            'total_commission': m['fee'],
            'shipping_fee_income': 0,
            'shipping_fee_cost': 0,
            'coupon_discount': m['coupon'],
            'point_discount': 0,
            'other_deductions': 0,
            'net_settlement': m['settle'],
            'fee_breakdown': {
                'source': 'coupang_wing_settlement',
                'order_count': m['cnt'],
                'dup_skipped': dup_count,
            },
        })
        summaries.append(f'{mk}: {m["sales"]:,}원→{m["settle"]:,}원')
    if dup_count:
        summaries.append(f'(중복 {dup_count}건 제외)')

    return {
        'settlements': settlements,
        'summary': ', '.join(summaries),
    }


def _parse_11st_settlement(file_bytes):
    """11번가 정산확정건 (.xls) 파싱.
    헤더 Row5(0-indexed): ...,정산금액[17],판매금액합계[18],...,공제금액합계[20],...
    서비스이용료(상품)[39],제휴마케팅(상품)[40],할인쿠폰이용료[43],후불광고비[53]
    주문일/결제일 기준 월별 집계.
    """
    import xlrd, io, re
    from collections import defaultdict

    wb = xlrd.open_workbook(file_contents=file_bytes)
    ws = wb.sheet_by_index(0)

    # 제목에서 기간 추출 (fallback용)
    title = str(ws.cell_value(1, 0)) if ws.nrows > 1 else ''
    match = re.search(r'(\d{4})/(\d{2})/\d{2}~\d{4}/(\d{2})/\d{2}', title)
    fallback_month = f'{match.group(1)}-{match.group(2)}' if match else 'unknown'

    ncols = ws.ncols

    def to_int(v):
        try:
            return int(float(v))
        except (ValueError, TypeError):
            return 0

    def _safe(r, c):
        return to_int(ws.cell_value(r, c)) if c < ncols else 0

    # 헤더(Row 5, 0-indexed)에서 주문일/결제일 컬럼 찾기
    date_col = None
    if ws.nrows > 5:
        for c in range(ncols):
            h = str(ws.cell_value(5, c)).strip()
            if '주문일' in h:
                date_col = c
                break
            if '결제일' in h or '결제확인' in h:
                date_col = c
                break

    monthly = defaultdict(lambda: {
        'sales': 0, 'fee': 0, 'coupon': 0, 'ad': 0, 'settle': 0, 'cnt': 0,
    })

    for r in range(6, ws.nrows):
        no = ws.cell_value(r, 0)
        if not no:
            continue

        try:
            # 날짜 결정: 주문일/결제일 컬럼 → fallback: 제목 월
            if date_col is not None and date_col < ncols:
                if ws.cell_type(r, date_col) == xlrd.XL_CELL_DATE:
                    dt = xlrd.xldate_as_tuple(ws.cell_value(r, date_col), wb.datemode)
                    raw_date = f'{dt[0]:04d}-{dt[1]:02d}-{dt[2]:02d}'
                else:
                    raw_date = str(ws.cell_value(r, date_col)).strip()[:10]
                    raw_date = raw_date.replace('/', '-')
                month_key = raw_date[:7] if len(raw_date) >= 7 else fallback_month
            else:
                month_key = fallback_month

            m = monthly[month_key]
            m['cnt'] += 1
            m['sales'] += _safe(r, 18)    # 판매금액합계
            m['settle'] += _safe(r, 17)    # 정산금액
            m['fee'] += (_safe(r, 39) + _safe(r, 40)
                         + _safe(r, 41) + _safe(r, 42))
            m['coupon'] += _safe(r, 43)    # 할인쿠폰이용료
            m['ad'] += _safe(r, 53)         # 후불광고비
        except Exception:
            continue  # 개별 행 오류 무시

    settlements = []
    summaries = []
    for mk in sorted(monthly):
        m = monthly[mk]
        if mk == 'unknown' or m['cnt'] == 0:
            continue
        settlements.append({
            'channel': '11번가',
            'settlement_date': f'{mk}-01',
            'settlement_id': f'11settle_{mk}',
            'gross_sales': m['sales'],
            'total_commission': m['fee'],
            'shipping_fee_income': 0,
            'shipping_fee_cost': 0,
            'coupon_discount': m['coupon'],
            'point_discount': 0,
            'other_deductions': m['ad'],
            'net_settlement': m['settle'],
            'fee_breakdown': {
                'source': '11st_settlement',
                'order_count': m['cnt'],
                'ad_cost': m['ad'],
            },
        })
        summaries.append(f'{mk} {m["cnt"]}건: {m["sales"]:,}원→{m["settle"]:,}원')

    return {
        'settlements': settlements,
        'summary': ', '.join(summaries) if summaries else f'{fallback_month} 데이터 없음',
    }


def _parse_toss_settlement(file_bytes):
    """토스페이 PG 정산내역 (정산내역_요약.xlsx) 파싱.
    Row 1 헤더: 정산계좌[1],정산액입금일[2],매출일[3],...,매출액(A)[6],PG수수료합(D)[9],당일정산액(E)[10]
    """
    import openpyxl, io

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb[wb.sheetnames[0]]

    settlements = []
    total_sales = 0
    total_settle = 0
    month_set = set()

    for r in range(2, ws.max_row + 1):
        sales_date = str(ws.cell(r, 3).value or '').strip()  # 매출일
        if not sales_date or len(sales_date) < 8:
            continue

        sales = int(float(ws.cell(r, 6).value or 0))    # 매출액
        fee = int(float(ws.cell(r, 9).value or 0))      # PG수수료 합 (B+C)
        settle = int(float(ws.cell(r, 10).value or 0))   # 당일 정산액

        settlements.append({
            'channel': '자사몰',
            'settlement_date': sales_date,
            'settlement_id': f'tsettle_{sales_date}',
            'gross_sales': sales,
            'total_commission': fee,
            'shipping_fee_income': 0,
            'shipping_fee_cost': 0,
            'coupon_discount': 0,
            'point_discount': 0,
            'other_deductions': 0,
            'net_settlement': settle,
            'fee_breakdown': {'source': 'toss_settlement'},
        })
        total_sales += sales
        total_settle += settle
        month_set.add(sales_date[:7])

    wb.close()
    months = ','.join(sorted(month_set))
    return {
        'settlements': settlements,
        'summary': f'{months} {len(settlements)}일, 매출 {total_sales:,}원 → 정산 {total_settle:,}원',
    }


def _parse_oasis_settlement(file_bytes):
    """오아시스 매출정산내역.xlsx 파싱.
    Row1=정산월, Row5 헤더: 주문번호[1],주문일[2],...,판매금액[8],배송비[9],합계금액[10]
    수수료 8% 고정 (판매+배송 합계 기준). 주문일 기준 월별 집계.
    """
    import openpyxl, io
    from datetime import datetime as _dt
    from collections import defaultdict

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb[wb.sheetnames[0]]

    # 메타: Row1 = 정산월 (fallback용)
    month_raw = str(ws.cell(1, 2).value or '').strip()
    fallback_month = month_raw[:7] if len(month_raw) >= 7 else 'unknown'

    # 헤더(Row 5)에서 주문일 컬럼 찾기 (기본: column 2)
    date_col = 2
    for c in range(1, min(ws.max_column + 1, 15)):
        h = str(ws.cell(5, c).value or '').strip()
        if '주문일' in h:
            date_col = c
            break

    monthly = defaultdict(lambda: {'sales': 0, 'ship': 0, 'cnt': 0})

    for r in range(6, ws.max_row + 1):
        sales = ws.cell(r, 8).value   # 판매금액
        ship = ws.cell(r, 9).value    # 배송비
        if not sales:
            continue

        date_val = ws.cell(r, date_col).value
        if isinstance(date_val, _dt):
            order_date = date_val.strftime('%Y-%m-%d')
        else:
            order_date = str(date_val or '').strip()[:10]
        month_key = order_date[:7] if len(order_date) >= 7 else fallback_month

        m = monthly[month_key]
        m['sales'] += int(float(sales))
        m['ship'] += int(float(ship or 0))
        m['cnt'] += 1

    wb.close()

    settlements = []
    summaries = []
    for mk in sorted(monthly):
        m = monthly[mk]
        if mk == 'unknown' or m['cnt'] == 0:
            continue
        gross = m['sales'] + m['ship']
        commission = int(gross * 0.08)
        net = gross - commission
        settlements.append({
            'channel': '오아시스',
            'settlement_date': f'{mk}-01',
            'settlement_id': f'osettle_{mk}',
            'gross_sales': gross,
            'total_commission': commission,
            'shipping_fee_income': m['ship'],
            'shipping_fee_cost': 0,
            'coupon_discount': 0,
            'point_discount': 0,
            'other_deductions': 0,
            'net_settlement': net,
            'fee_breakdown': {
                'source': 'oasis_settlement',
                'sales_only': m['sales'],
                'shipping': m['ship'],
                'commission_rate': 0.08,
                'order_count': m['cnt'],
            },
        })
        summaries.append(f'{mk} {m["cnt"]}건: {gross:,}원-{commission:,}원={net:,}원')

    return {
        'settlements': settlements,
        'summary': ', '.join(summaries) if summaries else f'{fallback_month} 데이터 없음',
    }


def _parse_auction_settlement(file_bytes, channel):
    """옥션/지마켓 정산서 파싱 — 헤더 기반 컬럼 자동 감지.

    옥션: 상품판매(47컬럼), 배송비(14컬럼)
    지마켓: 상품판매(51컬럼), 배송비(27컬럼)
    헤더명으로 결제금액·서비스이용료·최종정산금·배송비 등 컬럼 자동 매핑.
    주문일(옥션) / 체결일(지마켓) 기준 월별 집계.
    """
    import io, xlrd
    from collections import defaultdict

    wb = xlrd.open_workbook(file_contents=file_bytes)
    ws = wb.sheet_by_index(0)

    if ws.nrows < 2:
        raise ValueError('데이터가 없습니다')

    prefix = 'auction_' if channel == '옥션' else 'gmarket_'
    ncols = ws.ncols

    # ── 헤더(Row 0) 읽기 ──
    header_map = {}
    for c in range(ncols):
        h = str(ws.cell_value(0, c)).strip()
        header_map[h] = c

    def _find_col(*keywords):
        """정확 매칭 → substring 매칭 순서로 컬럼 인덱스 반환."""
        for kw in keywords:
            if kw in header_map:
                return header_map[kw]
        for kw in keywords:
            for h, c in header_map.items():
                if kw in h:
                    return c
        return None

    def _to_int(r, c):
        """셀 값을 정수로 변환 (콤마 문자열·숫자 모두 처리)."""
        if c is None or c >= ncols:
            return 0
        v = ws.cell_value(r, c)
        try:
            return int(float(str(v).replace(',', '') or 0))
        except (ValueError, TypeError):
            return 0

    # ── 파일 유형 판별 (헤더 기반) ──
    settle_ship_col = _find_col('배송비정산금액', '배송비 정산액')
    is_shipping = settle_ship_col is not None

    # ── 금액 컬럼 감지 ──
    if is_shipping:
        gross_col = header_map.get('배송비')       # 정확히 '배송비'
        comm_col = _find_col('배송수수료', '서비스이용료')
        settle_col = settle_ship_col
    else:
        gross_col = _find_col('결제금액', '고객결제금')
        comm_col = _find_col('서비스이용료')
        settle_col = _find_col('최종정산금')

    # ── 날짜 컬럼 감지: 주문일(옥션) / 체결일(지마켓) ──
    date_col = _find_col('주문일', '체결일')
    if date_col is None:
        # fallback: 정산완료일
        date_col = _find_col('정산완료일')

    file_type = 'ship' if is_shipping else 'sales'
    monthly = defaultdict(lambda: {'sales': 0, 'comm': 0, 'settle': 0, 'cnt': 0})

    for r in range(1, ws.nrows):
        try:
            sales = _to_int(r, gross_col)
            comm = _to_int(r, comm_col)
            settle = _to_int(r, settle_col)

            # 날짜 처리
            raw_date = ''
            if date_col is not None and date_col < ncols:
                if ws.cell_type(r, date_col) == xlrd.XL_CELL_DATE:
                    dt = xlrd.xldate_as_tuple(ws.cell_value(r, date_col), wb.datemode)
                    raw_date = f'{dt[0]:04d}-{dt[1]:02d}-{dt[2]:02d}'
                else:
                    raw_date = str(ws.cell_value(r, date_col)).strip()[:10]
                    raw_date = raw_date.replace('/', '-')

            month_key = raw_date[:7] if len(raw_date) >= 7 else 'unknown'
            m = monthly[month_key]
            m['sales'] += sales
            m['comm'] += comm
            m['settle'] += settle
            m['cnt'] += 1
        except Exception:
            continue

    settlements = []
    summaries = []
    for mk in sorted(monthly):
        m = monthly[mk]
        if mk == 'unknown' or m['cnt'] == 0:
            continue
        settle_id = f'{prefix}{file_type}_{mk}'
        settlements.append({
            'channel': channel,
            'settlement_date': f'{mk}-01',
            'settlement_id': settle_id,
            'gross_sales': m['sales'],
            'total_commission': abs(m['comm']),
            'coupon_discount': 0,
            'other_deductions': 0,
            'net_settlement': m['settle'],
            'fee_breakdown': {
                'source': f'{channel.lower()}_settlement',
                'file_type': file_type,
                'order_count': m['cnt'],
            },
        })
        summaries.append(f'{mk} {m["cnt"]}건: {m["sales"]:,}원→{m["settle"]:,}원')

    if not settlements:
        raise ValueError('날짜를 파싱할 수 없습니다')

    return {
        'settlements': settlements,
        'summary': f'{channel} {file_type} ' + ', '.join(summaries),
    }


# ── 온라인 플랫폼 정산 대시보드 ──

@marketplace_bp.route('/sales')
@role_required('admin', 'general')
def sales():
    """온라인 플랫폼 정산 대시보드."""
    import calendar
    from datetime import date

    month = request.args.get('month', '')
    if not month:
        t = date.today()
        month = t.strftime('%Y-%m')

    year, mon = int(month[:4]), int(month[5:7])
    last_day = calendar.monthrange(year, mon)[1]
    date_from = f'{month}-01'
    date_to = f'{month}-{last_day:02d}'

    return render_template('marketplace/sales.html',
                           month=month,
                           date_from=date_from,
                           date_to=date_to)


@marketplace_bp.route('/api/sales-data')
@role_required('admin', 'general')
def sales_data():
    """API 매출 집계 데이터 (AJAX)."""
    import calendar
    from datetime import date

    db = get_db()
    month = request.args.get('month', '')
    if not month:
        t = date.today()
        month = t.strftime('%Y-%m')

    year, mon = int(month[:4]), int(month[5:7])
    last_day = calendar.monthrange(year, mon)[1]
    date_from = f'{month}-01'
    date_to = f'{month}-{last_day:02d}'

    # 1) api_orders: 주문 데이터 (네이버/쿠팡/자사몰) — 필요 컬럼만 조회 (메모리 절약)
    _sales_cols = "channel,order_date,order_status,total_amount,commission,settlement_amount,shipping_fee,fee_detail"
    orders = db.query_api_orders(date_from=date_from, date_to=date_to, columns=_sales_cols)

    # 2) api_settlements: 로켓정산 + 광고비
    settlements = db.query_api_settlements(date_from=date_from, date_to=date_to)

    # ── 채널별 집계 ──
    # channel_key: (채널명, 서브타입)
    ch_daily = defaultdict(lambda: defaultdict(lambda: {
        'sales': 0, 'commission': 0, 'settlement': 0,
        'shipping': 0, 'orders_count': 0,
    }))

    # 취소/환불 상태 제외 함수
    def _is_cancelled(order_status, channel):
        """취소/환불/반품 주문 여부 판단."""
        s = (order_status or '').upper()
        if not s:
            return False
        # 네이버: CANCEL*, RETURN*, EXCHANGE*
        if 'CANCEL' in s or 'RETURN' in s or 'EXCHANGE' in s:
            return True
        # Cafe24: C로 시작하는 상태코드 (C00, C40 등)
        if channel == '자사몰' and s.startswith('C'):
            return True
        return False

    for o in orders:
        if _is_cancelled(o.get('order_status', ''), o.get('channel', '')):
            continue

        ch = o.get('channel', '')
        dt = o.get('order_date', '')[:10]
        fee = o.get('fee_detail') or {}

        # 네이버: delivery_type으로 N배송/일반 구분 (통합 후 합산)
        if ch in ('스마트스토어', '해미애찬'):
            d_type = fee.get('delivery_type', '')
            if d_type == 'ARRIVAL_GUARANTEE':
                key = f'{ch}_N배송'
            else:
                key = f'{ch}_일반'
        else:
            key = ch

        d = ch_daily[key][dt]
        d['sales'] += int(o.get('total_amount') or 0)
        d['commission'] += int(o.get('commission') or 0)
        d['settlement'] += int(o.get('settlement_amount') or 0)
        d['shipping'] += int(o.get('shipping_fee') or 0)
        d['orders_count'] += 1

    # 정산 데이터 반영 (로켓 엑셀 + 쿠팡 revenue_fees)
    for s in settlements:
        ch = s.get('channel', '')
        sid = s.get('settlement_id', '')
        dt = s.get('settlement_date', '')[:10]

        if sid.startswith('ad_cost_'):
            continue  # 광고비는 아래에서 별도 처리

        if ch == '쿠팡로켓':
            # 로켓 Billing Statement 업로드 데이터
            d = ch_daily['쿠팡로켓'][dt]
            d['sales'] += int(s.get('gross_sales') or 0)
            d['commission'] += int(s.get('total_commission') or 0)
            d['settlement'] += int(s.get('net_settlement') or 0)
            d['shipping'] += int(s.get('shipping_fee_income') or 0) - int(s.get('shipping_fee_cost') or 0)

        elif ch == '쿠팡' and sid.startswith('revenue_'):
            # 쿠팡 Wing revenue-history: 수수료/정산 데이터 덮어쓰기
            # orders에서 매출/건수는 이미 있으므로, 수수료/정산만 반영
            d = ch_daily['쿠팡'][dt]
            d['commission'] = int(s.get('total_commission') or 0)
            d['settlement'] = int(s.get('net_settlement') or 0)
            d['shipping'] = int(s.get('shipping_fee_income') or 0) - int(s.get('shipping_fee_cost') or 0)

    # 쿠팡 Wing 수수료 추정 (revenue-history 미도착분 → 10.8% 추정)
    # revenue-history 실데이터가 들어오면 위에서 덮어쓰므로, 여기서는 0인 날만 추정
    for dt, d in ch_daily.get('쿠팡', {}).items():
        if d['sales'] > 0 and d['commission'] == 0:
            d['commission'] = int(d['sales'] * 0.108)
            d['settlement'] = d['sales'] - d['commission']

    # 자사몰 PG 수수료 추정 (API 미제공 → 3.3% 추정)
    for dt, d in ch_daily.get('자사몰', {}).items():
        if d['sales'] > 0 and d['commission'] == 0:
            d['commission'] = int(d['sales'] * 0.033)
            d['settlement'] = d['sales'] - d['commission']

    # ── 광고비 집계 ──
    ad_daily = defaultdict(lambda: defaultdict(int))  # {channel: {date: cost}}
    ad_items = []  # 개별 광고비 내역

    # 차감항목 집계 (쿠폰차감/바우처적립)
    deductions = {'coupon': 0, 'voucher': 0}  # 월 합계

    for s in settlements:
        sid = s.get('settlement_id', '')
        if sid.startswith('ad_cost_'):
            ch = s.get('channel', '')
            dt = s.get('settlement_date', '')[:10]
            cost = int(s.get('other_deductions') or 0)
            fb = s.get('fee_breakdown') or {}
            ad_daily[ch][dt] += cost
            ad_items.append({
                'channel': ch,
                'date': dt,
                'cost': cost,
                'memo': fb.get('memo', ''),
                'source': fb.get('source', ''),
            })
        elif sid.startswith('deduct_coupon_'):
            deductions['coupon'] += int(s.get('other_deductions') or 0)
        elif sid.startswith('deduct_voucher_'):
            deductions['voucher'] += int(s.get('other_deductions') or 0)

    # ── 정산서 업로드 데이터 처리 ──
    # settlement_id prefix로 정산서 데이터 구분
    settle_data = {}  # {channel: {sales, commission, coupon, settle, source}}
    for s in settlements:
        sid = s.get('settlement_id', '')
        ch = s.get('channel', '')

        if sid.startswith('nsettle_'):
            # 네이버 정산서 (스마트스토어 or 해미애찬)
            if ch not in settle_data:
                settle_data[ch] = {'sales': 0, 'commission': 0, 'coupon': 0,
                                   'settle': 0, 'ad_cost': 0, 'source': 'settlement'}
            sd = settle_data[ch]
            sd['sales'] += int(s.get('gross_sales') or 0)
            sd['commission'] += int(s.get('total_commission') or 0)
            sd['coupon'] += int(s.get('coupon_discount') or 0)
            sd['settle'] += int(s.get('net_settlement') or 0)

        elif sid.startswith('wsettle_'):
            # 쿠팡Wing 정산서
            if '쿠팡' not in settle_data:
                settle_data['쿠팡'] = {'sales': 0, 'commission': 0, 'coupon': 0,
                                      'settle': 0, 'ad_cost': 0, 'source': 'settlement'}
            sd = settle_data['쿠팡']
            sd['sales'] += int(s.get('gross_sales') or 0)
            sd['commission'] += int(s.get('total_commission') or 0)
            sd['coupon'] += int(s.get('coupon_discount') or 0)
            sd['settle'] += int(s.get('net_settlement') or 0)

        elif sid.startswith('rocket_'):
            # 쿠팡로켓 Billing Statement
            if '쿠팡로켓' not in settle_data:
                settle_data['쿠팡로켓'] = {'sales': 0, 'commission': 0, 'coupon': 0,
                                          'settle': 0, 'ad_cost': 0, 'source': 'settlement'}
            sd = settle_data['쿠팡로켓']
            sd['sales'] += int(s.get('gross_sales') or 0)
            sd['commission'] += int(s.get('total_commission') or 0)
            sd['settle'] += int(s.get('net_settlement') or 0)

        elif sid.startswith('11settle_'):
            if '11번가' not in settle_data:
                settle_data['11번가'] = {'sales': 0, 'commission': 0, 'coupon': 0,
                                        'settle': 0, 'ad_cost': 0, 'source': 'settlement'}
            sd = settle_data['11번가']
            sd['sales'] += int(s.get('gross_sales') or 0)
            sd['commission'] += int(s.get('total_commission') or 0)
            sd['coupon'] += int(s.get('coupon_discount') or 0)
            sd['settle'] += int(s.get('net_settlement') or 0)
            sd['ad_cost'] += int(s.get('other_deductions') or 0)  # 후불광고비

        elif sid.startswith('tsettle_'):
            if '자사몰' not in settle_data:
                settle_data['자사몰'] = {'sales': 0, 'commission': 0, 'coupon': 0,
                                        'settle': 0, 'ad_cost': 0, 'source': 'settlement'}
            sd = settle_data['자사몰']
            sd['sales'] += int(s.get('gross_sales') or 0)
            sd['commission'] += int(s.get('total_commission') or 0)
            sd['settle'] += int(s.get('net_settlement') or 0)

        elif sid.startswith('osettle_'):
            if '오아시스' not in settle_data:
                settle_data['오아시스'] = {'sales': 0, 'commission': 0, 'coupon': 0,
                                          'settle': 0, 'ad_cost': 0, 'source': 'settlement'}
            sd = settle_data['오아시스']
            sd['sales'] += int(s.get('gross_sales') or 0)
            sd['commission'] += int(s.get('total_commission') or 0)
            sd['settle'] += int(s.get('net_settlement') or 0)

        elif sid.startswith('auction_'):
            if '옥션' not in settle_data:
                settle_data['옥션'] = {'sales': 0, 'commission': 0, 'coupon': 0,
                                       'settle': 0, 'ad_cost': 0, 'source': 'settlement'}
            sd = settle_data['옥션']
            sd['sales'] += int(s.get('gross_sales') or 0)
            sd['commission'] += int(s.get('total_commission') or 0)
            sd['settle'] += int(s.get('net_settlement') or 0)

        elif sid.startswith('gmarket_'):
            if '지마켓' not in settle_data:
                settle_data['지마켓'] = {'sales': 0, 'commission': 0, 'coupon': 0,
                                         'settle': 0, 'ad_cost': 0, 'source': 'settlement'}
            sd = settle_data['지마켓']
            sd['sales'] += int(s.get('gross_sales') or 0)
            sd['commission'] += int(s.get('total_commission') or 0)
            sd['settle'] += int(s.get('net_settlement') or 0)

    # ── 채널별 월간 합계 ──
    channel_order = [
        '스마트스토어', '해미애찬',
        '쿠팡', '쿠팡로켓',
        '11번가', '자사몰', '오아시스',
        '옥션', '지마켓',
    ]
    channel_labels = {
        '스마트스토어': '스마트스토어(배마마)',
        '해미애찬': '스마트스토어(해미애찬)',
        '쿠팡': '쿠팡(Wing)',
        '쿠팡로켓': '쿠팡(로켓)',
        '11번가': '11번가',
        '자사몰': '자사몰(Cafe24)',
        '오아시스': '오아시스',
        '옥션': '옥션',
        '지마켓': '지마켓',
    }

    # ── API 기반 네이버 N배송+일반 → 스토어별 통합 ──
    for store in ('스마트스토어', '해미애찬'):
        if f'{store}_N배송' in ch_daily or f'{store}_일반' in ch_daily:
            merged = defaultdict(lambda: {'sales': 0, 'commission': 0, 'settlement': 0,
                                          'shipping': 0, 'orders_count': 0})
            for sub_key in (f'{store}_N배송', f'{store}_일반'):
                for dt, dd in ch_daily.get(sub_key, {}).items():
                    m = merged[dt]
                    for k in ('sales', 'commission', 'settlement', 'shipping', 'orders_count'):
                        m[k] += dd[k]
            ch_daily[store] = dict(merged)
            # 서브키 제거 (daily_summary에서 중복 집계 방지)
            ch_daily.pop(f'{store}_N배송', None)
            ch_daily.pop(f'{store}_일반', None)

    # ── 채널별 매출 합계 (광고비 비율 배분용) ──
    ch_sales_totals = {}
    for key in channel_order:
        daily = ch_daily.get(key, {})
        ch_sales_totals[key] = sum(dd['sales'] for dd in daily.values())

    # 쿠팡 광고비: Wing + 로켓 매출 비율로 배분
    coupang_ad_total = sum(ad_daily.get('쿠팡', {}).values())
    coupang_wing_sales = ch_sales_totals.get('쿠팡', 0)
    coupang_rocket_sales = ch_sales_totals.get('쿠팡로켓', 0)
    # 정산서 매출이 있으면 그 값 사용 (API보다 정확)
    if settle_data.get('쿠팡'):
        coupang_wing_sales = settle_data['쿠팡']['sales']
    coupang_total_sales = coupang_wing_sales + coupang_rocket_sales

    if coupang_total_sales > 0 and coupang_ad_total > 0:
        coupang_wing_ad = round(coupang_ad_total * coupang_wing_sales / coupang_total_sales)
        coupang_rocket_ad = coupang_ad_total - coupang_wing_ad
    else:
        coupang_wing_ad = coupang_ad_total
        coupang_rocket_ad = 0

    # 네이버 광고비: 배마마 + 해미애찬 매출 비율로 배분
    naver_ad_total = (sum(ad_daily.get('스마트스토어', {}).values())
                      + sum(ad_daily.get('해미애찬', {}).values()))
    naver_bm_sales = ch_sales_totals.get('스마트스토어', 0)
    naver_hm_sales = ch_sales_totals.get('해미애찬', 0)
    # 정산서 매출이 있으면 그 값 사용
    if settle_data.get('스마트스토어'):
        naver_bm_sales = settle_data['스마트스토어']['sales']
    if settle_data.get('해미애찬'):
        naver_hm_sales = settle_data['해미애찬']['sales']
    naver_total_sales = naver_bm_sales + naver_hm_sales

    if naver_total_sales > 0 and naver_ad_total > 0:
        naver_bm_ad = round(naver_ad_total * naver_bm_sales / naver_total_sales)
        naver_hm_ad = naver_ad_total - naver_bm_ad
    else:
        naver_bm_ad = naver_ad_total
        naver_hm_ad = 0

    # 채널별 광고비 배분
    ad_cost_map = {
        '스마트스토어': naver_bm_ad,
        '해미애찬': naver_hm_ad,
        '쿠팡': coupang_wing_ad,
        '쿠팡로켓': coupang_rocket_ad,
        '11번가': 0,
        '자사몰': 0,
        '오아시스': 0,
    }

    # 네이버 차감항목 (쿠폰차감+바우처적립): 배마마/해미애찬 매출 비율 배분
    total_deduct = deductions['coupon'] + deductions['voucher']
    if naver_total_sales > 0 and total_deduct != 0:
        deduct_bm = round(total_deduct * naver_bm_sales / naver_total_sales)
        deduct_hm = total_deduct - deduct_bm
    else:
        deduct_bm = total_deduct
        deduct_hm = 0
    deduct_map = {
        '스마트스토어': deduct_bm,
        '해미애찬': deduct_hm,
        '쿠팡': 0, '쿠팡로켓': 0, '11번가': 0, '자사몰': 0, '오아시스': 0,
    }

    channel_summary = []
    total = {'sales': 0, 'commission': 0, 'settlement': 0, 'shipping': 0,
             'ad_cost': 0, 'deductions': 0, 'net': 0, 'orders_count': 0}

    for key in channel_order:
        # 정산서 데이터 우선, 없으면 API 추정
        sd = settle_data.get(key)
        daily = ch_daily.get(key, {})

        if sd:
            # 정산서 기준
            s_sales = sd['sales']
            s_commission = sd['commission']
            s_settlement = sd['settle']
            s_coupon = sd['coupon']
            s_shipping = 0
            s_orders = sum(dd['orders_count'] for dd in daily.values()) if daily else 0
            source = 'settlement'
            # 정산서에 포함된 광고비 (11번가 후불광고비 등)
            settle_ad = sd.get('ad_cost', 0)
        elif daily:
            # API 추정
            s_sales = sum(dd['sales'] for dd in daily.values())
            s_commission = sum(dd['commission'] for dd in daily.values())
            s_settlement = sum(dd['settlement'] for dd in daily.values())
            s_coupon = 0
            s_shipping = sum(dd['shipping'] for dd in daily.values())
            s_orders = sum(dd['orders_count'] for dd in daily.values())
            source = 'api_estimate'
            settle_ad = 0
        else:
            continue

        ad_cost = ad_cost_map.get(key, 0) + settle_ad
        deduct = deduct_map.get(key, 0)
        # 정산서의 쿠폰차감은 이미 정산금액에 반영됨 → deduct에 포함하지 않음
        # 단, 수동 입력 차감(deduct_map)은 API 추정 모드에서만 의미 있음
        if source == 'settlement':
            deduct = 0  # 정산서는 이미 차감 반영됨

        net = s_settlement - ad_cost - deduct
        comm_rate = (s_commission / s_sales * 100) if s_sales else 0
        profit_rate = (net / s_sales * 100) if s_sales else 0

        channel_summary.append({
            'key': key,
            'label': channel_labels.get(key, key),
            'sales': s_sales,
            'commission': s_commission,
            'commission_rate': round(comm_rate, 1),
            'settlement': s_settlement,
            'shipping': s_shipping,
            'ad_cost': ad_cost,
            'deductions': deduct + s_coupon,  # 표시용: 쿠폰차감 포함
            'net': net,
            'orders_count': s_orders,
            'source': source,
            'profit_rate': round(profit_rate, 1),
        })

        total['sales'] += s_sales
        total['commission'] += s_commission
        total['settlement'] += s_settlement
        total['shipping'] += s_shipping
        total['ad_cost'] += ad_cost
        total['deductions'] += deduct + s_coupon
        total['net'] += net
        total['orders_count'] += s_orders

    total['commission_rate'] = round(
        (total['commission'] / total['sales'] * 100) if total['sales'] else 0, 1
    )
    total['profit_rate'] = round(
        (total['net'] / total['sales'] * 100) if total['sales'] else 0, 1
    )

    # ── 일별 합계 ──
    all_dates = set()
    for daily in ch_daily.values():
        all_dates.update(daily.keys())
    for ch_ads in ad_daily.values():
        all_dates.update(ch_ads.keys())

    daily_summary = []
    for dt in sorted(all_dates):
        ds = {'date': dt, 'sales': 0, 'commission': 0, 'settlement': 0, 'ad_cost': 0}
        for daily in ch_daily.values():
            dd = daily.get(dt, {})
            ds['sales'] += dd.get('sales', 0)
            ds['commission'] += dd.get('commission', 0)
            ds['settlement'] += dd.get('settlement', 0)
        for ch_ads in ad_daily.values():
            ds['ad_cost'] += ch_ads.get(dt, 0)
        ds['net'] = ds['settlement'] - ds['ad_cost']
        daily_summary.append(ds)

    return jsonify({
        'month': month,
        'channels': channel_summary,
        'total': total,
        'daily': daily_summary,
        'ad_items': sorted(ad_items, key=lambda x: x['date']),
        'deductions': deductions,
    })
