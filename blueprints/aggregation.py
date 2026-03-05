"""
aggregation.py — 통합 집계 Blueprint.
DB 기반 보고서: stock_ledger(SALES_OUT) + daily_revenue 조회 → 엑셀 다운로드.
파일 업로드 불필요 — 이미 DB에 모든 데이터가 실시간 반영되어 있음.
"""
import os
import re
import unicodedata
from datetime import datetime
from services.tz_utils import now_kst


def _sanitize_filename(name: str) -> str:
    """파일명에 사용 불가한 문자를 _로 치환 (윈도우 호환)."""
    return re.sub(r'[\\/:*?"<>|]', '_', str(name).strip())


def _safe_int(val) -> int:
    """문자열 '3.0', None, '' 등을 안전하게 int로 변환."""
    try:
        return int(float(val or 0))
    except (ValueError, TypeError):
        return 0

import pandas as pd
from flask import (
    Blueprint, render_template, request, current_app,
    flash, redirect, url_for, send_file, jsonify, abort,
)
from flask_login import login_required, current_user

from auth import role_required

aggregation_bp = Blueprint('aggregation', __name__, url_prefix='/aggregation')


def _norm(text):
    return unicodedata.normalize('NFC', str(text).strip())


def _list_result_files(output_dir, limit=30):
    """최근 결과 파일 목록."""
    if not os.path.exists(output_dir):
        return []
    files = [f for f in os.listdir(output_dir)
             if f.endswith(('.xlsx', '.xls', '.csv'))]
    files.sort(key=lambda f: os.path.getmtime(os.path.join(output_dir, f)),
               reverse=True)
    return files[:limit]


@aggregation_bp.route('/')
@role_required('admin', 'manager', 'sales')
def index():
    """통합 집계 페이지"""
    output_dir = current_app.config['OUTPUT_FOLDER']
    result_files = _list_result_files(output_dir)
    return render_template('aggregation/index.html', result_files=result_files)


# ================================================================
# 날짜별 요약 API
# ================================================================

@aggregation_bp.route('/api/summary')
@role_required('admin', 'manager', 'sales')
def api_summary():
    """기간별 요약: 출고(stock_ledger SALES_OUT) + 매출(daily_revenue)"""
    date_from = request.args.get('date_from') or request.args.get('date')
    date_to = request.args.get('date_to') or date_from
    if not date_from:
        return jsonify({'error': '날짜를 지정하세요'}), 400

    db = current_app.db
    try:
        # 출고 현황 (SALES_OUT)
        outbound = db.query_stock_ledger(
            date_from=date_from, date_to=date_to,
            type_list=['SALES_OUT'])

        outbound_count = len(outbound)
        outbound_items = set()
        outbound_qty = 0
        location_counts = {}
        for o in outbound:
            outbound_items.add(o.get('product_name', ''))
            outbound_qty += abs(_safe_int(o.get('qty', 0)))
            loc = o.get('location', '기타')
            location_counts[loc] = location_counts.get(loc, 0) + 1

        # 거래처 출고 (SALES_OUT with note containing 거래처 etc.)
        # → stock_ledger에는 구분 없이 SALES_OUT으로 들어가므로 전체 포함됨

        # 매출 현황
        revenue = db.query_revenue(date_from=date_from, date_to=date_to)
        revenue_total = 0
        revenue_count = len(revenue or [])
        category_revenue = {}
        channel_revenue = {}
        for r in (revenue or []):
            amt = float(r.get('revenue', 0) or r.get('amount', 0) or 0)
            revenue_total += amt
            cat = r.get('category', '기타')
            category_revenue[cat] = category_revenue.get(cat, 0) + amt
            ch = r.get('channel', '')
            if ch:
                channel_revenue[ch] = channel_revenue.get(ch, 0) + amt

        # 입고 현황
        inbound = db.query_stock_ledger(
            date_from=date_from, date_to=date_to,
            type_list=['INBOUND'])
        inbound_count = len(inbound)

        # 생산 현황
        production = db.query_stock_ledger(
            date_from=date_from, date_to=date_to,
            type_list=['PRODUCTION', 'PROD_OUT'])
        production_count = len(production)

        return jsonify({
            'date_from': date_from,
            'date_to': date_to,
            'outbound_count': outbound_count,
            'outbound_items': len(outbound_items),
            'outbound_qty': outbound_qty,
            'locations': location_counts,
            'inbound_count': inbound_count,
            'production_count': production_count,
            'revenue_count': revenue_count,
            'revenue_total': int(revenue_total),
            'category_revenue': {k: int(v) for k, v in category_revenue.items()},
            'channel_revenue': {k: int(v) for k, v in channel_revenue.items()},
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ================================================================
# 채널별 주문수량 집계 API
# ================================================================

def _channel_group(channel):
    """채널명 → 집계 그룹 분류."""
    ch = (channel or '').strip()
    if ch in ('N배송_수동', 'N배송'):
        return 'N배송'
    if ch == '쿠팡':
        return '쿠팡매출'
    return '일반매출'


@aggregation_bp.route('/api/channel-orders')
@role_required('admin', 'manager', 'sales')
def api_channel_orders():
    """채널별 일자별 주문수량 집계.

    order_transactions(온라인) + daily_revenue(거래처매출/로켓) 통합.
    그룹: 일반매출, 쿠팡/로켓, N배송, 거래처매출, 합계.
    """
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to') or date_from
    if not date_from:
        return jsonify({'error': '날짜를 지정하세요'}), 400

    db = current_app.db
    try:
        from collections import defaultdict

        # 일자별 그룹별 {date: {group: qty}}
        agg = defaultdict(lambda: defaultdict(int))
        groups = ['일반매출', '쿠팡매출', '로켓', 'N배송', '거래처매출']

        # 1. order_transactions (온라인 채널)
        offset = 0
        while True:
            resp = db.client.table('order_transactions') \
                .select('order_date,channel,qty') \
                .eq('status', '정상') \
                .gte('order_date', date_from) \
                .lte('order_date', date_to) \
                .range(offset, offset + 999) \
                .execute()
            batch = resp.data or []
            for r in batch:
                d = r.get('order_date', '')
                if not d:
                    continue
                grp = _channel_group(r.get('channel', ''))
                agg[d][grp] += _safe_int(r.get('qty', 0))
            if len(batch) < 1000:
                break
            offset += 1000

        # 2. daily_revenue (거래처매출, 로켓)
        offset = 0
        while True:
            resp = db.client.table('daily_revenue') \
                .select('revenue_date,category,qty') \
                .in_('category', ['거래처매출', '로켓']) \
                .gte('revenue_date', date_from) \
                .lte('revenue_date', date_to) \
                .range(offset, offset + 999) \
                .execute()
            batch = resp.data or []
            for r in batch:
                d = r.get('revenue_date', '')
                if not d:
                    continue
                cat = (r.get('category') or '').strip()
                if cat == '로켓':
                    grp = '로켓'
                else:
                    grp = '거래처매출'
                agg[d][grp] += _safe_int(r.get('qty', 0))
            if len(batch) < 1000:
                break
            offset += 1000

        # 결과 구성
        dates = sorted(agg.keys())
        rows = []
        totals = {g: 0 for g in groups}
        totals['합계'] = 0

        for d in dates:
            row = {'date': d}
            row_total = 0
            for g in groups:
                v = agg[d].get(g, 0)
                row[g] = v
                row_total += v
                totals[g] += v
            row['합계'] = row_total
            totals['합계'] += row_total
            rows.append(row)

        return jsonify({
            'groups': groups,
            'rows': rows,
            'totals': totals,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ================================================================
# DB 기반 보고서 생성 (출고+매출 엑셀)
# ================================================================

def _load_sort_order(db):
    """option_master에서 품목별 sort_order 로드 → {product_name: sort_order}."""
    try:
        rows = db.query_option_master()
        if not rows:
            return {}
        return {
            _norm(r.get('product_name', '')): int(r.get('sort_order', 999) or 999)
            for r in rows if r.get('product_name')
        }
    except Exception:
        return {}


def _build_pivot_excel(data, sort_map, output_dir, date_file, ts,
                       filename_prefix, date_key, sheets):
    """품목(행) × 날짜(열) 피벗 엑셀 생성.

    Args:
        data: list of dicts (revenue rows or outbound rows)
        sort_map: {product_name: sort_order}
        sheets: [{'title', 'value_key', 'alt_key'?, 'abs_val'?, 'num_fmt'}]
    Returns:
        filepath or None
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # 날짜/품목 수집 + 시트별 데이터 집계
    dates = sorted(set(r.get(date_key, '') for r in data if r.get(date_key)))
    if not dates:
        return None

    # 시트별 집계: {sheet_idx: {product: {date: value}}}
    sheet_aggs = [{} for _ in sheets]
    for r in data:
        nm = _norm(r.get('product_name', ''))
        d = r.get(date_key, '')
        if not nm or not d:
            continue
        for si, sh in enumerate(sheets):
            val = _safe_int(r.get(sh['value_key'], 0)
                            or r.get(sh.get('alt_key', ''), 0) or 0)
            if sh.get('abs_val'):
                val = abs(val)
            if nm not in sheet_aggs[si]:
                sheet_aggs[si][nm] = {}
            sheet_aggs[si][nm][d] = sheet_aggs[si][nm].get(d, 0) + val

    # 품목 목록: sort_order 기준
    all_products = set()
    for sa in sheet_aggs:
        all_products.update(sa.keys())
    products = sorted(all_products,
                      key=lambda x: (sort_map.get(x, 999), x))

    if not products:
        return None

    # 스타일
    header_fill = PatternFill(start_color="D6EAF8", fill_type="solid")
    total_fill = PatternFill(start_color="FEF9E7", fill_type="solid")
    bold = Font(bold=True)
    right_align = Alignment(horizontal="right")
    center_align = Alignment(horizontal="center")
    thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))

    # 날짜 표시 (월/일)
    def _short_date(d):
        try:
            dt = datetime.strptime(d, '%Y-%m-%d')
            return f"{dt.month}/{dt.day}"
        except Exception:
            return d

    wb = Workbook()
    first_sheet = True

    for si, sh in enumerate(sheets):
        if first_sheet:
            ws = wb.active
            ws.title = sh['title']
            first_sheet = False
        else:
            ws = wb.create_sheet(title=sh['title'])

        num_fmt = sh.get('num_fmt', '#,##0')
        sa = sheet_aggs[si]

        # 헤더: 품목명 | 날짜1 | 날짜2 | ... | 합계
        ws.column_dimensions['A'].width = 30
        c = ws.cell(row=1, column=1, value='품목명')
        c.font = bold
        c.fill = header_fill
        c.alignment = center_align

        for di, d in enumerate(dates):
            col = di + 2
            cl = get_column_letter(col)
            ws.column_dimensions[cl].width = 12
            c = ws.cell(row=1, column=col, value=_short_date(d))
            c.font = bold
            c.fill = header_fill
            c.alignment = center_align

        total_col = len(dates) + 2
        cl = get_column_letter(total_col)
        ws.column_dimensions[cl].width = 14
        c = ws.cell(row=1, column=total_col, value='합계')
        c.font = bold
        c.fill = header_fill
        c.alignment = center_align

        # 데이터 행
        row_num = 2
        date_totals = {d: 0 for d in dates}
        grand_total = 0

        for nm in products:
            product_data = sa.get(nm, {})
            row_total = sum(product_data.values())
            if row_total <= 0:
                continue

            ws.cell(row=row_num, column=1, value=nm).border = thin_border
            for di, d in enumerate(dates):
                val = product_data.get(d, 0)
                c = ws.cell(row=row_num, column=di + 2,
                            value=val if val else '')
                c.border = thin_border
                if val:
                    c.number_format = num_fmt
                    c.alignment = right_align
                date_totals[d] += val

            c = ws.cell(row=row_num, column=total_col, value=row_total)
            c.number_format = num_fmt
            c.alignment = right_align
            c.font = bold
            c.border = thin_border
            grand_total += row_total
            row_num += 1

        # 합계 행
        c = ws.cell(row=row_num, column=1, value='합계')
        c.font = bold
        c.fill = total_fill
        for di, d in enumerate(dates):
            c = ws.cell(row=row_num, column=di + 2, value=date_totals[d])
            c.font = bold
            c.fill = total_fill
            c.number_format = num_fmt
            c.alignment = right_align
        c = ws.cell(row=row_num, column=total_col, value=grand_total)
        c.font = bold
        c.fill = total_fill
        c.number_format = num_fmt
        c.alignment = right_align

    filepath = os.path.join(
        output_dir, f"{filename_prefix}_{date_file}_{ts}.xlsx")
    wb.save(filepath)
    return filepath


def _write_styled_excel(filepath, headers, data_rows, total_row=None,
                        col_widths=None, num_cols=None):
    """openpyxl로 스타일링된 엑셀 작성 (이전 통합집계 양식).

    Args:
        filepath: 저장 경로
        headers: 헤더 리스트 ['품목명', '수량', ...]
        data_rows: [[val, val, ...], ...] 2D 리스트
        total_row: 합계 행 리스트 (None이면 생략)
        col_widths: {col_letter: width} (None이면 자동)
        num_cols: 숫자 포맷 적용할 컬럼 인덱스 set (0-based)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active

    header_fill = PatternFill(start_color="D6EAF8", fill_type="solid")
    total_fill = PatternFill(start_color="FEF9E7", fill_type="solid")
    bold = Font(bold=True)
    bold_white = Font(bold=True, color="FFFFFF")
    right_align = Alignment(horizontal="right")
    num_fmt = '#,##0'
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )
    if num_cols is None:
        num_cols = set()

    # 컬럼 너비
    if col_widths:
        for col_letter, w in col_widths.items():
            ws.column_dimensions[col_letter].width = w

    # 헤더
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = bold
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    # 데이터
    for ri, row_data in enumerate(data_rows, 2):
        for ci, val in enumerate(row_data):
            c = ws.cell(row=ri, column=ci + 1, value=val)
            c.border = thin_border
            if ci in num_cols and isinstance(val, (int, float)):
                c.number_format = num_fmt
                c.alignment = right_align

    # 합계 행
    if total_row:
        tr = len(data_rows) + 2
        for ci, val in enumerate(total_row):
            c = ws.cell(row=tr, column=ci + 1, value=val)
            c.font = bold
            c.fill = total_fill
            if ci in num_cols and isinstance(val, (int, float)):
                c.number_format = num_fmt
                c.alignment = right_align

    wb.save(filepath)


@aggregation_bp.route('/generate', methods=['POST'])
@role_required('admin', 'manager', 'sales')
def generate_report():
    """stock_ledger(SALES_OUT) + daily_revenue 조회 → 엑셀 생성"""
    date_from = request.form.get('date_from') or request.form.get('date')
    date_to = request.form.get('date_to') or date_from
    if not date_from:
        flash('날짜를 지정하세요.', 'danger')
        return redirect(url_for('aggregation.index'))

    date_label = date_from if date_from == date_to else f"{date_from}~{date_to}"

    db = current_app.db
    output_dir = current_app.config['OUTPUT_FOLDER']
    os.makedirs(output_dir, exist_ok=True)

    try:
        ts = now_kst().strftime("%Y%m%d_%H%M%S")
        generated_files = []
        date_file = date_from if date_from == date_to else f"{date_from}_{date_to}"

        # sort_order 로드
        sort_map = _load_sort_order(db)

        # ══════════════════════════════════════════════
        # 1. 통합집계표: stock_ledger SALES_OUT (이전 양식)
        # ══════════════════════════════════════════════
        outbound = db.query_stock_ledger(
            date_from=date_from, date_to=date_to,
            type_list=['SALES_OUT'])

        if not outbound:
            flash(f'{date_label} 기간의 출고(SALES_OUT) 데이터가 없습니다.', 'warning')
            return redirect(url_for('aggregation.index'))

        # 품목별/창고별 합계: {(product_name, location): total_qty}
        agg = {}
        for o in outbound:
            nm = _norm(o.get('product_name', ''))
            loc = o.get('location', '기타')
            qty = abs(_safe_int(o.get('qty', 0)))
            if not nm:
                continue
            key = (nm, loc)
            agg[key] = agg.get(key, 0) + qty

        if not agg:
            flash('유효한 출고 데이터가 없습니다.', 'warning')
            return redirect(url_for('aggregation.index'))

        # sort_order 기준 정렬
        sorted_items = sorted(
            agg.items(),
            key=lambda x: (sort_map.get(x[0][0], 999), x[0][1], x[0][0])
        )

        # 통합집계표 (합계본) — 출력순서 | 품목명 | 창고 | 수량
        headers = ['출력순서', '품목명', '창고', '수량']
        data_rows = []
        for (nm, loc), qty in sorted_items:
            if qty <= 0:
                continue
            so = sort_map.get(nm, '')
            data_rows.append([so if so != 999 else '', nm, loc, qty])

        total_items = len(data_rows)
        total_qty = sum(r[3] for r in data_rows)
        total_row = ['', '합계', f'{total_items}종', total_qty]

        agg_path = os.path.join(output_dir, f"통합집계표_{date_file}_{ts}.xlsx")
        _write_styled_excel(
            agg_path, headers, data_rows, total_row,
            col_widths={'A': 10, 'B': 30, 'C': 12, 'D': 12},
            num_cols={0, 3}
        )
        generated_files.append(agg_path)

        # 창고별 분리 — 출력순서 | 품목명 | 수량
        warehouses = sorted(set(loc for (nm, loc), qty in sorted_items if qty > 0))
        for wh_name in warehouses:
            wh_rows = []
            for (nm, loc), qty in sorted_items:
                if loc != wh_name or qty <= 0:
                    continue
                so = sort_map.get(nm, '')
                wh_rows.append([so if so != 999 else '', nm, qty])
            if not wh_rows:
                continue
            wh_total = sum(r[2] for r in wh_rows)
            wh_path = os.path.join(
                output_dir,
                f"통합출고_{_sanitize_filename(wh_name)}_{date_file}_{ts}.xlsx"
            )
            _write_styled_excel(
                wh_path,
                ['출력순서', '품목명', '수량'],
                wh_rows,
                ['', f'합계 {len(wh_rows)}종', wh_total],
                col_widths={'A': 10, 'B': 30, 'C': 12},
                num_cols={0, 2}
            )
            generated_files.append(wh_path)

        # ══════════════════════════════════════════════
        # 2. 일일매출: 일자별 × 채널별 피벗 테이블
        # ══════════════════════════════════════════════
        revenue = db.query_revenue(date_from=date_from, date_to=date_to)

        if revenue:
            # 채널 목록 수집 + 일자별×채널별 합산
            ch_totals = {}    # {channel: total_revenue}
            by_date_ch = {}   # {date: {channel: revenue}}

            from services.channel_config import normalize_channel_display
            for r in revenue:
                d = r.get('revenue_date', '')
                raw_ch = r.get('channel', '') or ''
                # 'None' 문자열, 빈값 → category로 폴백
                if not raw_ch or raw_ch in ('None', 'none', 'null'):
                    ch = r.get('category', '기타') or '기타'
                else:
                    ch = raw_ch
                ch = normalize_channel_display(ch)
                rev = _safe_int(r.get('revenue', 0) or r.get('amount', 0) or 0)
                if not d:
                    continue
                ch_totals[ch] = ch_totals.get(ch, 0) + rev
                if d not in by_date_ch:
                    by_date_ch[d] = {}
                by_date_ch[d][ch] = by_date_ch[d].get(ch, 0) + rev

            # 채널 순서: 매출 내림차순
            channels = [k for k, v in sorted(ch_totals.items(), key=lambda x: -x[1])]

            if by_date_ch and channels:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                wb = Workbook()
                ws = wb.active
                ws.title = "일일매출_채널별"

                header_fill = PatternFill(start_color="D6EAF8", fill_type="solid")
                total_fill = PatternFill(start_color="FEF9E7", fill_type="solid")
                grand_fill = PatternFill(start_color="AED6F1", fill_type="solid")
                bold = Font(bold=True)
                right_align = Alignment(horizontal="right")
                center_align = Alignment(horizontal="center")
                num_fmt = '#,##0'
                thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))

                # 헤더
                ws.column_dimensions['A'].width = 12
                rev_headers = ['날짜'] + channels + ['합계']
                for ci, h in enumerate(rev_headers, 1):
                    c = ws.cell(row=1, column=ci, value=h)
                    c.font = bold
                    c.fill = header_fill
                    c.alignment = center_align
                    if ci > 1:
                        from openpyxl.utils import get_column_letter
                        ws.column_dimensions[get_column_letter(ci)].width = 14

                # 날짜별 데이터
                row_num = 2
                for d in sorted(by_date_ch.keys()):
                    ch_data = by_date_ch[d]
                    ws.cell(row=row_num, column=1, value=d).font = Font(bold=True)
                    day_total = 0
                    for chi, ch in enumerate(channels, 2):
                        val = ch_data.get(ch, 0)
                        c = ws.cell(row=row_num, column=chi, value=val if val else '')
                        if val:
                            c.number_format = num_fmt
                            c.alignment = right_align
                        day_total += val
                    # 합계
                    c = ws.cell(row=row_num, column=len(channels) + 2, value=day_total)
                    c.number_format = num_fmt
                    c.alignment = right_align
                    c.font = bold
                    row_num += 1

                # 채널별 합계 행
                for ci, h in enumerate(['합계'] + [ch_totals.get(ch, 0) for ch in channels] + [sum(ch_totals.values())]):
                    c = ws.cell(row=row_num, column=ci + 1, value=h)
                    c.font = bold
                    c.fill = total_fill
                    if isinstance(h, (int, float)):
                        c.number_format = num_fmt
                        c.alignment = right_align

                rev_path = os.path.join(output_dir, f"일일매출_{date_file}_{ts}.xlsx")
                wb.save(rev_path)
                generated_files.append(rev_path)

        # ══════════════════════════════════════════════
        # 3. 매출집계표: 품목(행) × 날짜(열) 피벗 (판매수량 + 매출액)
        # ══════════════════════════════════════════════
        if revenue:
            rev_path = _build_pivot_excel(
                revenue, sort_map, output_dir, date_file, ts,
                filename_prefix='매출집계표',
                date_key='revenue_date',
                sheets=[
                    {'title': '판매수량', 'value_key': 'qty', 'num_fmt': '#,##0'},
                    {'title': '매출액', 'value_key': 'revenue', 'alt_key': 'amount',
                     'num_fmt': '#,##0'},
                ],
            )
            if rev_path:
                generated_files.append(rev_path)

        # ══════════════════════════════════════════════
        # 3-2. 출고집계표: 품목(행) × 날짜(열) 피벗 (출고수량)
        # ══════════════════════════════════════════════
        if outbound:
            out_path = _build_pivot_excel(
                outbound, sort_map, output_dir, date_file, ts,
                filename_prefix='출고집계표',
                date_key='transaction_date',
                sheets=[
                    {'title': '출고수량', 'value_key': 'qty', 'abs_val': True,
                     'num_fmt': '#,##0'},
                ],
            )
            if out_path:
                generated_files.append(out_path)

        # ══════════════════════════════════════════════
        # 4. 일자별 종합 보고서 (출고+매출 날짜별 구분)
        # ══════════════════════════════════════════════
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "일자별종합"

            # 스타일 정의
            date_fill = PatternFill(start_color="D6EAF8", fill_type="solid")
            section_fill = PatternFill(start_color="F2F3F4", fill_type="solid")
            subtotal_fill = PatternFill(start_color="FEF9E7", fill_type="solid")
            grand_fill = PatternFill(start_color="AED6F1", fill_type="solid")
            bold = Font(bold=True)
            bold_lg = Font(bold=True, size=13)
            right_align = Alignment(horizontal="right")
            num_fmt = '#,##0'

            # 컬럼 너비
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 28
            ws.column_dimensions['C'].width = 14
            ws.column_dimensions['D'].width = 14
            ws.column_dimensions['E'].width = 14
            ws.column_dimensions['F'].width = 16
            ws.column_dimensions['G'].width = 14

            # 날짜 목록 (출고+매출 합산)
            dates = sorted(set(
                [o.get('transaction_date', '') for o in outbound if o.get('transaction_date')] +
                [r.get('revenue_date', '') for r in (revenue or []) if r.get('revenue_date')]
            ))

            grand_out_kinds = 0
            grand_out_qty = 0
            grand_rev_count = 0
            grand_rev_total = 0
            row_num = 1

            for d in dates:
                # ── 날짜 헤더 ──
                ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
                cell = ws.cell(row=row_num, column=1, value=f"  {d}")
                cell.font = bold_lg
                cell.fill = date_fill
                for c in range(1, 8):
                    ws.cell(row=row_num, column=c).fill = date_fill
                row_num += 1

                # ── [출고] 헤더 ──
                for ci, hdr in enumerate(["[출고]", "품목명", "수량", "창고"], 1):
                    c = ws.cell(row=row_num, column=ci, value=hdr)
                    c.font = bold
                    c.fill = section_fill
                row_num += 1

                # 출고 데이터 (해당 날짜, sort_order 정렬)
                day_out = [o for o in outbound if o.get('transaction_date') == d]
                day_agg = {}
                for o in day_out:
                    nm = _norm(o.get('product_name', ''))
                    loc = o.get('location', '기타')
                    qty = abs(_safe_int(o.get('qty', 0)))
                    if nm:
                        key = (nm, loc)
                        day_agg[key] = day_agg.get(key, 0) + qty

                for (nm, loc), qty in sorted(day_agg.items(),
                        key=lambda x: (sort_map.get(x[0][0], 999), x[0][0])):
                    ws.cell(row=row_num, column=2, value=nm)
                    c_qty = ws.cell(row=row_num, column=3, value=qty)
                    c_qty.number_format = num_fmt
                    c_qty.alignment = right_align
                    ws.cell(row=row_num, column=4, value=loc)
                    row_num += 1

                # 출고 소계
                day_out_qty = sum(day_agg.values())
                for ci, val in enumerate([
                    "", f"소계: {len(day_agg)}종", day_out_qty, ""
                ], 1):
                    c = ws.cell(row=row_num, column=ci, value=val)
                    c.font = bold
                    c.fill = subtotal_fill
                    if ci == 3:
                        c.number_format = num_fmt
                        c.alignment = right_align
                row_num += 1

                # ── [매출] 헤더 ──
                for ci, hdr in enumerate(["[매출]", "품목명", "채널", "수량", "단가", "매출액"], 1):
                    c = ws.cell(row=row_num, column=ci, value=hdr)
                    c.font = bold
                    c.fill = section_fill
                row_num += 1

                # 매출 데이터 (채널별 그룹핑)
                day_rev = [r for r in (revenue or []) if r.get('revenue_date') == d]
                day_rev_total = 0
                for r in sorted(day_rev, key=lambda x: (
                        x.get('channel', ''), x.get('product_name', ''))):
                    amt = _safe_int(r.get('revenue', 0) or r.get('amount', 0) or 0)
                    ws.cell(row=row_num, column=2, value=r.get('product_name', ''))
                    ws.cell(row=row_num, column=3, value=r.get('channel', ''))
                    c4 = ws.cell(row=row_num, column=4, value=_safe_int(r.get('qty', 0)))
                    c4.number_format = num_fmt
                    c4.alignment = right_align
                    c5 = ws.cell(row=row_num, column=5, value=_safe_int(r.get('unit_price', 0)))
                    c5.number_format = num_fmt
                    c5.alignment = right_align
                    c6 = ws.cell(row=row_num, column=6, value=amt)
                    c6.number_format = num_fmt
                    c6.alignment = right_align
                    day_rev_total += amt
                    row_num += 1

                # 매출 소계
                for ci, val in enumerate([
                    "", f"소계: {len(day_rev)}건", "", "", "", day_rev_total
                ], 1):
                    c = ws.cell(row=row_num, column=ci, value=val)
                    c.font = bold
                    c.fill = subtotal_fill
                    if ci == 6:
                        c.number_format = num_fmt
                        c.alignment = right_align
                row_num += 1

                row_num += 1  # 빈 행

                grand_out_kinds += len(day_agg)
                grand_out_qty += day_out_qty
                grand_rev_count += len(day_rev)
                grand_rev_total += day_rev_total

            # ── 총계 ──
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
            c = ws.cell(row=row_num, column=1, value="  TOTAL")
            c.font = bold_lg
            c.fill = grand_fill
            for ci in range(1, 7):
                ws.cell(row=row_num, column=ci).fill = grand_fill
            row_num += 1

            for ci, val in enumerate(["", f"총 출고: {grand_out_kinds}종", grand_out_qty, "", "", ""], 1):
                c = ws.cell(row=row_num, column=ci, value=val)
                c.font = bold
                if ci == 3:
                    c.number_format = num_fmt
                    c.alignment = right_align
            row_num += 1

            for ci, val in enumerate(["", f"총 매출: {grand_rev_count}건", "", "", "", grand_rev_total], 1):
                c = ws.cell(row=row_num, column=ci, value=val)
                c.font = bold
                if ci == 6:
                    c.number_format = num_fmt
                    c.alignment = right_align

            daily_path = os.path.join(output_dir, f"일자별종합_{date_file}_{ts}.xlsx")
            wb.save(daily_path)
            generated_files.append(daily_path)

        except Exception as daily_err:
            import traceback
            traceback.print_exc()
            flash(f'일자별 종합 보고서 생성 오류: {daily_err}', 'warning')

        # ══════════════════════════════════════════════
        # 7. 채널별 주문수량: 세트BOM 풀기 + 창고 표시
        #    (N배송은 세트 풀기 제외)
        # ══════════════════════════════════════════════
        try:
            from collections import defaultdict as _dd
            from services.order_to_stock_service import (
                _load_bom_map, _load_option_map, _get_warehouse, _decompose
            )
            from services.channel_config import CHANNEL_REVENUE_MAP

            # BOM + 옵션마스터 로드
            bom_map = _load_bom_map(db)
            opt_map = _load_option_map(db)
            bom_all = bom_map.get('모든채널', {})
            bom_coupang = bom_map.get('쿠팡전용', {})

            # {date: {(product_name, warehouse): {channel_group: qty}}}
            ch_agg = _dd(lambda: _dd(lambda: _dd(int)))
            ch_groups = ['일반매출', '쿠팡매출', '로켓', 'N배송', '거래처매출']

            # order_transactions (온라인 채널)
            ot_offset = 0
            while True:
                ot_resp = db.client.table('order_transactions') \
                    .select('order_date,product_name,channel,qty') \
                    .eq('status', '정상') \
                    .gte('order_date', date_from) \
                    .lte('order_date', date_to) \
                    .range(ot_offset, ot_offset + 999) \
                    .execute()
                ot_batch = ot_resp.data or []
                for r in ot_batch:
                    d = r.get('order_date', '')
                    pn = _norm(r.get('product_name', ''))
                    if not d or not pn:
                        continue
                    qty = _safe_int(r.get('qty', 0))
                    if qty <= 0:
                        continue
                    grp = _channel_group(r.get('channel', ''))
                    ch_raw = r.get('channel', '')
                    rev_cat = CHANNEL_REVENUE_MAP.get(ch_raw, '일반매출')
                    is_n = (grp == 'N배송')

                    # N배송: 세트 안 풀기 + 창고 CJ용인 고정
                    if is_n:
                        wh = 'CJ용인'
                        ch_agg[d][(pn, wh)][grp] += qty
                    else:
                        # 세트 BOM 풀기 (쿠팡/로켓은 쿠팡전용 BOM 우선)
                        if grp in ('쿠팡매출', '로켓') or rev_cat in ('쿠팡매출', '로켓'):
                            decomposed = _decompose(pn, qty, bom_coupang, bom_all)
                        else:
                            decomposed = _decompose(pn, qty, bom_all, None)
                        # 분해된 단품별 창고 결정
                        for item_name, item_qty in decomposed.items():
                            wh = _get_warehouse(item_name, opt_map)
                            ch_agg[d][(_norm(item_name), wh)][grp] += item_qty

                if len(ot_batch) < 1000:
                    break
                ot_offset += 1000

            # daily_revenue (거래처매출, 로켓)
            dr_offset = 0
            while True:
                dr_resp = db.client.table('daily_revenue') \
                    .select('revenue_date,product_name,category,qty') \
                    .in_('category', ['거래처매출', '로켓']) \
                    .gte('revenue_date', date_from) \
                    .lte('revenue_date', date_to) \
                    .range(dr_offset, dr_offset + 999) \
                    .execute()
                dr_batch = dr_resp.data or []
                for r in dr_batch:
                    d = r.get('revenue_date', '')
                    pn = _norm(r.get('product_name', ''))
                    if not d or not pn:
                        continue
                    qty = _safe_int(r.get('qty', 0))
                    if qty <= 0:
                        continue
                    cat = (r.get('category') or '').strip()
                    grp = '로켓' if cat == '로켓' else '거래처매출'

                    # 거래처매출/로켓도 세트 BOM 풀기
                    if grp == '로켓':
                        decomposed = _decompose(pn, qty, bom_coupang, bom_all)
                    else:
                        decomposed = _decompose(pn, qty, bom_all, None)
                    for item_name, item_qty in decomposed.items():
                        wh = _get_warehouse(_norm(item_name), opt_map)
                        ch_agg[d][(_norm(item_name), wh)][grp] += item_qty

                if len(dr_batch) < 1000:
                    break
                dr_offset += 1000

            if ch_agg:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter

                # 창고별 색상 구분
                wh_fills = {
                    '넥스원': PatternFill(start_color="FFFFFF", fill_type="solid"),
                    '해서': PatternFill(start_color="FFF3E0", fill_type="solid"),
                    'CJ용인': PatternFill(start_color="E8F5E9", fill_type="solid"),
                }

                wb_ch = Workbook()
                first_sheet = True

                header_fill = PatternFill(start_color="D6EAF8", fill_type="solid")
                total_fill = PatternFill(start_color="FEF9E7", fill_type="solid")
                wh_subtotal_fill = PatternFill(start_color="EBF5FB", fill_type="solid")
                ch_bold = Font(bold=True)
                ch_right = Alignment(horizontal="right")
                ch_center = Alignment(horizontal="center")
                ch_border = Border(bottom=Side(style="thin", color="CCCCCC"))

                for d in sorted(ch_agg.keys()):
                    day_data = ch_agg[d]
                    if not day_data:
                        continue

                    # 시트 생성
                    if first_sheet:
                        ws = wb_ch.active
                        ws.title = d
                        first_sheet = False
                    else:
                        ws = wb_ch.create_sheet(title=d)

                    # 헤더: 순서|품목명|창고|채널그룹들|합계
                    headers_ch = ['순서', '품목명', '창고'] + ch_groups + ['합계']
                    ws.column_dimensions['A'].width = 8
                    ws.column_dimensions['B'].width = 30
                    ws.column_dimensions['C'].width = 10
                    for ci, h in enumerate(headers_ch):
                        col = ci + 1
                        if ci >= 3:
                            ws.column_dimensions[get_column_letter(col)].width = 12
                        c = ws.cell(row=1, column=col, value=h)
                        c.font = ch_bold
                        c.fill = header_fill
                        c.alignment = ch_center

                    # 품목+창고 정렬: 창고별 그룹 → sort_order → 품목명
                    wh_order = {'넥스원': 0, '해서': 1, 'CJ용인': 2}
                    sorted_keys = sorted(
                        day_data.keys(),
                        key=lambda x: (wh_order.get(x[1], 9), sort_map.get(x[0], 999), x[0])
                    )

                    row_num = 2
                    col_totals = {g: 0 for g in ch_groups}
                    grand_total = 0
                    prev_wh = None
                    wh_subtotals = {g: 0 for g in ch_groups}
                    wh_grand = 0
                    wh_count = 0

                    def _write_wh_subtotal(ws, row, wh_name, wh_cnt, wh_subs, wh_grd):
                        """창고별 소계 행 작성."""
                        ws.cell(row=row, column=1, value='').fill = wh_subtotal_fill
                        c = ws.cell(row=row, column=2,
                                    value=f'[{wh_name}] 소계 {wh_cnt}종')
                        c.font = ch_bold
                        c.fill = wh_subtotal_fill
                        ws.cell(row=row, column=3, value='').fill = wh_subtotal_fill
                        for gi, g in enumerate(ch_groups):
                            c = ws.cell(row=row, column=gi + 4, value=wh_subs[g])
                            c.font = ch_bold
                            c.fill = wh_subtotal_fill
                            c.number_format = '#,##0'
                            c.alignment = ch_right
                        c = ws.cell(row=row, column=len(ch_groups) + 4, value=wh_grd)
                        c.font = ch_bold
                        c.fill = wh_subtotal_fill
                        c.number_format = '#,##0'
                        c.alignment = ch_right
                        return row + 1

                    for (pn, wh) in sorted_keys:
                        # 창고 변경 시 이전 창고 소계 출력
                        if prev_wh is not None and wh != prev_wh and wh_count > 0:
                            row_num = _write_wh_subtotal(
                                ws, row_num, prev_wh, wh_count, wh_subtotals, wh_grand)
                            wh_subtotals = {g: 0 for g in ch_groups}
                            wh_grand = 0
                            wh_count = 0

                        prev_wh = wh
                        wh_count += 1
                        so = sort_map.get(pn, 999)
                        row_fill = wh_fills.get(wh)

                        c = ws.cell(row=row_num, column=1,
                                    value=so if so != 999 else '')
                        c.border = ch_border
                        if row_fill:
                            c.fill = row_fill
                        c = ws.cell(row=row_num, column=2, value=pn)
                        c.border = ch_border
                        if row_fill:
                            c.fill = row_fill
                        c = ws.cell(row=row_num, column=3, value=wh)
                        c.border = ch_border
                        if row_fill:
                            c.fill = row_fill

                        row_total = 0
                        for gi, g in enumerate(ch_groups):
                            v = day_data[(pn, wh)].get(g, 0)
                            c = ws.cell(row=row_num, column=gi + 4,
                                        value=v if v else '')
                            c.border = ch_border
                            if row_fill:
                                c.fill = row_fill
                            if v:
                                c.number_format = '#,##0'
                                c.alignment = ch_right
                            col_totals[g] += v
                            wh_subtotals[g] += v
                            row_total += v

                        c = ws.cell(row=row_num, column=len(ch_groups) + 4,
                                    value=row_total)
                        c.number_format = '#,##0'
                        c.alignment = ch_right
                        c.font = ch_bold
                        c.border = ch_border
                        if row_fill:
                            c.fill = row_fill
                        grand_total += row_total
                        wh_grand += row_total
                        row_num += 1

                    # 마지막 창고 소계
                    if prev_wh is not None and wh_count > 0:
                        row_num = _write_wh_subtotal(
                            ws, row_num, prev_wh, wh_count, wh_subtotals, wh_grand)

                    # 총 합계 행
                    total_items = len(sorted_keys)
                    ws.cell(row=row_num, column=1, value='').fill = total_fill
                    c = ws.cell(row=row_num, column=2,
                                value=f'합계 {total_items}종')
                    c.font = ch_bold
                    c.fill = total_fill
                    ws.cell(row=row_num, column=3, value='').fill = total_fill
                    for gi, g in enumerate(ch_groups):
                        c = ws.cell(row=row_num, column=gi + 4, value=col_totals[g])
                        c.font = ch_bold
                        c.fill = total_fill
                        c.number_format = '#,##0'
                        c.alignment = ch_right
                    c = ws.cell(row=row_num, column=len(ch_groups) + 4,
                                value=grand_total)
                    c.font = ch_bold
                    c.fill = total_fill
                    c.number_format = '#,##0'
                    c.alignment = ch_right

                ch_path = os.path.join(output_dir, f"채널별주문수량_{date_file}_{ts}.xlsx")
                wb_ch.save(ch_path)
                generated_files.append(ch_path)

        except Exception as ch_err:
            import traceback
            traceback.print_exc()
            flash(f'채널별 주문수량 생성 오류: {ch_err}', 'warning')

        flash(f"[{date_label}] 보고서 생성 완료 — 출고 {total_items}종 {total_qty:,}개, "
              f"파일 {len(generated_files)}개", 'success')

        downloads = []
        for fpath in generated_files:
            fname = os.path.basename(fpath)
            downloads.append({
                'name': fname,
                'url': url_for('aggregation.download', filename=fname),
            })

        result_files = _list_result_files(output_dir)
        return render_template('aggregation/index.html',
                               result={'downloads': downloads},
                               result_files=result_files)

    except Exception as e:
        flash(f'보고서 생성 오류: {e}', 'danger')
        import traceback
        traceback.print_exc()
        return redirect(url_for('aggregation.index'))


# ================================================================
# 파일 다운로드/삭제
# ================================================================

@aggregation_bp.route('/download/<path:filename>')
@role_required('admin', 'manager', 'sales')
def download(filename):
    """파일 다운로드"""
    output_dir = os.path.abspath(current_app.config['OUTPUT_FOLDER'])
    safe_name = os.path.basename(filename)
    filepath = os.path.join(output_dir, safe_name)

    if not os.path.abspath(filepath).startswith(output_dir):
        abort(403)
    if not os.path.exists(filepath):
        flash('파일을 찾을 수 없습니다.', 'danger')
        return redirect(url_for('aggregation.index'))

    return send_file(
        filepath,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=safe_name,
    )


@aggregation_bp.route('/delete-file', methods=['POST'])
@role_required('admin')
def delete_file():
    """파일 삭제"""
    filenames = request.form.getlist('delete_files')
    if not filenames:
        flash('삭제할 파일을 선택하세요.', 'danger')
        return redirect(url_for('aggregation.index'))

    output_dir = os.path.abspath(current_app.config['OUTPUT_FOLDER'])
    deleted = 0
    for fname in filenames:
        safe_name = os.path.basename(fname)
        filepath = os.path.join(output_dir, safe_name)
        if os.path.abspath(filepath).startswith(output_dir) and os.path.exists(filepath):
            os.remove(filepath)
            deleted += 1

    if deleted > 0:
        flash(f'파일 {deleted}건 삭제 완료', 'success')
    else:
        flash('삭제할 파일이 없습니다.', 'warning')

    return redirect(url_for('aggregation.index'))
