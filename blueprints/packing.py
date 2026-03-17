"""
blueprints/packing.py — 패킹센터 (위탁업체 + 내부직원)
별도 로그인/회원가입 + 전용 GUI + 택배 송장 관리
"""
import json
import time
import io
from functools import wraps
from datetime import datetime, timezone

from flask import (
    Blueprint, render_template, redirect, url_for,
    flash, request, current_app, session, jsonify,
)
from flask_wtf.csrf import validate_csrf
from flask_login import login_user, logout_user, login_required, current_user
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField
from wtforms.validators import DataRequired, Length, EqualTo, Regexp

from models import User
from db_utils import get_db

packing_bp = Blueprint('packing', __name__, url_prefix='/packing')

# 내부직원 역할 (패킹센터 접근 허용)
_INTERNAL_ROLES = ('admin', 'manager', 'logistics', 'sales', 'general')
_ALL_PACKING_ROLES = ('packing',) + _INTERNAL_ROLES


# ── Forms ──

class PackingLoginForm(FlaskForm):
    username = StringField('아이디', validators=[DataRequired(), Length(max=80)])
    password = PasswordField('비밀번호', validators=[DataRequired()])


class PackingRegisterForm(FlaskForm):
    username = StringField('아이디', validators=[
        DataRequired(), Length(min=4, max=80),
        Regexp(r'^[a-zA-Z0-9_]+$', message='영문, 숫자, 밑줄만 사용 가능합니다.')
    ])
    name = StringField('이름', validators=[DataRequired(), Length(max=100)])
    company_name = StringField('업체명', validators=[DataRequired(), Length(max=200)])
    password = PasswordField('비밀번호', validators=[
        DataRequired(), Length(min=8, message='비밀번호는 8자 이상이어야 합니다.')
    ])
    password2 = PasswordField('비밀번호 확인', validators=[
        DataRequired(), EqualTo('password', message='비밀번호가 일치하지 않습니다.')
    ])


class PackingChangePasswordForm(FlaskForm):
    current_password = PasswordField('현재 비밀번호', validators=[DataRequired()])
    new_password = PasswordField('새 비밀번호', validators=[
        DataRequired(), Length(min=8)
    ])
    new_password2 = PasswordField('새 비밀번호 확인', validators=[
        DataRequired(), EqualTo('new_password', message='비밀번호가 일치하지 않습니다.')
    ])


# ── Decorator ──

def packing_required(f):
    """패킹센터 접근 제어 (packing + 운영자 허용)"""
    @wraps(f)
    @login_required
    def wrapped(*args, **kwargs):
        # packing 역할 + 관리자/책임자는 패킹센터 접근 가능
        if current_user.role not in _ALL_PACKING_ROLES:
            flash('패킹센터 접근 권한이 없습니다.', 'danger')
            return redirect(url_for('main.dashboard'))
        return f(*args, **kwargs)
    return wrapped


# ── Helper: auth.py 함수 재사용 ──

def _packing_log_action(action, target=None, detail=None, user_id=None):
    """감사 로그 기록"""
    from auth import _log_action
    _log_action(action, target=target, detail=detail, user_id=user_id)


# ── Routes ──

@packing_bp.route('/login', methods=['GET', 'POST'])
def packing_login():
    """패킹센터 전용 로그인"""
    if current_user.is_authenticated:
        if current_user.role in _ALL_PACKING_ROLES:
            return redirect(url_for('packing.index'))
        return redirect(url_for('main.dashboard'))

    form = PackingLoginForm()

    if form.validate_on_submit():
        from auth import _get_client_ip, _check_ip_rate_limit, _record_ip_attempt
        client_ip = _get_client_ip()

        # IP 차단 확인
        blocked_seconds = _check_ip_rate_limit(client_ip)
        if blocked_seconds > 0:
            minutes = blocked_seconds // 60 + 1
            flash(f'너무 많은 로그인 시도. {minutes}분 후 다시 시도하세요.', 'danger')
            return render_template('packing/login.html', form=form)

        row = get_db().query_user_by_username(form.username.data)
        user = User(row) if row else None

        # 계정 잠금 확인
        if user and user.is_locked():
            flash('계정이 잠겼습니다. 잠시 후 다시 시도해주세요.', 'danger')
            return render_template('packing/login.html', form=form)

        if user and user.check_password(form.password.data):
            # 패킹 + 운영자(admin/manager) 허용
            if user.role not in _ALL_PACKING_ROLES:
                flash('패킹센터 접근 권한이 없습니다.', 'danger')
                return render_template('packing/login.html', form=form)

            if not user.is_active_user:
                flash('비활성화된 계정입니다. 관리자에게 문의하세요.', 'danger')
                return render_template('packing/login.html', form=form)

            if not user.is_approved:
                flash('관리자 승인 대기 중입니다.', 'warning')
                return render_template('packing/login.html', form=form)

            # 로그인 성공
            get_db().update_user(user.id, {
                'failed_login_count': 0,
                'locked_until': None,
                'last_login': datetime.now(timezone.utc).isoformat(),
            })

            login_user(user, remember=False)
            session.permanent = True
            session['_last_active'] = time.time()
            _packing_log_action('packing_login', target=user.username,
                                detail=f'IP: {client_ip}')

            return redirect(url_for('packing.index'))
        else:
            # 로그인 실패
            _record_ip_attempt(client_ip)
            if user:
                new_count = user.failed_login_count + 1
                update_data = {'failed_login_count': new_count}
                max_attempts = current_app.config.get('LOGIN_MAX_ATTEMPTS', 5)
                if new_count >= max_attempts:
                    from datetime import timedelta
                    lockout = current_app.config.get('LOGIN_LOCKOUT_MINUTES', 15)
                    update_data['locked_until'] = (
                        datetime.now(timezone.utc) + timedelta(minutes=lockout)
                    ).isoformat()
                    flash(f'로그인 {max_attempts}회 실패. {lockout}분간 잠금됩니다.', 'danger')
                get_db().update_user(user.id, update_data)
            flash('아이디 또는 비밀번호가 올바르지 않습니다.', 'danger')

    return render_template('packing/login.html', form=form)


@packing_bp.route('/register', methods=['GET', 'POST'])
def packing_register():
    """패킹센터 위탁업체 회원가입"""
    if current_user.is_authenticated:
        if current_user.role == 'packing':
            return redirect(url_for('packing.index'))
        return redirect(url_for('main.dashboard'))

    form = PackingRegisterForm()
    if form.validate_on_submit():
        existing = get_db().query_user_by_username(form.username.data)
        if existing:
            flash('이미 사용 중인 아이디입니다.', 'danger')
            return render_template('packing/register.html', form=form)

        temp_user = User()
        temp_user.set_password(form.password.data)

        get_db().insert_user({
            'username': form.username.data,
            'name': form.name.data,
            'company_name': form.company_name.data,
            'password_hash': temp_user.password_hash,
            'role': 'packing',
            'is_approved': False,
            'is_active_user': True,
        })

        created = get_db().query_user_by_username(form.username.data)
        created_id = created['id'] if created else None
        _packing_log_action('packing_register', target=form.username.data,
                            user_id=created_id,
                            detail=f'업체: {form.company_name.data}')

        flash('가입 신청이 완료되었습니다. 관리자 승인 후 이용 가능합니다.', 'success')
        return redirect(url_for('packing.packing_login'))

    return render_template('packing/register.html', form=form)


@packing_bp.route('/logout')
@login_required
def packing_logout():
    """패킹센터 로그아웃"""
    is_internal = current_user.role in _INTERNAL_ROLES
    _packing_log_action('packing_logout', target=current_user.username)
    if is_internal:
        # 내부직원은 로그아웃하지 않고 통합시스템으로 복귀
        flash('통합시스템으로 돌아갑니다.', 'info')
        return redirect(url_for('main.dashboard'))
    logout_user()
    session.clear()
    flash('로그아웃 되었습니다.', 'info')
    return redirect(url_for('packing.packing_login'))


@packing_bp.route('/')
@packing_required
def index():
    """패킹센터 홈"""
    return render_template('packing/index.html')


@packing_bp.route('/change-password', methods=['GET', 'POST'])
@packing_required
def change_password():
    """패킹센터 비밀번호 변경"""
    form = PackingChangePasswordForm()
    if form.validate_on_submit():
        if not current_user.check_password(form.current_password.data):
            flash('현재 비밀번호가 올바르지 않습니다.', 'danger')
            return render_template('packing/change_password.html', form=form)

        current_user.set_password(form.new_password.data)
        get_db().update_user(current_user.id, {
            'password_hash': current_user.password_hash,
            'password_changed_at': current_user.password_changed_at,
        })
        _packing_log_action('packing_change_password', target=current_user.username)
        flash('비밀번호가 변경되었습니다.', 'success')
        return redirect(url_for('packing.index'))

    return render_template('packing/change_password.html', form=form)


# ── Phase 2: 녹화모드 + 작업이력 ──────────────────────────

@packing_bp.route('/recording')
@packing_required
def recording():
    """녹화모드 페이지"""
    return render_template('packing/recording.html')


@packing_bp.route('/api/lookup-barcode', methods=['POST'])
@packing_required
def api_lookup_barcode():
    """송장번호(바코드)로 주문 검색."""
    data = request.get_json(silent=True) or {}
    barcode = data.get('barcode', '').strip()
    if not barcode:
        return jsonify({'ok': False, 'error': '바코드를 입력해주세요.'})

    db = get_db()

    # invoice_no_clean (하이픈 제거) 컬럼으로 1회 exact match 검색
    barcode_clean = barcode.replace('-', '')
    try:
        res = db.client.table("order_shipping").select("*") \
            .eq("invoice_no_clean", barcode_clean) \
            .eq("is_anonymized", False) \
            .limit(100).execute()
        shipping_list = res.data or []
    except Exception:
        shipping_list = []

    # fallback: invoice_no_clean이 없는 구 데이터 대비
    if not shipping_list:
        shipping_list = db.search_order_shipping(barcode, field='invoice')
    if not shipping_list:
        return jsonify({'ok': False,
                        'error': f'송장번호 "{barcode}"에 해당하는 주문이 없습니다.'})

    ship = shipping_list[0]
    channel = ship.get('channel', '')

    # 동일 송장의 모든 order_no 수집 (스마트스토어 등 상품별 order_no 분리)
    all_order_nos = list({s.get('order_no', '') for s in shipping_list if s.get('order_no')})
    order_no = all_order_nos[0] if all_order_nos else ''

    # order_transactions에서 전체 주문 상세 조회 — in_() 한 번으로 조회
    order_rows = []
    try:
        res = db.client.table("order_transactions") \
            .select("product_name,qty,original_option,barcode") \
            .eq("channel", channel).in_("order_no", all_order_nos) \
            .eq("status", "정상").execute()
        order_rows = res.data or []
    except Exception as e:
        import logging
        logging.error(f'[PACKING] order_transactions 조회 에러: {e}')
        order_rows = []

    # 수취인 마스킹
    name = ship.get('name', '')
    if name and len(name) > 1:
        masked = name[0] + '*' * (len(name) - 1)
    else:
        masked = '***'

    # option_master에서 바코드 보강 (order_transactions.barcode가 비어있는 경우)
    product_names_need_barcode = [
        o.get('product_name', '') for o in order_rows
        if not (o.get('barcode') or '').strip()
    ]
    barcode_map = {}
    if product_names_need_barcode:
        try:
            opt_list = db.query_option_master_as_list()  # 캐시 사용
            for opt in opt_list:
                pn = opt.get('품목명', '')
                bc = (opt.get('바코드') or '').strip()
                if pn and bc and pn in product_names_need_barcode:
                    barcode_map[pn] = bc
        except Exception:
            pass

    # 대표 품목명 (바코드 포함)
    items = []
    for o in order_rows:
        bc = (o.get('barcode') or '').strip()
        if not bc:
            bc = barcode_map.get(o.get('product_name', ''), '')
        items.append({
            'product_name': o.get('product_name', ''),
            'qty': o.get('qty', 0),
            'option_name': o.get('original_option', ''),
            'barcode': bc,
        })

    product_summary = ', '.join(
        f"{it['product_name']} x{it['qty']}" for it in items[:5]
    ) if items else '(품목 없음)'
    if len(items) > 5:
        product_summary += f' 외 {len(items) - 5}건'

    return jsonify({
        'ok': True,
        'data': {
            'channel': channel,
            'order_no': order_no,
            'order_nos': all_order_nos,  # 전체 order_no 목록
            'recipient_name': masked,
            'courier': ship.get('courier', ''),
            'items': items,
            'product_summary': product_summary,
            'total_qty': sum(it['qty'] for it in items),
        }
    })


@packing_bp.route('/api/start-job', methods=['POST'])
@packing_required
def api_start_job():
    """녹화 시작 — packing_jobs 레코드 생성."""
    data = request.get_json(silent=True) or {}
    barcode = data.get('barcode', '').strip()
    if not barcode:
        return jsonify({'ok': False, 'error': '바코드 없음'})

    db = get_db()
    job = {
        'user_id': current_user.id,
        'username': current_user.username,
        'company_name': getattr(current_user, 'company_name', '') or '',
        'scanned_barcode': barcode,
        'channel': data.get('channel', ''),
        'order_no': data.get('order_no', ''),
        'product_name': data.get('product_summary', ''),
        'recipient_name': data.get('recipient_name', ''),
        'order_info': json.dumps({
            'items': data.get('items', []),
            'order_nos': data.get('order_nos', []),
        }, ensure_ascii=False),
        'status': 'recording',
        'started_at': datetime.now(timezone.utc).isoformat(),
    }

    result = db.insert_packing_job(job)
    if not result:
        return jsonify({'ok': False, 'error': '작업 생성 실패'})

    _packing_log_action('packing_start_recording',
                        target=barcode,
                        detail=f'job_id={result["id"]}')
    return jsonify({'ok': True, 'job_id': result['id']})


@packing_bp.route('/api/complete-job', methods=['POST'])
@packing_required
def api_complete_job():
    """녹화 완료 — 영상 업로드 + 상태 갱신."""
    # CSRF 검증 (multipart이라 FlaskForm 미사용)
    csrf_token = request.form.get('csrf_token') or request.headers.get('X-CSRFToken')
    try:
        validate_csrf(csrf_token)
    except Exception:
        return jsonify({'ok': False, 'error': 'CSRF 검증 실패'}), 400

    job_id = request.form.get('job_id')
    video_file = request.files.get('video')
    duration_ms = request.form.get('duration_ms', 0, type=int)
    scanned_items_raw = request.form.get('scanned_items', '')

    if not job_id or not video_file:
        return jsonify({'ok': False, 'error': '필수 데이터 누락'})

    db = get_db()
    job = db.get_packing_job(int(job_id))
    if not job:
        return jsonify({'ok': False, 'error': '작업을 찾을 수 없습니다.'})

    # 권한 확인: 본인 작업 또는 운영자
    if current_user.role not in _INTERNAL_ROLES and job['user_id'] != current_user.id:
        return jsonify({'ok': False, 'error': '권한 없음'})

    # 영상 읽기
    video_bytes = video_file.read()
    max_size = current_app.config.get('PACKING_VIDEO_MAX_BYTES', 100 * 1024 * 1024)
    if len(video_bytes) > max_size:
        return jsonify({'ok': False, 'error': f'영상 크기 초과 ({len(video_bytes) // 1024 // 1024}MB)'})

    # Supabase Storage 업로드
    now = datetime.now(timezone.utc)
    path = (f"{now.strftime('%Y/%m/%d')}/"
            f"{current_user.id}_{job['scanned_barcode']}_{int(now.timestamp())}.webm")

    try:
        db.upload_packing_video(path, video_bytes)
    except Exception as e:
        return jsonify({'ok': False, 'error': f'영상 업로드 실패: {e}'})

    # scanned_items 파싱
    scanned_items = None
    if scanned_items_raw:
        try:
            import json as _json
            scanned_items = _json.loads(scanned_items_raw)
        except Exception:
            pass

    # order_info에 scanned_items 병합
    update_data = {
        'status': 'completed',
        'video_path': path,
        'video_size_bytes': len(video_bytes),
        'video_duration_ms': duration_ms,
        'completed_at': now.isoformat(),
    }
    if scanned_items is not None:
        existing_info = job.get('order_info') or {}
        if isinstance(existing_info, str):
            try:
                import json as _json
                existing_info = _json.loads(existing_info)
            except Exception:
                existing_info = {}
        if isinstance(existing_info, list):
            existing_info = {'items': existing_info}
        existing_info['scanned_items'] = scanned_items
        update_data['order_info'] = json.dumps(existing_info, ensure_ascii=False)

    # Job 업데이트
    db.update_packing_job(int(job_id), update_data)

    _packing_log_action('packing_complete_recording',
                        target=job['scanned_barcode'],
                        detail=f'job_id={job_id}, size={len(video_bytes)}')

    # ── 출고 처리 ──
    outbound_result = _process_outbound(db, job)
    return jsonify({'ok': True, 'outbound': outbound_result})


def _process_outbound(db, job):
    """패킹 완료 후 출고 처리 — 재고차감 + is_outbound_done."""
    channel = job.get('channel', '')
    order_no = job.get('order_no', '')
    order_nos = [order_no] if order_no else []
    try:
        info = json.loads(job.get('order_info', '{}')) if isinstance(job.get('order_info'), str) else (job.get('order_info') or {})
        if isinstance(info, dict) and info.get('order_nos'):
            order_nos = info['order_nos']
    except Exception:
        pass

    if not channel or not order_nos:
        return None

    try:
        from services.order_to_stock_service import process_packing_outbound
        ot_rows = db.client.table("order_transactions").select("id") \
            .eq("channel", channel).in_("order_no", order_nos) \
            .eq("status", "정상").eq("is_outbound_done", False) \
            .execute()
        ot_ids = [r['id'] for r in (ot_rows.data or [])]

        if ot_ids:
            outbound_results = []
            for oid in ot_ids:
                try:
                    res = process_packing_outbound(db, oid)
                    outbound_results.append(res)
                except Exception as e2:
                    outbound_results.append({'outbound_count': 0, 'error': str(e2)})
            total_out = sum(r.get('outbound_count', 0) for r in outbound_results)

            # order_shipping 배송상태 갱신 (전체 order_no)
            for ono in order_nos:
                try:
                    db.client.table("order_shipping").update({"shipping_status": "출고완료"}) \
                        .eq("channel", channel).eq("order_no", ono).execute()
                except Exception:
                    pass

            return {'outbound_count': total_out, 'orders': len(ot_ids)}
    except Exception as e:
        return {'error': str(e)}

    return None


@packing_bp.route('/api/complete-job-no-video', methods=['POST'])
@packing_required
def api_complete_job_no_video():
    """영상 없이 작업 완료 (카메라 없는 검증 모드)."""
    data = request.get_json(silent=True) or {}
    job_id = data.get('job_id')
    if not job_id:
        return jsonify({'ok': False, 'error': 'job_id 누락'})

    db = get_db()
    job = db.get_packing_job(int(job_id))
    if not job:
        return jsonify({'ok': False, 'error': '작업을 찾을 수 없습니다.'})

    if current_user.role not in _INTERNAL_ROLES and job['user_id'] != current_user.id:
        return jsonify({'ok': False, 'error': '권한 없음'})

    duration_ms = data.get('duration_ms', 0)
    scanned_items = data.get('scanned_items')

    now = datetime.now(timezone.utc)
    update_data = {
        'status': 'completed',
        'video_path': None,
        'video_size_bytes': 0,
        'video_duration_ms': duration_ms,
        'completed_at': now.isoformat(),
    }
    if scanned_items is not None:
        existing_info = job.get('order_info') or {}
        if isinstance(existing_info, str):
            try:
                existing_info = json.loads(existing_info)
            except Exception:
                existing_info = {}
        if isinstance(existing_info, list):
            existing_info = {'items': existing_info}
        existing_info['scanned_items'] = scanned_items
        update_data['order_info'] = json.dumps(existing_info, ensure_ascii=False)

    db.update_packing_job(int(job_id), update_data)

    _packing_log_action('packing_complete_no_video',
                        target=job['scanned_barcode'],
                        detail=f'job_id={job_id}')

    outbound_result = _process_outbound(db, job)
    return jsonify({'ok': True, 'outbound': outbound_result})


@packing_bp.route('/api/cancel-job', methods=['POST'])
@packing_required
def api_cancel_job():
    """녹화 취소."""
    data = request.get_json(silent=True) or {}
    job_id = data.get('job_id')
    if not job_id:
        return jsonify({'ok': False, 'error': 'job_id 누락'})

    db = get_db()
    db.update_packing_job(int(job_id), {'status': 'cancelled'})
    return jsonify({'ok': True})


@packing_bp.route('/history')
@packing_required
def history():
    """작업이력 페이지"""
    return render_template('packing/history.html')


@packing_bp.route('/api/history')
@packing_required
def api_history():
    """작업이력 JSON API."""
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    search = request.args.get('search', '')
    page = request.args.get('page', 1, type=int)
    per_page = 20

    db = get_db()
    # packing=본인, admin/manager=전체
    user_id = None if current_user.role in _INTERNAL_ROLES else current_user.id

    rows = db.query_packing_jobs(
        user_id=user_id, date_from=date_from, date_to=date_to,
        search=search, limit=per_page, offset=(page - 1) * per_page,
    )
    total = db.count_packing_jobs(
        user_id=user_id, date_from=date_from, date_to=date_to, search=search,
    )

    return jsonify({
        'ok': True,
        'data': rows,
        'total': total,
        'page': page,
        'pages': max(1, (total + per_page - 1) // per_page),
    })


@packing_bp.route('/api/video-url/<int:job_id>')
@packing_required
def api_video_url(job_id):
    """영상 서명 URL 반환."""
    db = get_db()
    job = db.get_packing_job(job_id)
    if not job:
        return jsonify({'ok': False, 'error': '작업 없음'})

    if current_user.role not in _INTERNAL_ROLES and job['user_id'] != current_user.id:
        return jsonify({'ok': False, 'error': '권한 없음'})

    if not job.get('video_path'):
        return jsonify({'ok': False, 'error': '영상 없음'})

    url = db.get_packing_video_signed_url(job['video_path'], expires_in=3600)
    if not url:
        return jsonify({'ok': False, 'error': '서명 URL 생성 실패'})

    return jsonify({'ok': True, 'url': url})


# ── Phase 3: 택배 관리 ──────────────────────────

@packing_bp.route('/courier')
@packing_required
def courier():
    """택배 관리 페이지 (송장 등록/관리)."""
    return render_template('packing/courier.html')


@packing_bp.route('/api/courier/pending')
@packing_required
def api_courier_pending():
    """송장 미등록 대기 주문 목록 조회."""
    channel = request.args.get('channel', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    db = get_db()
    try:
        # order_shipping에서 shipping_status='대기', invoice_no 없는 건 조회
        q = db.client.table("order_shipping").select(
            "id,channel,order_no,name,phone,address,memo,courier,invoice_no,"
            "shipping_status,created_at"
        ).eq("shipping_status", "대기")

        if channel:
            q = q.eq("channel", channel)
        if date_from:
            q = q.gte("created_at", f"{date_from}T00:00:00")
        if date_to:
            q = q.lte("created_at", f"{date_to}T23:59:59")

        q = q.order("created_at", desc=True).limit(500)
        res = q.execute()
        rows = res.data or []

        # 수취인 마스킹
        for r in rows:
            name = r.get('name', '')
            if name and len(name) > 1:
                r['name'] = name[0] + '*' * (len(name) - 1)
            else:
                r['name'] = '***'
            # 주소 마스킹 (시/구까지만)
            addr = r.get('address', '')
            if addr:
                parts = addr.split()
                r['address'] = ' '.join(parts[:3]) + ' ...' if len(parts) > 3 else addr

        return jsonify({'ok': True, 'data': rows, 'total': len(rows)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


@packing_bp.route('/api/courier/upload-excel', methods=['POST'])
@packing_required
def api_courier_upload_excel():
    """엑셀 파일로 송장번호 일괄 업로드."""
    # CSRF
    csrf_token = request.form.get('csrf_token') or request.headers.get('X-CSRFToken')
    try:
        validate_csrf(csrf_token)
    except Exception:
        return jsonify({'ok': False, 'error': 'CSRF 검증 실패'}), 400

    file = request.files.get('file')
    if not file:
        return jsonify({'ok': False, 'error': '파일이 없습니다.'})

    filename = file.filename.lower()
    if not filename.endswith(('.xlsx', '.xls')):
        return jsonify({'ok': False, 'error': 'xlsx 또는 xls 파일만 지원합니다.'})

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), read_only=True)
        ws = wb.active

        # 헤더 행 감지
        headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]

        # 컬럼 매핑 (유연한 컬럼명 감지)
        col_map = {}
        for i, h in enumerate(headers):
            hl = h.lower().replace(' ', '')
            if '주문번호' in h or 'order_no' in hl:
                col_map['order_no'] = i
            elif '송장번호' in h or 'invoice' in hl or '운송장' in h:
                col_map['invoice_no'] = i
            elif '택배사' in h or 'courier' in hl or '배송업체' in h:
                col_map['courier'] = i
            elif '채널' in h or 'channel' in hl:
                col_map['channel'] = i

        if 'order_no' not in col_map or 'invoice_no' not in col_map:
            return jsonify({'ok': False,
                            'error': '필수 컬럼이 없습니다. "주문번호"와 "송장번호" 컬럼이 필요합니다.'})

        # 데이터 파싱
        updates = []
        skip_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_list = list(row)
            order_no = str(row_list[col_map['order_no']] or '').strip()
            invoice_no = str(row_list[col_map['invoice_no']] or '').strip()

            if not order_no or not invoice_no:
                skip_count += 1
                continue

            channel = ''
            if 'channel' in col_map:
                channel = str(row_list[col_map['channel']] or '').strip()

            courier = 'CJ대한통운'  # 기본값
            if 'courier' in col_map:
                c = str(row_list[col_map['courier']] or '').strip()
                if c:
                    courier = c

            updates.append({
                'channel': channel,
                'order_no': order_no,
                'invoice_no': invoice_no,
                'courier': courier,
            })

        wb.close()

        if not updates:
            return jsonify({'ok': False, 'error': '유효한 데이터가 없습니다.'})

        # 채널 미지정 시 order_shipping에서 자동 검색
        db = get_db()
        for u in updates:
            if not u['channel']:
                ship = db.search_order_shipping(u['order_no'], field='order_no')
                if ship:
                    u['channel'] = ship[0].get('channel', '')

        # 일괄 업데이트
        success_count = db.bulk_update_shipping_invoices(updates)

        _packing_log_action('courier_excel_upload',
                            target=file.filename,
                            detail=f'total={len(updates)}, success={success_count}, skip={skip_count}')

        return jsonify({
            'ok': True,
            'total': len(updates),
            'success': success_count,
            'skipped': skip_count,
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': f'파일 처리 오류: {e}'})


# ── Phase 4: CJ대한통운 API 연동 ──────────────────────────

def _get_cj_client():
    """CJ 클라이언트 인스턴스 (lazy)."""
    from services.courier.cj_client import CJCourierClient
    cfg = current_app.config
    return CJCourierClient(
        api_key=cfg.get('CJ_API_KEY', ''),
        customer_id=cfg.get('CJ_CUSTOMER_ID', ''),
        base_url=cfg.get('CJ_API_BASE_URL', ''),
        test_mode=cfg.get('CJ_API_TEST_MODE', True),
    )


@packing_bp.route('/api/courier/register', methods=['POST'])
@packing_required
def api_courier_register():
    """선택한 주문을 CJ API로 송장 등록."""
    data = request.get_json(silent=True) or {}
    order_ids = data.get('order_ids', [])  # order_shipping id 목록

    if not order_ids:
        return jsonify({'ok': False, 'error': '등록할 주문을 선택해주세요.'})

    db = get_db()
    cj = _get_cj_client()

    # 발송인 정보 (사업장 기본)
    sender = {
        'name': '배마마',
        'phone': '02-0000-0000',
        'zipcode': '00000',
        'address': '서울특별시',
    }

    # 선택된 주문의 배송정보 조회
    results = []
    success_count = 0

    for oid in order_ids:
        try:
            ship_res = db.client.table("order_shipping").select("*") \
                .eq("id", oid).single().execute()
            ship = ship_res.data
            if not ship:
                results.append({'id': oid, 'ok': False, 'error': '주문 없음'})
                continue

            # 이미 송장 있으면 스킵
            if ship.get('invoice_no'):
                results.append({
                    'id': oid, 'ok': True,
                    'invoice_no': ship['invoice_no'],
                    'message': '이미 등록됨',
                })
                continue

            receiver = {
                'name': ship.get('name', ''),
                'phone': ship.get('phone', ''),
                'zipcode': '',
                'address': ship.get('address', ''),
            }

            # CJ API 호출
            reg_result = cj.register_shipment(
                sender=sender,
                receiver=receiver,
                items=[{'product_name': '배마마 이유식', 'qty': 1}],
                memo=ship.get('memo', ''),
            )

            if reg_result.get('ok'):
                # DB 업데이트
                db.update_order_shipping_invoice(
                    ship['channel'], ship['order_no'],
                    reg_result['invoice_no'], 'CJ대한통운'
                )
                success_count += 1
                results.append({
                    'id': oid, 'ok': True,
                    'invoice_no': reg_result['invoice_no'],
                    'order_no': ship['order_no'],
                })
            else:
                results.append({
                    'id': oid, 'ok': False,
                    'error': reg_result.get('error', '등록 실패'),
                })
        except Exception as e:
            results.append({'id': oid, 'ok': False, 'error': str(e)})

    _packing_log_action('courier_api_register',
                        detail=f'total={len(order_ids)}, success={success_count}')

    return jsonify({
        'ok': True,
        'total': len(order_ids),
        'success': success_count,
        'results': results,
    })


@packing_bp.route('/api/courier/label', methods=['POST'])
@packing_required
def api_courier_label():
    """운송장 라벨 PDF 다운로드."""
    from flask import Response

    data = request.get_json(silent=True) or {}
    invoice_nos = data.get('invoice_nos', [])

    if not invoice_nos:
        return jsonify({'ok': False, 'error': '송장번호가 없습니다.'})

    cj = _get_cj_client()
    result = cj.get_label_pdf(invoice_nos)

    if not result.get('ok'):
        return jsonify({'ok': False, 'error': result.get('error', '라벨 생성 실패')})

    return Response(
        result['pdf_bytes'],
        mimetype='application/pdf',
        headers={
            'Content-Disposition': f'attachment; filename=labels_{len(invoice_nos)}.pdf',
        }
    )


@packing_bp.route('/api/courier/tracking/<invoice_no>')
@packing_required
def api_courier_tracking(invoice_no):
    """배송 추적 조회."""
    cj = _get_cj_client()
    result = cj.get_tracking(invoice_no)
    return jsonify(result)
