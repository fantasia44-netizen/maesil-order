"""
adjustment.py — 재고 조정 Blueprint.
양수/음수 수량으로 재고 증감 조정, 사유(memo) 필수.
엑셀 실사 일괄조정 (기준일 역산) + 배치 되돌리기 포함.
"""
import uuid
import time
from datetime import datetime
from services.tz_utils import today_kst

from flask import (
    Blueprint, render_template, request, current_app,
    jsonify, send_file,
)
from flask_login import login_required, current_user

from auth import role_required, _log_action
from db_utils import get_db

adjustment_bp = Blueprint('adjustment', __name__, url_prefix='/adjustment')


@adjustment_bp.route('/')
@role_required('admin', 'manager', 'production', 'logistics', 'general')
def index():
    """재고 조정 페이지"""
    db = get_db()
    locations = []
    try:
        locations, _ = db.query_filter_options()
    except Exception:
        pass
    return render_template('adjustment/index.html', locations=locations)


@adjustment_bp.route('/api/products')
@role_required('admin', 'manager', 'production', 'logistics', 'general')
def api_products():
    """창고별 재고 품목 목록 JSON (자동완성용)"""
    location = request.args.get('location', '')
    if not location:
        return jsonify([])
    try:
        from services.excel_io import build_stock_snapshot
        all_data = get_db().query_stock_by_location(location)
        snapshot = build_stock_snapshot(all_data)
        products = []
        for name, info in snapshot.items():
            if info['total'] > 0:
                products.append({
                    'name': name,
                    'qty': info['total'],
                    'unit': info.get('unit', '개'),
                    'category': info.get('category', ''),
                    'storage_method': info.get('storage_method', ''),
                })
        products.sort(key=lambda x: x['name'])
        return jsonify(products)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@adjustment_bp.route('/api/history')
@role_required('admin', 'manager', 'production', 'logistics', 'general')
def api_history():
    """재고 조정 이력 조회 JSON"""
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    if not date_from or not date_to:
        return jsonify([])
    try:
        data = get_db().query_stock_ledger(
            date_from=date_from, date_to=date_to, type_list=['ADJUST'])
        rows = []
        for r in data:
            rows.append({
                'id': r.get('id'),
                'date': r.get('transaction_date', ''),
                'product_name': r.get('product_name', ''),
                'qty': r.get('qty', 0),
                'location': r.get('location', ''),
                'storage_method': r.get('storage_method', ''),
                'unit': r.get('unit', ''),
                'memo': r.get('memo', ''),
                'category': r.get('category', ''),
            })
        rows.sort(key=lambda x: (x['date'], x['product_name']))
        return jsonify(rows)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── API: 개별 삭제 (admin 전용) ──

@adjustment_bp.route('/api/delete/<int:record_id>', methods=['POST'])
@role_required('admin')
def api_delete(record_id):
    """개별 조정 이력 블라인드 처리 (admin 전용)"""
    try:
        old_record = get_db().query_stock_ledger_by_id(record_id)
        get_db().blind_stock_ledger(record_id, blinded_by=current_user.username)
        _log_action('blind_adjustment', target=str(record_id),
                     old_value=old_record)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── API: 개별 수정 (admin 전용) ──

@adjustment_bp.route('/api/update/<int:record_id>', methods=['POST'])
@role_required('admin')
def api_update(record_id):
    """개별 조정 이력 수정 (admin 전용)"""
    data = request.get_json(silent=True)
    if not data:
        return jsonify({'error': '수정 데이터가 없습니다.'}), 400
    allowed = {'product_name', 'qty', 'location', 'memo', 'storage_method', 'unit', 'category'}
    update_data = {k: v for k, v in data.items() if k in allowed}
    if 'qty' in update_data:
        try:
            update_data['qty'] = float(update_data['qty'])
            if update_data['qty'] == 0:
                raise ValueError
        except (ValueError, TypeError):
            return jsonify({'error': '수량은 0이 아니어야 합니다.'}), 400
    if 'memo' in update_data and not update_data['memo'].strip():
        return jsonify({'error': '사유를 입력하세요.'}), 400
    # 빈 문자열 → None 변환 (PostgreSQL TEXT 컬럼 호환)
    for key in ('storage_method', 'category'):
        if key in update_data and update_data[key] == '':
            update_data[key] = None
    if not update_data:
        return jsonify({'error': '수정할 항목이 없습니다.'}), 400

    original = get_db().query_stock_ledger_by_id(record_id)
    if not original:
        return jsonify({'error': '레코드를 찾을 수 없습니다.'}), 404

    skip_fields = {'id', 'status', 'replaced_by', 'replaces',
                   'created_at', 'updated_at', 'updated_by', 'created_by',
                   'is_deleted', 'deleted_at', 'deleted_by'}
    new_payload = {k: v for k, v in original.items() if k not in skip_fields}
    new_payload.update(update_data)

    try:
        new_id = get_db().replace_stock_ledger(
            record_id, new_payload, replaced_by_user=current_user.username)
        _log_action('replace_adjustment', target=str(record_id),
                     old_value={k: original.get(k) for k in update_data},
                     new_value=update_data)
        return jsonify({'success': True, 'new_id': new_id})
    except Exception as e:
        _log_action('replace_adjustment_error', target=str(record_id),
                     detail=f'조정 수정 오류: {str(e)}', new_value=update_data)
        return jsonify({'error': str(e)}), 500


# ═══ 엑셀 실사 일괄조정 ═══

@adjustment_bp.route('/survey/sample-excel')
@role_required('admin', 'manager', 'production', 'logistics')
def survey_sample_excel():
    """재고실사 샘플 엑셀 다운로드."""
    import io
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({'error': 'openpyxl 미설치'}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = '재고실사'
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)

    headers = ['품목명(필수)', '창고위치(필수)', '실사수량(필수)', '단위', '보관방법', '카테고리', '사유']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    ws.append(['건해삼채200g', '냉동실', 85, '개', '냉동', '해산물', ''])
    ws.append(['유기농사과즙', '냉장실', 120, '개', '냉장', '음료', ''])

    for i, w in enumerate([20, 14, 14, 8, 10, 12, 16], 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='재고실사_양식.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@adjustment_bp.route('/survey/export-stock')
@role_required('admin', 'manager', 'production', 'logistics')
def survey_export_stock():
    """기준일 시점 재고를 엑셀로 내보내기 (실사수량 컬럼 비워둠)."""
    import io
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({'error': 'openpyxl 미설치'}), 500

    location = request.args.get('location', '').strip()
    survey_date = request.args.get('survey_date', '').strip()

    if not location:
        return jsonify({'error': '창고위치를 선택해주세요.'}), 400

    db = get_db()
    from services.excel_io import build_stock_snapshot, normalize_location
    location = normalize_location(location)

    # 현재 재고 조회
    try:
        stock_data = db.query_stock_by_location(location)
        snapshot = build_stock_snapshot(stock_data)
    except Exception as e:
        return jsonify({'error': f'재고 조회 실패: {e}'}), 500

    # product_costs에서 category/storage_method 기본값 보강
    try:
        pc_map = db.query_product_costs()
        for pname, info in snapshot.items():
            pc = pc_map.get(pname) or pc_map.get(pname.replace(' ', '')) or {}
            if not info.get('storage_method') and pc.get('storage_method'):
                info['storage_method'] = pc['storage_method']
            if not info.get('category') and pc.get('category'):
                info['category'] = pc['category']
    except Exception:
        pass

    # 기준일 역산
    after_movements = {}
    if survey_date:
        try:
            survey_next = survey_date + 'T23:59:59'
            all_mvs = db.query_stock_ledger(
                date_from=survey_next, date_to='2099-12-31',
                location=location)
            for mv in all_mvs:
                name = mv.get('product_name', '')
                after_movements[name] = after_movements.get(name, 0) + float(mv.get('qty', 0))
        except Exception:
            pass

    wb = Workbook()
    ws = wb.active
    ws.title = '재고실사'

    # 스타일
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    input_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['품목명(필수)', '창고위치(필수)', '시스템재고', '실사수량(입력)', '단위', '보관방법', '카테고리', '사유']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # 데이터
    row_num = 2
    for product_name in sorted(snapshot.keys()):
        info = snapshot[product_name]
        current_qty = info.get('total', 0)
        if current_qty <= 0:
            continue

        after_mv = after_movements.get(product_name, 0)
        # 공백 제거 매칭도 시도
        if after_mv == 0:
            norm_name = product_name.replace(' ', '')
            for k, v in after_movements.items():
                if k.replace(' ', '') == norm_name:
                    after_mv = v
                    break

        system_qty = current_qty - after_mv if survey_date else current_qty

        ws.cell(row=row_num, column=1, value=product_name).border = thin_border
        ws.cell(row=row_num, column=2, value=location).border = thin_border
        c3 = ws.cell(row=row_num, column=3, value=system_qty)
        c3.border = thin_border
        c3.alignment = Alignment(horizontal='right')
        # 실사수량 — 노란색 입력 칸
        c4 = ws.cell(row=row_num, column=4, value=None)
        c4.fill = input_fill
        c4.border = thin_border
        c4.alignment = Alignment(horizontal='right')
        ws.cell(row=row_num, column=5, value=info.get('unit', '개')).border = thin_border
        ws.cell(row=row_num, column=6, value=info.get('storage_method', '')).border = thin_border
        ws.cell(row=row_num, column=7, value=info.get('category', '')).border = thin_border
        ws.cell(row=row_num, column=8, value='').border = thin_border
        row_num += 1

    # 컬럼 너비
    widths = [22, 12, 12, 14, 8, 10, 12, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # 메타 정보 시트
    ws2 = wb.create_sheet('_메타(수정금지)')
    ws2.append(['기준일', survey_date or '현재'])
    ws2.append(['창고위치', location])
    ws2.append(['내보내기일시', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    ws2.append(['품목수', row_num - 2])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"재고실사_{location}_{survey_date or 'current'}.xlsx"
    return send_file(buf, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@adjustment_bp.route('/survey/preview', methods=['POST'])
@role_required('admin', 'manager', 'production', 'logistics')
def survey_preview():
    """엑셀 업로드 → 기준일 시점 시스템재고 역산 → 미리보기 JSON."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return jsonify({'ok': False, 'error': 'openpyxl 미설치'}), 500

    f = request.files.get('file')
    if not f or not f.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'ok': False, 'error': '엑셀 파일(.xlsx)을 선택해주세요.'})

    survey_date = request.form.get('survey_date', '').strip()
    location_filter = request.form.get('location', '').strip()

    db = get_db()
    from services.excel_io import build_stock_snapshot, normalize_location

    wb = load_workbook(f, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    preview = []
    errors = []

    # 헤더 감지: 내보내기 엑셀(8컬럼: 품목명/위치/시스템재고/실사수량/...) vs 수동양식(7컬럼)
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0] if ws.max_row else []
    is_export_format = len(header_row) >= 4 and header_row[2] and '시스템' in str(header_row[2])

    # ── 성능 최적화: 재고/이력을 루프 밖에서 1번만 조회 (캐시) ──
    _snapshot_cache = {}  # {location: snapshot}
    _after_mv_cache = {}  # {location: {product_name: 변동합계}}

    def _get_snapshot(loc):
        if loc not in _snapshot_cache:
            try:
                _snapshot_cache[loc] = build_stock_snapshot(db.query_stock_by_location(loc))
            except Exception:
                _snapshot_cache[loc] = {}
        return _snapshot_cache[loc]

    def _get_after_movements(loc):
        if loc not in _after_mv_cache:
            _after_mv_cache[loc] = {}
            if survey_date:
                try:
                    survey_next = survey_date + 'T23:59:59'
                    all_mvs = db.query_stock_ledger(
                        date_from=survey_next, date_to='2099-12-31', location=loc)
                    for mv in all_mvs:
                        name = mv.get('product_name', '')
                        _after_mv_cache[loc][name] = _after_mv_cache[loc].get(name, 0) + float(mv.get('qty', 0))
                        # 공백 제거 버전도 등록
                        norm = name.replace(' ', '')
                        if norm != name:
                            _after_mv_cache[loc][norm] = _after_mv_cache[loc].get(norm, 0) + float(mv.get('qty', 0))
                except Exception:
                    pass
        return _after_mv_cache[loc]

    for i, row in enumerate(rows, 2):
        if not row or not row[0]:
            continue
        product_name = str(row[0] or '').strip()
        location = str(row[1] or '').strip() if len(row) > 1 else location_filter
        if not location:
            errors.append(f'행 {i}: 창고위치가 비어있습니다.')
            continue
        location = normalize_location(location)

        if is_export_format:
            # 엑셀에 기준일 시스템재고(row[2])가 포함됨 → 그대로 사용
            try:
                excel_system_qty = float(row[2]) if len(row) > 2 and row[2] is not None else None
            except (ValueError, TypeError):
                excel_system_qty = None
            try:
                actual_qty = float(row[3]) if len(row) > 3 and row[3] is not None else None
            except (ValueError, TypeError):
                errors.append(f'행 {i}: 실사수량이 숫자가 아닙니다.')
                continue
            if actual_qty is None:
                continue
            unit = str(row[4] or '').strip() if len(row) > 4 else ''
            storage_method = str(row[5] or '').strip() if len(row) > 5 else ''
            category = str(row[6] or '').strip() if len(row) > 6 else ''
            memo = str(row[7] or '').strip() if len(row) > 7 else ''
        else:
            try:
                actual_qty = float(row[2]) if len(row) > 2 and row[2] is not None else None
            except (ValueError, TypeError):
                errors.append(f'행 {i}: 실사수량이 숫자가 아닙니다.')
                continue
            if actual_qty is None:
                errors.append(f'행 {i}: 실사수량이 비어있습니다.')
                continue
            unit = str(row[3] or '').strip() if len(row) > 3 else ''
            storage_method = str(row[4] or '').strip() if len(row) > 4 else ''
            category = str(row[5] or '').strip() if len(row) > 5 else ''
            memo = str(row[6] or '').strip() if len(row) > 6 else ''

        # 캐시된 스냅샷에서 조회 (DB 호출 없음)
        snapshot = _get_snapshot(location)

        # 품목명 매칭
        current_qty = 0
        normalized_name = product_name.replace(' ', '')
        for sname, sinfo in snapshot.items():
            if sname == product_name or sname.replace(' ', '') == normalized_name:
                current_qty = sinfo.get('total', 0)
                if not unit:
                    unit = sinfo.get('unit', '개')
                if not category:
                    category = sinfo.get('category', '')
                if not storage_method:
                    storage_method = sinfo.get('storage_method', '')
                break

        # 기준일 시스템재고 결정 (서버 재계산 우선, 엑셀 값은 교차검증용):
        after_mv_map = _get_after_movements(location)
        after_movements_server = after_mv_map.get(product_name, 0) or after_mv_map.get(normalized_name, 0)
        server_system_qty = current_qty - after_movements_server if survey_date else current_qty

        # 엑셀에 박혀있는 시스템재고와 서버 재계산값이 다르면 경고
        warn = None
        if is_export_format and excel_system_qty is not None:
            if abs(excel_system_qty - server_system_qty) > max(1, server_system_qty * 0.01):
                warn = (f'엑셀 시스템재고({excel_system_qty:g})와 '
                        f'현재 재계산값({server_system_qty:g}) 불일치 — 서버값 사용')

        # 항상 서버 재계산값 사용 (엑셀 값 신뢰 안함)
        system_qty_at_date = server_system_qty
        after_movements = after_movements_server

        delta = actual_qty - system_qty_at_date

        # 큰 변동 경고 (현재 재고의 50% 이상)
        large_delta = abs(delta) > max(10, current_qty * 0.5) if current_qty else abs(delta) > 10

        # 조정 후 음수 예상 경고
        would_be_negative = (current_qty + delta) < 0

        preview.append({
            'row': i,
            'product_name': product_name,
            'location': location,
            'unit': unit or '개',
            'storage_method': storage_method,
            'category': category,
            'memo': memo,
            'current_qty': current_qty,
            'system_qty_at_date': system_qty_at_date,
            'excel_system_qty': excel_system_qty if is_export_format else None,
            'after_movements': after_movements,
            'actual_qty': actual_qty,
            'delta': delta,
            'large_delta': large_delta,
            'would_be_negative': would_be_negative,
            'warning': warn,
            'survey_date': survey_date or None,
        })

    return jsonify({
        'ok': True,
        'preview': preview,
        'errors': errors,
        'total_items': len(preview),
        'increase_count': sum(1 for p in preview if p['delta'] > 0),
        'decrease_count': sum(1 for p in preview if p['delta'] < 0),
        'no_change_count': sum(1 for p in preview if p['delta'] == 0),
    })


@adjustment_bp.route('/survey/apply', methods=['POST'])
@role_required('admin', 'manager', 'production', 'logistics')
def survey_apply():
    """미리보기 확인 후 일괄 적용."""
    data = request.get_json(silent=True)
    if not data or 'items' not in data:
        return jsonify({'ok': False, 'error': '적용할 데이터가 없습니다.'})

    items = data['items']
    memo_prefix = data.get('memo', '재고실사 일괄조정')
    survey_date = data.get('survey_date', today_kst())
    batch_id = f"SURVEY-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6].upper()}"

    db = get_db()
    from services.adjustment_service import process_adjustment_batch

    # 각 항목에 배치ID와 상세 사유 부여
    batch_items = []
    for item in items:
        delta = item.get('delta', 0)
        if delta == 0:
            continue
        batch_items.append({
            'product_name': item['product_name'],
            'location': item['location'],
            'qty': delta,
            'memo': f"[{batch_id}] {memo_prefix} (실사:{item.get('actual_qty')} 시스템:{item.get('system_qty_at_date')} 차이:{delta:+g})",
            'unit': item.get('unit', '개'),
            'storage_method': item.get('storage_method', ''),
            'category': item.get('category', ''),
        })

    if not batch_items:
        return jsonify({'ok': True, 'batch_id': batch_id, 'success': 0, 'fail': 0, 'skipped': len(items)})

    try:
        result = process_adjustment_batch(
            db, survey_date, batch_items,
            created_by=current_user.username)
        _log_action('survey_adjustment',
                     detail=f'재고실사 일괄조정 {batch_id}: {result.get("count", 0)}건',
                     new_value={'batch_id': batch_id, 'survey_date': survey_date,
                                'count': result.get('count', 0)})
        return jsonify({
            'ok': True,
            'batch_id': batch_id,
            'success': result.get('count', 0),
            'fail': 0,
            'skipped': len(items) - len(batch_items),
            'increase_count': result.get('increase_count', 0),
            'decrease_count': result.get('decrease_count', 0),
        })
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)}), 400
    except Exception as e:
        return jsonify({'ok': False, 'error': f'일괄 조정 오류: {e}'}), 500


@adjustment_bp.route('/survey/batch-history')
@role_required('admin', 'manager', 'production', 'logistics')
def survey_batch_history():
    """배치 실사 조정 이력 조회."""
    db = get_db()
    # 최근 60일 조정 이력에서 SURVEY- 배치만 추출
    from services.tz_utils import today_kst
    from datetime import timedelta
    end = today_kst()
    start = (datetime.strptime(end, '%Y-%m-%d') - timedelta(days=60)).strftime('%Y-%m-%d')

    movements = db.query_stock_ledger(date_from=start, date_to=end, type_list=['ADJUST'])
    batches = {}
    for m in movements:
        memo = m.get('memo', '')
        if not memo.startswith('[SURVEY-') and not memo.startswith('[ROLLBACK-SURVEY-'):
            continue
        bid = memo.split(']')[0][1:]  # [SURVEY-xxx] → SURVEY-xxx
        if bid not in batches:
            batches[bid] = {
                'batch_id': bid,
                'created_at': m.get('transaction_date', ''),
                'items': [],
                'total_increase': 0,
                'total_decrease': 0,
            }
        qty = float(m.get('qty', 0))
        batches[bid]['items'].append({
            'id': m.get('id'),
            'product_name': m.get('product_name', ''),
            'location': m.get('location', ''),
            'qty': qty,
            'memo': memo,
        })
        if qty > 0:
            batches[bid]['total_increase'] += qty
        else:
            batches[bid]['total_decrease'] += qty

    batch_list = sorted(batches.values(), key=lambda b: b['batch_id'], reverse=True)
    return jsonify({'ok': True, 'batches': batch_list[:20]})


@adjustment_bp.route('/survey/rollback', methods=['POST'])
@role_required('admin', 'manager')
def survey_rollback():
    """배치 되돌리기: 원본 ADJUST를 유지하고, 같은 날짜에 offset ADJUST 생성.

    원칙:
      - 원본 보존 (audit trail)
      - offset 레코드 transaction_date = 원본 transaction_date
      - 과거 기준일 화면에서도 net=0 이 되어 재고 일관성 유지
    """
    data = request.get_json(silent=True)
    batch_id = data.get('batch_id') if data else None
    if not batch_id:
        return jsonify({'ok': False, 'error': '배치 ID가 필요합니다.'})

    db = get_db()
    from services.tz_utils import today_kst
    from datetime import timedelta
    end = today_kst()
    start = (datetime.strptime(end, '%Y-%m-%d') - timedelta(days=180)).strftime('%Y-%m-%d')

    movements = db.query_stock_ledger(date_from=start, date_to=end, type_list=['ADJUST'])
    batch_items = [m for m in movements
                   if m.get('memo', '').startswith(f'[{batch_id}]')
                   and not m.get('is_deleted')]
    if not batch_items:
        return jsonify({'ok': False, 'error': f'배치 "{batch_id}" 이력을 찾을 수 없습니다.'})

    # 이미 offset 생성된 배치인지 확인
    rollback_check = [m for m in movements
                      if m.get('memo', '').startswith(f'[ROLLBACK-{batch_id}]')
                      and not m.get('is_deleted')]
    if rollback_check:
        return jsonify({'ok': False, 'error': '이미 되돌린 배치입니다.'})

    # 원본 날짜별 groupby → offset ADJUST 각각 같은 날짜로 생성
    from services.adjustment_service import process_adjustment_batch
    success = 0
    fail = 0
    errors = []
    dates_processed = set()

    # 원본 transaction_date 별로 묶어서 process_adjustment_batch 호출
    from collections import defaultdict
    by_date: dict[str, list] = defaultdict(list)
    for item in batch_items:
        qty = float(item.get('qty', 0))
        if qty == 0:
            continue
        d = str(item.get('transaction_date', ''))[:10]
        by_date[d].append({
            'product_name': item['product_name'],
            'location': item.get('location', ''),
            'qty': -qty,  # 역방향
            'memo': f"[ROLLBACK-{batch_id}] 되돌리기 ({-qty:+g})",
            'unit': item.get('unit', '개'),
            'storage_method': item.get('storage_method', ''),
            'category': item.get('category', ''),
        })

    for txn_date, items in by_date.items():
        try:
            result = process_adjustment_batch(
                db, txn_date, items,
                created_by=current_user.username)
            success += result.get('count', 0)
            dates_processed.add(txn_date)
        except Exception as e:
            fail += len(items)
            errors.append(f'{txn_date}: {str(e)[:100]}')

    _log_action('survey_rollback',
                 detail=f'재고실사 되돌리기 {batch_id}: {success}건 (원본 날짜 기준)',
                 new_value={'batch_id': batch_id, 'success': success,
                            'dates': sorted(dates_processed)})
    return jsonify({
        'ok': True,
        'batch_id': batch_id,
        'rollback_success': success,
        'rollback_fail': fail,
        'dates_processed': sorted(dates_processed),
        'errors': errors[:5] if errors else [],
    })


@adjustment_bp.route('/batch', methods=['POST'])
@role_required('admin', 'manager', 'production', 'logistics', 'general')
def batch():
    """다건 일괄 재고 조정 (JSON)"""
    data = request.get_json(silent=True)
    if not data:
        return jsonify({'error': '요청 데이터가 없습니다.'}), 400

    items = data.get('items', [])
    date_str = data.get('date', today_kst())

    if not items:
        return jsonify({'error': '조정 항목이 없습니다.'}), 400

    # 유효성 검증
    for i, item in enumerate(items):
        name = str(item.get('product_name', '')).strip()
        location = str(item.get('location', '')).strip()
        qty = item.get('qty', 0)
        memo = str(item.get('memo', '')).strip()
        if not name:
            return jsonify({'error': f'{i+1}번째 항목: 품목명을 입력하세요.'}), 400
        if not location:
            return jsonify({'error': f'{i+1}번째 항목: 창고위치를 선택하세요.'}), 400
        try:
            if float(qty) == 0:
                raise ValueError
        except (ValueError, TypeError):
            return jsonify({'error': f'{i+1}번째 항목 ({name}): 수량은 0이 아니어야 합니다.'}), 400
        if not memo:
            return jsonify({'error': f'{i+1}번째 항목 ({name}): 사유를 입력하세요.'}), 400

    try:
        from services.adjustment_service import process_adjustment_batch
        result = process_adjustment_batch(
            get_db(), date_str, items,
            created_by=current_user.username)
        _log_action('batch_adjustment',
                     detail=f'{date_str} 재고조정 {result.get("count", 0)}건 '
                            f'(증가 {result.get("increase_count", 0)}건, '
                            f'감소 {result.get("decrease_count", 0)}건, '
                            f'항목 {len(items)}건)',
                     new_value={'date': date_str, 'batch_ts': result.get('batch_ts'),
                                'count': result.get('count', 0)})
        return jsonify({
            'success': True,
            'count': result.get('count', 0),
            'increase_count': result.get('increase_count', 0),
            'decrease_count': result.get('decrease_count', 0),
            'warnings': result.get('warnings', []),
        })
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': f'재고 조정 중 오류: {e}'}), 500
